#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""입력 정규화 폴백 추출기 — 외부 MCP 없이 학생 원본에서 텍스트를 뽑는다.

지원: hwpx(zip+xml, 표준 라이브러리), xlsx(openpyxl), 확장자 위장 HTML, 일반 텍스트.
지원 밖 포맷은 SKILL.md의 폴백 절차를 따른다: 이미지·스캔 pdf는 Claude가 직접
읽어 전사하고, hwp 바이너리는 한글에서 hwpx나 pdf로 다시 저장하도록 안내한다.
"""
from __future__ import annotations

import sys
import zipfile
from html.parser import HTMLParser
from pathlib import Path
from xml.etree import ElementTree

GUIDANCE = {
    "hwp-or-ole": "hwp 바이너리(또는 구형 xls) 파일입니다. 한글이나 엑셀에서 hwpx, xlsx, pdf 형식으로 다시 저장해 주세요.",
    "image": "이미지 파일입니다. Claude가 파일을 직접 읽어 전사하세요(외부 OCR 불필요).",
    "pdf": "pdf 파일입니다. Claude가 직접 읽어 전사하거나(스캔 포함) 텍스트 pdf는 pypdf로 추출하세요.",
    "docx": "docx 파일입니다. 사용 가능한 문서 파싱 도구를 활용하거나 pdf로 변환해 주세요.",
    "zip": "알 수 없는 zip 계열 파일입니다. 원본 형식을 확인해 주세요.",
}


class UnsupportedFormatError(Exception):
    def __init__(self, fmt: str):
        self.fmt = fmt
        super().__init__(GUIDANCE.get(fmt, f"지원하지 않는 형식입니다: {fmt}"))


def sniff_format(path) -> str:
    """확장자가 아니라 파일 헤더로 실제 포맷을 판정한다."""
    head = Path(path).open("rb").read(512)
    if head[:4] == b"PK\x03\x04":
        with zipfile.ZipFile(path) as z:
            names = z.namelist()
        if any(n.startswith("Contents/") for n in names):
            return "hwpx"
        if any(n.startswith("xl/") for n in names):
            return "xlsx"
        if any(n.startswith("word/") for n in names):
            return "docx"
        return "zip"
    if head[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
        return "hwp-or-ole"
    if head[:5] == b"%PDF-":
        return "pdf"
    if head[:4] == b"\x89PNG" or head[:3] == b"\xff\xd8\xff":
        return "image"
    upper = head.upper()
    if head.lstrip()[:1] == b"<" or b"<!DOCTYPE" in upper or b"<HTML" in upper:
        return "html"
    return "text"


def _extract_hwpx(path) -> str:
    paragraphs = []
    with zipfile.ZipFile(path) as z:
        sections = sorted(n for n in z.namelist() if n.startswith("Contents/section"))
        for name in sections:
            root = ElementTree.fromstring(z.read(name))
            for p in root.iter():
                if not p.tag.endswith("}p"):
                    continue
                text = "".join(t.text or "" for t in p.iter() if t.tag.endswith("}t"))
                if text.strip():
                    paragraphs.append(text)
    return "\n".join(paragraphs)


class _HtmlText(HTMLParser):
    """테이블은 탭 구분 행으로, 나머지 텍스트는 그대로 수집한다."""

    def __init__(self):
        super().__init__()
        self.rows = []
        self.row = None
        self.cell = None
        self.loose = []

    def handle_starttag(self, tag, attrs):
        if tag == "tr":
            self.row = []
        elif tag in ("td", "th"):
            self.cell = []
        elif tag == "br" and self.cell is not None:
            self.cell.append("\n")

    def handle_endtag(self, tag):
        if tag in ("td", "th") and self.row is not None and self.cell is not None:
            self.row.append("".join(self.cell).strip())
            self.cell = None
        elif tag == "tr" and self.row:
            self.rows.append(self.row)
            self.row = None

    def handle_data(self, data):
        if self.cell is not None:
            self.cell.append(data)
        elif data.strip():
            self.loose.append(data.strip())

    def text(self) -> str:
        parts = ["\t".join(cells).replace("\n", " ") for cells in self.rows]
        if not parts:
            parts = self.loose
        return "\n".join(parts)


def _extract_html(path) -> str:
    parser = _HtmlText()
    parser.feed(Path(path).read_text(encoding="utf-8", errors="replace"))
    return parser.text()


def _extract_xlsx(path) -> str:
    from openpyxl import load_workbook

    lines = []
    wb = load_workbook(path, read_only=True, data_only=True)
    for ws in wb.worksheets:
        lines.append(f"## 시트: {ws.title}")
        for row in ws.iter_rows(values_only=True):
            cells = ["" if c is None else str(c) for c in row]
            if any(c.strip() for c in cells):
                lines.append("\t".join(cells))
    wb.close()
    return "\n".join(lines)


def extract(path):
    """(포맷, 텍스트)를 반환한다. 지원 밖 포맷은 UnsupportedFormatError."""
    fmt = sniff_format(path)
    if fmt == "hwpx":
        return fmt, _extract_hwpx(path)
    if fmt == "html":
        return fmt, _extract_html(path)
    if fmt == "xlsx":
        return fmt, _extract_xlsx(path)
    if fmt == "text":
        return fmt, Path(path).read_text(encoding="utf-8", errors="replace")
    raise UnsupportedFormatError(fmt)


def main(argv=None) -> int:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")

    import argparse

    parser = argparse.ArgumentParser(description="학생 원본 텍스트 추출기(폴백)")
    parser.add_argument("file", help="추출할 파일 경로")
    args = parser.parse_args(argv)

    try:
        fmt, text = extract(args.file)
    except UnsupportedFormatError as e:
        print(f"[{e.fmt}] {e}")
        return 2
    print(f"[형식: {fmt}]")
    print(text)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
