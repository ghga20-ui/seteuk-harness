#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""가명처리 파이프라인 — 명렬 인식, 토큰 발급·치환, 재결합, 유출 검사.

식별정보(실명·학번)는 이 로컬 모듈과 매핑표에만 존재한다. LLM에는 토큰만 노출된다.
매칭 키는 이름이 아니라 학번이다(동명이인 원천 차단).
"""
from __future__ import annotations

import json
import re
import secrets
from pathlib import Path

STUDENT_ID = re.compile(r"\b\d{5}\b")
# 패턴 경로(헤더 없음) 전용 — 추측이므로 오탐 방지가 우선이라 형식을 엄격히 본다.
# {2,4} -> {2,8}: 다문화 성명(예: 응우옌티탄흐엉, 7자)을 놓치지 않기 위해 넓혔다.
# 패턴 경로는 이미 확인필요=True가 붙으므로 사용자 확인으로 오탐을 보완한다.
KOREAN_NAME = re.compile(r"^[가-힣]{2,8}$")
MAX_HEADER_NAME_LEN = 20  # 헤더 경로(이름 열이 명시된 경우)의 과도한 길이 방지용 상한선


def _char_spaced_pattern(s: str) -> str:
    """문자(숫자) 사이에 임의의 공백을 허용하는 정규식 패턴 문자열을 만든다.

    예: '김하윤' -> '김\\s*하\\s*윤'. 학생이 '김 하윤'처럼 띄어 쓰거나 학번을
    '301 01'처럼 띄어 쓰는 표기 변형을 놓치지 않기 위한 공용 유틸이다(골든리포트
    §6 FIX A). 치환(pseudonymize_text)과 탐지(scan_leak)가 같은 패턴을 써야
    "치환은 하는데 검사는 못 잡는" 게이트 누수가 생기지 않는다.
    """
    return r"\s*".join(re.escape(ch) for ch in s)


def _name_pattern(name: str) -> re.Pattern:
    """이름을 공백 내성 정규식으로 컴파일한다(치환·탐지 공용)."""
    return re.compile(_char_spaced_pattern(name))


def _id_pattern(sid: str) -> re.Pattern:
    """학번을 공백 내성 정규식으로 컴파일한다.

    숫자 경계((?<!\\d)/(?!\\d))는 공백 허용과 무관하게 그대로 유지한다 —
    '101010' 같은 무관한 숫자열 안의 부분 문자열을 오탐하지 않기 위해서다.
    """
    return re.compile(rf"(?<!\d){_char_spaced_pattern(sid)}(?!\d)")


# 성 제외 이름(given name) 뒤에 붙는 인명 전용 호칭·조사. 실측(학생 답안
# 69,034자)에서 성 제외 이름이 맨몸으로 등장한 4건은 전부 시어·일반어와의
# 충돌(예: 성 제외 이름이 '하늘'인 경우 "하늘을 우러러")이었고, 조사가 붙어
# 인명임이 확실한 경우는 0건이었다. 그래서 성 제외 이름은 이 접미가 뒤따를
# 때만 치환·탐지 대상으로 삼는다 — 맨몸 등장은 의도적으로 건드리지 않는다.
GIVEN_NAME_HONORIFIC_SUFFIXES = sorted(
    ["이는", "이가", "이를", "이랑", "이한테", "이와", "이도", "이의", "이네",
     "아", "야", "쌤", "샘", "선배", "누나", "형", "언니", "오빠"],
    key=len, reverse=True,
)


def _given_name_suffix_pattern(given: str) -> re.Pattern:
    """성 제외 이름 + 인명 전용 접미(호칭 조사)를 캡처 그룹으로 나눠 컴파일한다.

    그룹1=이름, 그룹2=접미. 치환 시 접미(그룹2)는 그대로 보존해야 문장이
    깨지지 않는다('하윤이는' -> 'S-XXXX이는', '이는'까지 삼키면 안 됨).
    """
    suffix_alt = "|".join(re.escape(s) for s in GIVEN_NAME_HONORIFIC_SUFFIXES)
    return re.compile(f"({_char_spaced_pattern(given)})({suffix_alt})")


def _sub_given_name_with_suffix(text: str, given: str, replacement: str):
    """성 제외 이름을 접미 보존 조건으로 치환한다. 반환은 (결과, 치환건수)."""
    pattern = _given_name_suffix_pattern(given)
    return pattern.subn(lambda m: replacement + m.group(2), text)


def _cell_text(c) -> str:
    """셀 값을 문자열로 만든다. 정수인 실수는 소수점을 떼어 학번을 살린다.

    엑셀이 학번을 숫자 서식으로 저장하면(설문 응답 시트 등) 셀이 30101.0으로
    들어와 str()이 "30101.0"이 되고 isdigit()이 False가 되어 행 전체가 버려졌다.
    실측에서 건너뛴 947행 중 441행(46.6%)이 이 한 가지 원인이었다.
    3.5 같은 진짜 소수는 그대로 둔다 — 점수 열이 훼손되면 안 된다.
    """
    if c is None:
        return ""
    if isinstance(c, float) and c.is_integer():
        return str(int(c))
    return str(c).strip()


def _sheets_from_xlsx(path):
    """워크북의 시트를 (시트명, 행 목록) 쌍의 목록으로 반환한다. 시트를 이어붙이지 않는다.

    이전 구현(_rows_from_xlsx)은 전 시트를 한 리스트로 이어붙였다 — 그래서 반별로
    나뉜 채점표가 한 반으로 합쳐지고, 시트1의 헤더 열 인덱스로 시트2의 데이터를
    읽어(열 인덱스는 시트마다 다르다) 엉뚱한 값을 가져오는 사고가 났다(실측 341개
    중 19~29개 파일, 최악 사례는 라벨-데이터 불일치 87/117쌍=74.4%). 호출부가
    시트별로 처리한 뒤 합칠지·멈출지 스스로 판단해야 하므로 시트 경계를 보존해
    반환한다.
    """
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    sheets = []
    for ws in wb.worksheets:
        rows = [[_cell_text(c) for c in row] for row in ws.iter_rows(values_only=True)]
        sheets.append((ws.title, rows))
    wb.close()
    return sheets


def _rows_from_text(path):
    # utf-8-sig: 메모장·PowerShell 5.1 -Encoding utf8 기본값은 UTF-8 BOM을 붙인다.
    # BOM이 있으면 제거하고, 없으면 그대로 읽으므로 항상 안전하다.
    text = Path(path).read_text(encoding="utf-8-sig", errors="replace")
    return [line.split() for line in text.splitlines() if line.strip()]


def _find_header_indices(rows):
    """헤더 행을 찾아 (id_idx, name_idx, surname_idx, 헤더행번호)를 반환.
    없으면 (None, None, None, -1).

    surname_idx는 "성" 또는 "성씨" 열이 이름 열과 함께 있을 때만 채워진다.
    성·이름이 분리된 명렬(흔한 형식)에서 이름 열만 보면 1자 이름으로 오인되어
    본문 치환에서 제외되고, 그 이름이 그대로 LLM에 노출되는 문제가 있었다 —
    성 열을 찾아 결합하면 이 문제가 근본적으로 사라진다.
    """
    for i, row in enumerate(rows):
        id_idx = next((j for j, c in enumerate(row) if c in ("학번", "번호")), None)
        name_idx = next((j for j, c in enumerate(row) if c == "이름"), None)
        surname_idx = next((j for j, c in enumerate(row) if c in ("성", "성씨")), None)
        if id_idx is not None and name_idx is not None:
            return (id_idx, name_idx, surname_idx, i)
    return (None, None, None, -1)


STATUS_WORDS = {
    "자퇴", "전출", "전입", "미등록", "휴학", "제적", "유예", "면제", "결시", "미응시",
}


def _is_status_word(name: str) -> bool:
    """이름 칸 값이 학생 상태 표기어인지 확인한다(공백 제거 후 완전 일치).

    실제 채점표에서 학번은 살아 있는데 이름 칸에 '자퇴' 같은 상태 문구가 적힌
    행이 있었다 — 이런 행은 학생이 아니므로 명렬에 포함하면 안 된다.
    """
    return re.sub(r"\s+", "", name or "") in STATUS_WORDS


def _pairs_from_rows_with_indices(rows, id_idx, name_idx, header_row_idx, surname_idx=None):
    """헤더 행의 열 인덱스를 사용해 (학번, 이름) 쌍을 추출.

    헤더가 "이름"이라고 명시적으로 지목한 열이므로 이름 형식(한글 2~4자 등)을
    검사하지 않는다 — 형식 검사는 헤더 없이 추측하는 패턴 경로에만 필요하다.
    이 검사 때문에 다문화 성명(응우옌티탄흐엉), 공백 포함 이름(박 서준), 영문
    이름(Nguyen), 1자 이름(봄) 등 실재하는 학생이 명렬에서 조용히 누락되는
    버그가 있었다. 이름은 비어 있지 않고 과도하게 길지 않으면 그대로 채택한다.
    학번은 헤더가 지목한 열이므로 숫자인지만 느슨히 확인한다.

    surname_idx가 주어지면(성 열이 이름 열과 함께 있으면) 그 행의 성 셀이
    비어 있지 않은 한 성 + 이름을 합쳐 전체 이름으로 사용한다. 한글 성명은
    성 1자 이상 + 이름 1자 이상이 결합되어야 완전하므로, 분리된 명렬에서
    이름 열만 보면 1자 이름으로 오인되어 본문에서 안전하게 가릴 수 없는
    상태가 된다 — 결합이 그 문제를 원천적으로 없앤다.

    학번·이름 중 하나만 비어 있는 데이터 행은 건너뛰고 건너뛴 행 수를 함께
    반환한다(둘 다 비어 있는 완전한 공백 행은 집계하지 않는다). 이름 칸이
    상태 표기어(자퇴 등)인 행도 학생이 아니므로 건너뜀에 포함한다 — 반환값에
    상태 표기로 건너뛴 건수를 별도로 함께 돌려준다(안내 문구 분기용).

    학번이 앞 행과 겹치는 행은 채택하지 않고 중복 건수로 따로 센다. '번호'
    열(출석번호)을 학번으로 쓰는 표는 반이 바뀔 때마다 번호가 되풀이되므로
    이 경로로 대량의 행이 사라진다 — 실측에서 738행짜리 표가 32명으로,
    805행이 38명으로 보고됐다. 세지 않으면 교사가 알아챌 방법이 없다.

    반환: (쌍 목록, 건너뜀, 상태표기건너뜀, 중복)
    """
    pairs, seen = [], set()
    skipped = 0
    status_skipped = 0
    duplicate = 0
    for i, row in enumerate(rows):
        if i == header_row_idx:  # 헤더 행 자체는 데이터에서 제외
            continue
        # 인덱스 범위 확인
        if id_idx >= len(row) or name_idx >= len(row):
            continue
        sid = row[id_idx]
        name = row[name_idx]
        if surname_idx is not None and surname_idx < len(row):
            surname = row[surname_idx]
            if surname:
                name = surname + name
        if not sid and not name:
            continue  # 완전히 빈 행은 건너뜀 집계 대상이 아니다
        sid_ok = bool(sid) and sid.isdigit()
        name_ok = bool(name) and len(name) <= MAX_HEADER_NAME_LEN
        if not (sid_ok and name_ok):
            skipped += 1
            continue
        if _is_status_word(name):
            skipped += 1
            status_skipped += 1
            continue
        if sid in seen:
            duplicate += 1  # 조용히 버리면 738행이 32명이 된다 — 반드시 센다
            continue
        seen.add(sid)
        pairs.append({"학번": sid, "이름": name})
    return pairs, skipped, status_skipped, duplicate


def _pairs_from_rows_pattern(rows):
    """행 목록에서 (학번, 이름) 쌍을 뽑는다(패턴 기반, 낮은 신뢰도). 긴 본문 셀은 이름 후보에서 제외."""
    pairs, seen = [], set()
    for row in rows:
        sid = next((c for c in row if STUDENT_ID.fullmatch(c or "")), None)
        if not sid or sid in seen:
            continue
        name = next((c for c in row if c and KOREAN_NAME.fullmatch(c)), None)
        if name:
            seen.add(sid)
            pairs.append({"학번": sid, "이름": name})
    return pairs


def _looks_like_legacy_xls(path) -> bool:
    """구형 .xls(OLE2 복합 문서)인지 매직바이트로 판정한다.

    openpyxl은 .xls를 읽지 못하는데, 확장자만 보고 텍스트 경로로 보내면
    바이너리가 깨진 문자열로 읽혀 예외도 없이 0명이 된다. 실측 코퍼스의
    17.3%(59파일)가 여기 해당했고 전부 "0명"이라는 같은 얼굴로 나왔다.
    """
    try:
        with open(path, "rb") as f:
            return f.read(8) == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"
    except OSError:
        return False


def _confirm_reasons(students, skipped, duplicate, id_label) -> list[str]:
    """헤더 경로 결과에 확인이 필요한 사유를 모은다(없으면 빈 목록).

    종전에는 헤더가 있으면 확인필요를 False로 박아 두었는데, 실측에서 나온
    조용한 오답 52건이 전부 이 경로에서 나왔다 — 정의상 한 건도 잡을 수
    없는 플래그였다. 사유를 계산해 붙이고, 깨끗한 표에는 붙이지 않는다.

    건너뜀은 사유에 넣지 않는다. 빈 행은 정상 명렬에도 흔해서 사유로 삼으면
    거의 모든 표에 딱지가 붙고, 늘 켜진 경고는 꺼진 경고만큼 쓸모없어진다.
    건너뜀은 별도 경고로 이미 건수와 함께 안내된다.
    """
    reasons = []
    if duplicate > 0:
        reasons.append(
            f"학번 중복으로 {duplicate}개 행을 채택하지 않았습니다. "
            "출석번호를 학번으로 쓰는 표라면 여러 반이 한 반으로 합쳐졌을 수 있습니다."
        )
    if students and re.sub(r"\s+", "", id_label or "") == "번호":
        reasons.append(
            "'번호' 열을 학번으로 사용했습니다. 출석번호라면 학번이 아니므로 "
            "학번 열이 따로 있는지 확인해 주세요."
        )
    return reasons


def _detect_from_rows(rows) -> dict:
    """단일 행 집합(시트 하나 또는 텍스트 파일 전체)에서 명렬을 추출한다.

    detect_roster의 옛 단일-표 로직을 시트별로 재사용할 수 있게 뽑아낸 것이다.
    반환 키: students, header_found(헤더 경로 사용 여부), 건너뜀, 상태표기건너뜀,
    중복, 성이름결합, id_label.
    """
    id_idx, name_idx, surname_idx, header_row_idx = _find_header_indices(rows)
    if id_idx is not None and name_idx is not None:
        students, skipped, status_skipped, duplicate = _pairs_from_rows_with_indices(
            rows, id_idx, name_idx, header_row_idx, surname_idx
        )
        id_label = ""
        if 0 <= header_row_idx < len(rows) and id_idx < len(rows[header_row_idx]):
            id_label = rows[header_row_idx][id_idx]
        return {
            "students": students, "header_found": True,
            "건너뜀": skipped, "상태표기건너뜀": status_skipped, "중복": duplicate,
            "성이름결합": surname_idx is not None, "id_label": id_label,
        }
    students = _pairs_from_rows_pattern(rows)
    return {
        "students": students, "header_found": False,
        "건너뜀": 0, "상태표기건너뜀": 0, "중복": 0,
        "성이름결합": False, "id_label": "",
    }


_SHEET_REPEAT_OVERLAP_RATIO = 0.8  # 교집합/작은쪽 크기 >= 이 값이면 "같은 명렬 반복"


def _merge_sheet_rosters(named_results: list) -> dict:
    """시트별 명렬 추출 결과를 학번 집합 비교로 병합할지·하나만 쓸지 판단한다.

    named_results: [(시트명, _detect_from_rows 반환값), ...] — 학생이 1명 이상
    잡힌(=데이터) 시트만 담겨 있어야 한다.

    실측(다중 데이터시트 17개)에서 9개는 시트 간 학번 교집합이 0이었다(선택과목을
    반별로 나눈 표) — 이 경우 시트 하나만 고르면 14~137명을 잃으므로 합집합이
    정답이다. 반대로 교집합이 큰 경우는 같은 명렬이 반복 등재된 것이므로 가장 큰
    시트만 쓴다. 그 사이는 확인을 요구하며 합집합을 쓴다.

    반환 키: students, extra_reasons(list[str]), sheet_names(합쳤으면 시트명 목록,
    아니면 None), 건너뜀, 상태표기건너뜀, 중복, header_found, 성이름결합, id_label.
    """
    if len(named_results) == 1:
        name, r = named_results[0]
        return {
            "students": r["students"], "extra_reasons": [], "sheet_names": None,
            "건너뜀": r["건너뜀"], "상태표기건너뜀": r["상태표기건너뜀"], "중복": r["중복"],
            "header_found": r["header_found"], "성이름결합": r["성이름결합"],
            "id_label": r["id_label"],
        }

    def ids_of(r):
        return {str(s.get("학번", "")) for s in r["students"]}

    ordered = sorted(named_results, key=lambda nr: len(nr[1]["students"]), reverse=True)
    primary_name, primary = ordered[0]
    primary_ids = ids_of(primary)

    all_disjoint = True
    all_repeat = True
    total_overlap = 0
    for _, r in ordered[1:]:
        ids = ids_of(r)
        inter = primary_ids & ids
        total_overlap += len(inter)
        if inter:
            all_disjoint = False
        smaller = min(len(primary_ids), len(ids)) or 1
        if len(inter) / smaller < _SHEET_REPEAT_OVERLAP_RATIO:
            all_repeat = False

    def _union(order):
        merged, seen = [], set()
        skipped = status_skipped = duplicate = 0
        for _, r in order:
            skipped += r["건너뜀"]
            status_skipped += r["상태표기건너뜀"]
            duplicate += r["중복"]
            for s in r["students"]:
                sid = str(s.get("학번", ""))
                if sid in seen:
                    continue
                seen.add(sid)
                merged.append(s)
        return merged, skipped, status_skipped, duplicate

    sheet_names = [name for name, _ in named_results]
    성이름결합 = any(r["성이름결합"] for _, r in named_results)

    if all_disjoint:
        merged, skipped, status_skipped, duplicate = _union([(primary_name, primary)] + ordered[1:])
        return {
            "students": merged, "extra_reasons": [], "sheet_names": sheet_names,
            "건너뜀": skipped, "상태표기건너뜀": status_skipped, "중복": duplicate,
            "header_found": primary["header_found"], "성이름결합": 성이름결합,
            "id_label": primary["id_label"],
        }

    if all_repeat:
        reason = [
            f"{len(named_results)}개 시트에 같은 명렬이 반복돼 가장 큰 시트만 사용했습니다."
        ]
        return {
            "students": primary["students"], "extra_reasons": reason, "sheet_names": None,
            "건너뜀": primary["건너뜀"], "상태표기건너뜀": primary["상태표기건너뜀"],
            "중복": primary["중복"], "header_found": primary["header_found"],
            "성이름결합": primary["성이름결합"], "id_label": primary["id_label"],
        }

    # 부분 겹침 — 합집합을 쓰되 확인을 요구한다.
    merged, skipped, status_skipped, duplicate = _union([(primary_name, primary)] + ordered[1:])
    reason = [f"시트 간 학번이 {total_overlap}건 겹쳐 합집합으로 처리했습니다. 확인해 주세요."]
    return {
        "students": merged, "extra_reasons": reason, "sheet_names": sheet_names,
        "건너뜀": skipped, "상태표기건너뜀": status_skipped, "중복": duplicate,
        "header_found": primary["header_found"], "성이름결합": 성이름결합,
        "id_label": primary["id_label"],
    }


def detect_roster(path) -> dict:
    """명단 파일에서 학번·이름 쌍을 감지한다. LLM을 거치지 않는 로컬 판정.

    엑셀은 시트를 이어붙이지 않고 시트별로 추출한 뒤 학번 집합을 비교해 합칠지
    (교집합 0), 하나만 쓸지(교집합 커서 반복 등재로 판단), 확인을 요구하며 합칠지
    (부분 겹침)를 결정한다(_merge_sheet_rosters). 데이터가 잡힌 시트가 1개면
    회귀 없이 기존과 동일하게 동작한다.
    """
    path = Path(path)
    읽기오류 = ""
    sheets: list = []
    if _looks_like_legacy_xls(path):
        # 텍스트 경로로 보내면 OLE2 바이너리가 깨진 문자열로 읽혀 예외 없이 0명이 된다.
        읽기오류 = "구형 .xls 형식은 읽을 수 없습니다. 엑셀에서 .xlsx로 다시 저장해 주세요."
    else:
        try:
            if path.suffix.lower() == ".xlsx":
                sheets = _sheets_from_xlsx(path)
            else:
                sheets = [("", _rows_from_text(path))]
        except Exception as exc:
            # 사유를 버리면 손상 파일·형식 미지원·학생 없음이 전부 "0명"으로 똑같이 보인다.
            sheets = []
            읽기오류 = f"{type(exc).__name__}: {exc}"

    per_sheet = [(name, _detect_from_rows(rows)) for name, rows in sheets]
    data_sheets = [(name, r) for name, r in per_sheet if r["students"]]

    if data_sheets:
        merged = _merge_sheet_rosters(data_sheets)
        students = merged["students"]
        skipped = merged["건너뜀"]
        status_skipped = merged["상태표기건너뜀"]
        duplicate = merged["중복"]
        성이름결합 = merged["성이름결합"]
        시트합침 = merged["sheet_names"]
        if merged["header_found"]:
            방식 = "표헤더"
            확인사유 = list(merged["extra_reasons"]) + _confirm_reasons(
                students, skipped, duplicate, merged["id_label"]
            )
        else:
            방식 = "패턴"
            확인사유 = ["표 헤더를 찾지 못해 패턴으로 추정했습니다."] + list(merged["extra_reasons"])
    else:
        students = []
        시트합침 = None
        성이름결합 = False
        skipped = sum(r["건너뜀"] for _, r in per_sheet)
        status_skipped = sum(r["상태표기건너뜀"] for _, r in per_sheet)
        duplicate = sum(r["중복"] for _, r in per_sheet)
        방식 = "실패"
        header_found_any = any(r["header_found"] for _, r in per_sheet)
        if header_found_any:
            id_label = next((r["id_label"] for _, r in per_sheet if r["header_found"]), "")
            확인사유 = _confirm_reasons(students, skipped, duplicate, id_label)
        else:
            확인사유 = ["표 헤더를 찾지 못해 패턴으로 추정했습니다."]

    result = {
        "students": students, "출처": str(path), "방식": 방식,
        "확인필요": bool(확인사유), "확인사유": 확인사유,
        "건너뜀": skipped, "상태표기건너뜀": status_skipped, "중복": duplicate,
    }
    if 읽기오류:
        result["읽기오류"] = 읽기오류
    if 성이름결합:
        result["성이름결합"] = True
    if 시트합침:
        result["시트합침"] = 시트합침
    return result


def issue_tokens(roster: dict, submitted_ids, existing: dict | None = None) -> dict:
    """제출자에게만 무작위 토큰을 발급한다. 순번이 아닌 난수여야 역추적이 어렵다."""
    mapping = {"활동": (existing or {}).get("활동"), "map": dict((existing or {}).get("map", {}))}
    submitted = {str(s) for s in submitted_ids}
    used = set(mapping["map"].values())
    for student in roster.get("students", []):
        sid = str(student.get("학번", ""))
        if sid not in submitted or sid in mapping["map"]:
            continue
        while True:
            token = "S-" + secrets.token_hex(2).upper()
            if token not in used:
                break
        used.add(token)
        mapping["map"][sid] = token
    return mapping


SHORT_NAME_WARNING = "1자 이름은 일반명사 충돌 위험으로 본문 치환에서 제외함"

# CLI 계층 fail-closed 정책용. pseudonymize_text 자체는(품질 보호를 위해) 1자
# 이름을 치환하지 않는 현재 동작을 유지하지만, 그 결과 실명이 본문에 그대로
# 남아 LLM에 전송되는 것은 별개의 심각한 개인정보 문제다. 따라서 CLI(mask,
# memo)는 파이프라인 진입 전에 명렬을 검사해, 성 결합 후에도 1자 이름이
# 남아 있으면 저장 자체를 차단한다(경고-후-진행이 아니라 중단).
DETECTION_LIMIT_WARNING = (
    "⚠ 탐지 한계: 한자·영문 표기, 별명, 3자 이상 띄어쓰기, 명렬에 없는 사람"
    "(교사·가족·친구)의 이름은 탐지하지 못합니다. 이 도구는 위험을 줄일 뿐 "
    "0으로 만들지 않습니다."
)
# 골든리포트 §6 — "경고 0건"을 "안전하다"로 오인하지 않도록, mask 실행마다
# 경고 건수와 무관하게 이 문구를 항상 stdout에 출력한다(_cmd_mask 참고).

SHORT_NAME_BLOCK_MESSAGE = (
    "중단: 1자 이름 {n}건이 명렬에 있습니다. 1자 이름은 본문에서 안전하게 가릴 수 "
    "없어(일반명사 충돌) 그대로 전송될 위험이 있습니다. 명렬의 이름 열을 성이 포함된 "
    "전체 이름으로 고쳐 주세요(성·이름이 분리된 명렬이면 합쳐 주세요)."
)


def count_short_names(roster: dict) -> int:
    """명렬에서 2자 미만(1자) 이름이 몇 건인지 센다.

    한글 성명은 성 1자 이상 + 이름 1자 이상이 결합되어야 완전하므로, 1자
    이름은 성·이름이 분리된 명렬이 아직 결합되지 않았거나 데이터 오류임을
    뜻한다 — 어느 경우든 그 이름은 일반명사와 충돌해 본문에서 안전하게 가릴
    수 없으므로 CLI 계층에서 이 값을 fail-closed 판단에 사용한다.
    """
    return sum(1 for s in roster.get("students", []) if len(s.get("이름", "")) < 2)


def pseudonymize_text(text: str, roster: dict, mapping: dict, owner_id=None):
    """학번과 명렬 이름을 토큰으로 치환한다. 본문 이름 치환은 경고로 보고한다.

    동명이인 오귀속 방지: 명렬 이름을 다른 학생의 토큰으로 치환하지 않는다.
    - owner_id가 주어지면(그 본문의 주인 학번), 그 학번의 이름만 그 학번의
      "자기 토큰"으로 치환한다. 그 외 명렬 이름(동명이인 포함)은 절대 남의
      토큰이 아니라 중립어 "급우"로 치환한다.
    - owner_id가 없으면(기존 호출부 호환) 명렬 이름은 전부 "급우"로 치환한다
      — 남의 토큰이 본문에 박히는 경로를 완전히 없앤다.

    1자 이름은 일반명사와 충돌할 위험이 압도적이므로(실제 명렬의 전체 이름은
    2자 이상) 본문 치환 대상에서 제외하고, 제외 사실만 경고에 남긴다(이름 값
    자체는 경고 문자열에 남기지 않는다 — SHORT_NAME_WARNING은 고정 문구다).

    공백 내성(골든리포트 §6 FIX A): '김 하윤'처럼 글자 사이에 공백을 넣어 쓴
    이름, '301 01'처럼 숫자 사이에 공백을 넣어 쓴 학번도 놓치지 않는다. 학번은
    숫자 경계를 유지해 '101010' 같은 무관한 숫자열은 그대로 보존한다.

    성 제외 이름(FIX B, 실측으로 범위 축소): 명렬 이름이 3자 이상이면 성을 뺀
    나머지(2자 이상, 예: '김하윤' -> '하윤')를 치환 대상에 추가하되, **인명
    전용 접미(호칭 조사: 이는/이가/이를/아/야/쌤/선배 등, GIVEN_NAME_HONORIFIC_
    SUFFIXES)가 뒤따를 때만** 치환한다 — '하윤이는'은 치환하지만 맨몸 '하윤'은
    건드리지 않는다. 실제 학생 답안 69,034자 실측에서 성 제외 이름이 맨몸으로
    등장한 4건은 전부 시어·일반어 충돌(예: 성 제외 이름이 '하늘'인 경우
    "하늘을 우러러")이었고 조사 동반 등장은 0건이었다 — 맨몸까지 치환하면
    골든리포트 §5의 결함 #3('봄봄' -> 'S-7880S-7880')과 같은 종류의 품질
    훼손만 낳고 얻는 것이 없다. 접미는 치환 후에도 보존한다('하윤이는' ->
    'S-XXXX이는', 접미까지 삼키면 문장이 깨진다). 2자 이름(예: '박봄')은 성을
    빼면 1자('봄')가 되므로 추가하지 않는다(1자 이름 차단 원칙과 동일 —
    일반명사·작품명 오탐 방지). **반드시 전체 이름을 먼저, 성 제외 이름을 그
    다음에 치환한다** — 순서가 뒤집히면 '김'+토큰처럼 성 글자가 홀로 남는
    잔재가 생긴다. 전체 이름 자체는 오탐 위험이 없으므로 지금처럼 맨몸으로도
    치환한다(이번 범위 축소는 성 제외 형태에만 적용).
    """
    warnings: list[str] = []
    out = text
    owner_id = str(owner_id) if owner_id is not None else None

    for sid, token in mapping.get("map", {}).items():
        # 학번은 고유 식별자라 동명이인 같은 오귀속 위험이 없다 — 항상 자기 토큰으로.
        # 앞뒤가 숫자가 아닐 때만 치환 (경계 보호, 공백 내성)
        out = _id_pattern(sid).sub(token, out)

    students = roster.get("students", [])
    if owner_id is not None:
        # 주인 학생의 명렬 항목을 먼저 처리해, 동명이인이 있어도 주인 이름이
        # 남의 급우 치환에 먼저 소비되지 않고 반드시 자기 토큰을 받도록 한다.
        students = sorted(students, key=lambda s: str(s.get("학번", "")) != owner_id)

    for student in students:
        name = student.get("이름", "")
        sid = str(student.get("학번", ""))
        if not name:
            continue
        if len(name) < 2:
            # 1자 이름('봄' 등 일반명사와 충돌)은 치환하지 않는다 — 작품명 등
            # 무관한 단어를 파괴하는 것이 실명 노출보다 더 심각한 품질 훼손이다.
            if _name_pattern(name).search(out):
                warnings.append(SHORT_NAME_WARNING)
            continue

        given = name[1:] if len(name) >= 3 else None

        if owner_id is not None and sid == owner_id:
            token = mapping.get("map", {}).get(sid)
            if not token:
                continue
            # 전체 이름은 맨몸으로도 치환한다 — 오탐 위험이 없다(경계 없이 무조건
            # 치환. 과소탐 방지 — 개인정보 누락이 최악. 한글은 조사로 어절 경계가
            # 흐려지므로 경계 검사를 하면 안 됨). 반드시 이 치환이 먼저 끝나야
            # 성 제외 이름 치환에서 '김'+토큰 잔재가 남지 않는다.
            out, count = _name_pattern(name).subn(token, out)
            if count > 0:
                warnings.append(f"본문에서 이름 '{name}'을 토큰으로 치환함(학번 {sid}, {count}회)")
            if given:
                # 성 제외 이름은 인명 전용 접미(호칭 조사)가 뒤따를 때만 치환한다
                # — 맨몸 등장은 시어·일반어와 충돌 위험이 실재하므로 건드리지
                # 않는다(실측 근거). 접미는 보존한다.
                out, count2 = _sub_given_name_with_suffix(out, given, token)
                if count2 > 0:
                    warnings.append(f"본문에서 성 제외 이름 '{given}'을 토큰으로 치환함(학번 {sid}, {count2}회)")
            continue

        # 자신이 아닌 명렬 이름(동명이인 포함) — 남의 토큰이 아니라 중립어로 치환
        out, count = _name_pattern(name).subn("급우", out)
        if count > 0:
            if mapping.get("map", {}).get(sid) is not None:
                warnings.append(f"본문에서 급우 이름 '{name}'을 급우로 치환함(학번 {sid}, {count}회)")
            else:
                # 토큰이 없는 명렬 이름 = 미제출자. 제출자 본문에 미제출 급우 이름이
                # 나오면 실명이 그대로 전송되므로 중립 대체어로 치환한다.
                warnings.append(f"본문에서 미제출자 이름 '{name}'을 급우로 치환함")
        if given:
            out, count2 = _sub_given_name_with_suffix(out, given, "급우")
            if count2 > 0:
                if mapping.get("map", {}).get(sid) is not None:
                    warnings.append(f"본문에서 급우 이름(성 제외) '{given}'을 급우로 치환함(학번 {sid}, {count2}회)")
                else:
                    warnings.append(f"본문에서 미제출자 급우 이름(성 제외) '{given}'을 급우로 치환함")
    return out, warnings


def reidentify(text: str, mapping: dict) -> str:
    """토큰을 학번으로 되돌린다(로컬 재결합).

    LLM이 토큰 코어를 소문자로 출력할 수 있으므로 대소문자를 무시하고 치환한다.
    """
    out = text
    for sid, token in mapping.get("map", {}).items():
        # 토큰도 앞뒤가 숫자가 아닐 때만 치환 (경계 보호)
        out = re.sub(rf"(?<!\d){re.escape(token)}(?!\d)", sid, out, flags=re.IGNORECASE)
    return out


TOKEN_PATTERN = re.compile(r"(?<![0-9A-Za-z])S-[0-9A-Fa-f]{4}(?![0-9A-Fa-f])")


def scan_leak(text: str, roster: dict, scope: str = "구조"):
    """LLM 전송 전후 텍스트에서 식별정보를 찾는다.

    학번은 어디서 발견되든 FAIL(오탐이 없는 강한 신호). 공백을 넣어 쓴 학번
    ('301 01')도 같은 공백 내성 패턴으로 잡는다(FIX A) — 치환은 하는데 검사는
    못 잡으면 게이트가 새는 셈이라 pseudonymize_text와 같은 관용도를 쓴다.
    이름은 구조 필드에서만 FAIL이고 본문에서는 WARN이다 — 일반명사와 겹치는
    이름('봄' 등)은 원리적으로 100% 탐지가 불가능하므로 게이트로 삼지 않는다.
    성 제외 이름(3자 이상 이름의 나머지, 예: '하윤')은 인명 전용 접미(호칭
    조사)가 뒤따를 때만 검사한다 — pseudonymize_text의 치환 조건과 반드시
    일치시킨다(치환은 안 하는데 검사만 하면, 혹은 그 반대면 게이트가 새거나
    무의미한 경고가 남발된다). 맨몸 등장은 시어·일반어 충돌 실측 근거로
    검사하지 않는다. 접미가 있어 검사 대상이 되어도 스코프와 무관하게 항상
    WARN이다(FAIL 아님, FIX B) — 오탐 위험이 여전히 남기 때문이다.
    """
    issues: list[tuple[str, str, str]] = []
    for sid in {str(s.get("학번", "")) for s in roster.get("students", [])}:
        # 학번은 숫자 경계로 검사한다 — pseudonymize_text가 일부러 보존하는
        # "101010번" 같은 무관한 숫자열 안의 부분 문자열을 오탐하면 교착이 생긴다.
        if sid and _id_pattern(sid).search(text):
            issues.append(("FAIL", "ID_LEAK", f"학번 {sid} 노출"))
    level = "FAIL" if scope == "구조" else "WARN"
    for name in {s.get("이름", "") for s in roster.get("students", [])}:
        if not name:
            continue
        if _name_pattern(name).search(text):
            issues.append((level, "NAME_LEAK", f"이름 '{name}' 노출({scope})"))
        if len(name) >= 3:
            given = name[1:]
            if _given_name_suffix_pattern(given).search(text):
                issues.append(("WARN", "NAME_LEAK", f"성 제외 이름 '{given}' 노출({scope})"))
    return issues


def scan_token_residue(text: str) -> list[str]:
    """최종 산출물에 남은 토큰을 찾는다(재결합 누락 감지)."""
    return TOKEN_PATTERN.findall(text)


def scan_id_in_narrative(text: str, roster: dict) -> list[str]:
    """최종 서술문에 학번이 그대로 남아 있는지 찾는다.

    재결합은 토큰을 학번으로 되돌리므로, 본문 이름 자리에 있던 토큰이
    학번 숫자로 복원되어 문장에 남을 수 있다. 토큰 잔존 검사로는 잡히지
    않으므로(S-XXXX가 이미 사라진 상태) 별도로 검사한다.
    """
    found: list[str] = []
    for student in roster.get("students", []):
        sid = str(student.get("학번", ""))
        if sid and re.search(rf"(?<!\d){re.escape(sid)}(?!\d)", text):
            found.append(sid)
    return found


MEMO_HEADER_ALIASES = {"관찰메모", "메모", "특기사항"}


def _normalize_header(cell) -> str:
    """헤더 셀 비교용 정규화 — 공백을 무시한다('관찰 메모' == '관찰메모')."""
    return re.sub(r"\s+", "", cell or "")


def _find_memo_header_indices(rows):
    """헤더 행에서 (학번열, 메모열, 헤더행번호)를 찾는다. 없으면 (None, None, -1)."""
    for i, row in enumerate(rows):
        id_idx = next((j for j, c in enumerate(row) if c in ("학번", "번호")), None)
        memo_idx = next(
            (j for j, c in enumerate(row) if _normalize_header(c) in MEMO_HEADER_ALIASES), None
        )
        if id_idx is not None and memo_idx is not None:
            return (id_idx, memo_idx, i)
    return (None, None, -1)


def _pick_data_sheet(sheets, has_data, sheet=None):
    """시트 목록에서 처리할 시트를 고른다.

    학번을 키로 학생별 값을 읽는 작업(점수·등급·메모)은 열 인덱스가 시트마다
    다르므로 시트를 넘나들며 인덱스를 재사용하면 엉뚱한 열의 값을 가져온다.
    그래서 시트별로는 항상 독립적으로 처리한다. 다만 데이터가 있는 시트가
    여럿이면(반별로 나뉜 채점표 등) roster와 같은 논리로 자동 병합을 시도할
    여지를 호출부에 남기기 위해 "multi" 결과에는 시트명이 아니라 (시트명, rows)
    쌍을 그대로 담아 반환한다 — 호출부가 학번 교집합·라벨 일치를 검사해 병합
    여부를 최종 판단한다(_merge_keyed_sheets).

    sheets: [(시트명, rows), ...]
    has_data(rows) -> bool: 이 시트가 처리 대상 데이터를 담고 있는지 판정.
    sheet: 교사가 --sheet로 지정한 시트명(있으면 존재 여부만 확인하고 그 시트를 쓴다).

    반환: (종류, payload)
    - ("ok", rows): 처리할 시트 하나가 정해짐(단일 데이터시트 또는 --sheet 지정)
    - ("invalid_sheet", [시트명, ...]): --sheet로 지정한 이름이 없음
    - ("multi", [(시트명, rows), ...]): --sheet 없이 데이터가 있는 시트가 2개 이상
    - ("none", None): 데이터가 있는 시트가 하나도 없음
    """
    if sheet is not None:
        match = next((rows for name, rows in sheets if name == sheet), None)
        if match is None:
            return ("invalid_sheet", [name for name, _ in sheets])
        return ("ok", match)

    data_sheets = [(name, rows) for name, rows in sheets if has_data(rows)]
    if not data_sheets:
        return ("none", None)
    if len(data_sheets) == 1:
        return ("ok", data_sheets[0][1])
    return ("multi", data_sheets)


def _sheet_student_ids(rows, id_idx, header_row_idx) -> set:
    """시트에서 학번 형식(다섯 자리 숫자)에 맞는 값만 학번 집합으로 모은다."""
    ids = set()
    for i, row in enumerate(rows):
        if i == header_row_idx or id_idx >= len(row):
            continue
        sid = row[id_idx]
        if sid and STUDENT_ID.fullmatch(sid):
            ids.add(sid)
    return ids


def _merge_keyed_sheets(resolved, pairs_fn):
    """학번 키로 값을 추출하는 시트 여러 개를 병합할지 판단한다(roster의
    _merge_sheet_rosters와 같은 취지 — 점수·등급·메모는 학번이 유일 키이므로
    시트 간 학번이 겹치지 않으면 합집합이 안전하다).

    점수 열 다중 후보(_resolve_score_column의 "ambiguous")를 막던 진짜 이유는
    "시트1의 열 인덱스로 시트2를 읽는 것"이었고, 그건 시트별 독립 해석으로 이미
    해결됐다. 합치는 것 자체는 위험하지 않으므로, 학번이 겹치지 않고 각 시트가
    가리키는 열의 의미(라벨)가 같으면 자동으로 합친다. 학번이 겹치면(같은
    학생에 값이 둘이라 어느 쪽이 맞는지 알 수 없음) 또는 라벨이 다르면(열 문자만
    같고 의미가 다른 경우가 가장 위험하다) 합치지 않고 확인을 요구한다.

    resolved: [(시트명, rows, resolution), ...] — resolution은 최소
    id_idx, header_row_idx, letter, label 키를 가져야 한다.
    pairs_fn(rows, resolution) -> (pairs, no_count)

    반환:
    - ("id_overlap", [시트명, ...])
    - ("label_mismatch", [{"sheet","letter","label"}, ...])
    - ("ok", 병합된 pairs, no_count 합, letter, label, [시트명, ...])
    """
    ids_by_sheet = [
        (name, _sheet_student_ids(rows, r["id_idx"], r["header_row_idx"]))
        for name, rows, r in resolved
    ]
    for i in range(len(ids_by_sheet)):
        for j in range(i + 1, len(ids_by_sheet)):
            if ids_by_sheet[i][1] & ids_by_sheet[j][1]:
                return ("id_overlap", [name for name, _, _ in resolved])

    labels = {r["label"] for _, _, r in resolved}
    if len(labels) > 1:
        entries = [{"sheet": name, "letter": r["letter"], "label": r["label"]} for name, _, r in resolved]
        return ("label_mismatch", entries)

    merged_pairs = []
    total_no = 0
    for name, rows, r in resolved:
        pairs, no_count = pairs_fn(rows, r)
        merged_pairs.extend(pairs)
        total_no += no_count
    letter = resolved[0][2]["letter"]
    label = resolved[0][2]["label"]
    sheet_names = [name for name, _, _ in resolved]
    return ("ok", merged_pairs, total_no, letter, label, sheet_names)


def _resolve_memo_column(rows) -> dict:
    """메모 열을 해석한다. 반환은 최소 status 키를 포함하며, status가 "ok"이면
    id_idx, header_row_idx, memo_idx, letter, label도 채워진다.

    label은 실제 헤더 셀 원문('관찰 메모'/'메모' 등)이다 — 시트마다 이 텍스트가
    다르면 같은 열 문자라도 의미가 다를 수 있으므로 병합 여부 판단(_merge_
    keyed_sheets)에 이 값을 쓴다.
    """
    id_idx, memo_idx, header_row_idx = _find_memo_header_indices(rows)
    if id_idx is None or memo_idx is None:
        return {"status": "not_found"}
    row = rows[header_row_idx] if header_row_idx < len(rows) else []
    raw = row[memo_idx].strip() if memo_idx < len(row) and row[memo_idx] else ""
    label = raw or _col_letter(memo_idx)
    return {
        "status": "ok", "id_idx": id_idx, "header_row_idx": header_row_idx,
        "memo_idx": memo_idx, "letter": _col_letter(memo_idx), "label": label,
    }


def _memo_has_data(rows) -> bool:
    return _resolve_memo_column(rows)["status"] == "ok"


def _memo_pairs_from_resolution(rows, resolution):
    id_idx = resolution["id_idx"]
    memo_idx = resolution["memo_idx"]
    header_row_idx = resolution["header_row_idx"]
    pairs = []
    for i, row in enumerate(rows):
        if i == header_row_idx:
            continue
        if id_idx >= len(row) or memo_idx >= len(row):
            continue
        sid = row[id_idx]
        memo = row[memo_idx]
        if not sid or not STUDENT_ID.fullmatch(sid):
            continue
        if not memo:
            continue
        pairs.append((sid, memo))
    return pairs, 0


def _parse_memo_from_rows(rows):
    resolution = _resolve_memo_column(rows)
    if resolution["status"] != "ok":
        return None
    pairs, _ = _memo_pairs_from_resolution(rows, resolution)
    return pairs, None


def parse_memo_xlsx(path, sheet=None):
    """관찰 메모 xlsx에서 (학번, 메모) 쌍을 추출한다. 시트별로 독립 처리한다 —
    열 인덱스는 시트마다 다르므로 시트를 넘나들며 재사용하지 않는다.

    --sheet 없이 메모 데이터가 있는 시트가 2개 이상이면 명렬(roster)과 같은
    논리로 판단한다: 학번 교집합이 없고 메모 열 라벨이 같으면 자동 병합하고,
    학번이 겹치거나 라벨이 다르면 exit 1로 멈춘다(호출부가 안내).

    반환:
    - None: 메모 열 헤더('관찰 메모'/'관찰메모'/'메모'/'특기사항' 중 하나, 공백
      무시)를 찾지 못함(호출부가 exit 1 + 안내로 처리)
    - {"invalid_sheet": True, "sheets": [...]}: --sheet로 지정한 시트가 없음
    - {"sheet_ambiguous": True, "sheets": [...]}: 학번이 겹쳐 자동 병합할 수
      없음(어느 시트 값이 맞는지 알 수 없음)
    - {"label_mismatch": True, "sheets": [{"sheet","letter","label"}, ...]}:
      시트마다 메모 열의 헤더 텍스트가 달라 같은 의미인지 확인이 필요함
    - (pairs, 병합시트목록 또는 None): 정상 — pairs는 (학번, 메모) 리스트
    """
    sheets = _sheets_from_xlsx(path)
    kind, payload = _pick_data_sheet(sheets, _memo_has_data, sheet=sheet)
    if kind == "invalid_sheet":
        return {"invalid_sheet": True, "sheets": payload}
    if kind == "none":
        return None
    if kind == "ok":
        return _parse_memo_from_rows(payload)

    resolved = [(name, rows, _resolve_memo_column(rows)) for name, rows in payload]
    merge_kind = _merge_keyed_sheets(resolved, _memo_pairs_from_resolution)
    if merge_kind[0] == "id_overlap":
        return {"sheet_ambiguous": True, "sheets": merge_kind[1]}
    if merge_kind[0] == "label_mismatch":
        return {"label_mismatch": True, "sheets": merge_kind[1]}
    _, merged_pairs, _total_no, _letter, _label, sheet_names = merge_kind
    return merged_pairs, sheet_names


_MEMO_LINE = re.compile(r"^(\d+)\s*[:]?\s*(.*)$")


def parse_memo_text(path):
    """`학번: 메모` 또는 `학번 메모` 형식의 텍스트에서 (학번, 메모) 쌍을 추출한다."""
    # utf-8-sig: BOM 붙은 파일(메모장·PowerShell 기본 저장)도 조용히 실패하지 않게 한다.
    text = Path(path).read_text(encoding="utf-8-sig", errors="replace")
    pairs = []
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        m = _MEMO_LINE.match(line)
        if not m:
            continue
        sid, memo = m.group(1), m.group(2).strip()
        if not memo:
            continue
        pairs.append((sid, memo))
    return pairs


SCORE_HEADER_ALIASES = {"점수", "총점", "합계", "평가점수"}

# '총점'은 그 자체로 전체 합계를 뜻하므로 상위 병합 라벨(예: "소설 비평하기")을
# 붙이지 않는다. 반대로 '점수'/'합계'/'평가점수'는 여러 하위 영역(소설 비평하기,
# 시 비평하기 등)에 반복해서 나타날 수 있는 일반 라벨이라, 왼쪽에서 가장 가까운
# 상위 구간 라벨과 결합해야만("소설 비평하기 > 점수") 서로 구별된다.
GROUP_TERMINAL_ALIASES = {"총점"}

_COLUMN_LETTER_RE = re.compile(r"[A-Za-z]{1,3}")


def _looks_like_column_letter(s: str) -> bool:
    """'--column'에 엑셀 열 문자(J, R, AA 등)가 주어졌는지 판별한다."""
    return bool(_COLUMN_LETTER_RE.fullmatch((s or "").strip()))


def _col_letter(idx: int) -> str:
    """0-based 열 인덱스를 엑셀 열 문자로 변환한다(0->A, 25->Z, 26->AA)."""
    idx += 1
    letters = ""
    while idx > 0:
        idx, rem = divmod(idx - 1, 26)
        letters = chr(65 + rem) + letters
    return letters


def _col_index_from_letter(letter: str) -> int:
    """엑셀 열 문자를 0-based 인덱스로 변환한다(A->0, Z->25, AA->26)."""
    letter = letter.strip().upper()
    idx = 0
    for ch in letter:
        idx = idx * 26 + (ord(ch) - 64)
    return idx - 1


def _find_label_row(rows):
    """학번/번호 열이 있는 행을 라벨 행(실제 열 헤더가 나열되는 행)으로 판정한다.

    제목 행(전체 병합)이나 하위 문항 행(문항1, 문항2 등)과 구분하기 위한
    기준이다 — 학번 열이 있는 행이 곧 "이 표의 열 이름이 나열된 행"이다.
    없으면 (None, -1).
    """
    for i, row in enumerate(rows):
        id_idx = next((j for j, c in enumerate(row) if c in ("학번", "번호")), None)
        if id_idx is not None:
            return (id_idx, i)
    return (None, -1)


def _find_score_candidates(rows):
    """라벨 행에서 점수 열 후보를 전부 찾는다(다중 헤더의 병합 라벨 결합 포함).

    라벨 행을 왼쪽에서 오른쪽으로 훑으며 "현재 구간 라벨"을 갱신한다(학번·이름
    열과 점수 별칭 자체는 구간 라벨 후보에서 제외). 점수 별칭 열을 만나면 그
    시점의 구간 라벨을 붙여 후보로 기록한다 — 병합 셀은 read_only 모드에서
    왼쪽 상단 셀에만 값이 있고 나머지는 빈 문자열로 읽히므로, 빈 셀을 만나도
    구간 라벨을 그대로 유지하면 자연히 병합 범위 전체에 적용된다.

    반환: [(col_idx, label), ...]. col_idx는 0-based.
    """
    id_idx, header_row_idx = _find_label_row(rows)
    if id_idx is None:
        return []
    row = rows[header_row_idx]
    candidates = []
    current_group = None
    for j, cell in enumerate(row):
        if j == id_idx:
            continue
        normalized = _normalize_header(cell)
        if not normalized or normalized == "이름":
            continue
        if normalized in SCORE_HEADER_ALIASES:
            label = cell.strip()
            if normalized not in GROUP_TERMINAL_ALIASES and current_group:
                label = f"{current_group} > {label}"
            candidates.append((j, label))
        else:
            current_group = cell.strip()
    return candidates


def _resolve_score_column(rows, column=None):
    """점수 열을 결정한다(다중 후보 시 자동 선택 금지 — fail-closed).

    반환은 항상 "status" 키를 포함한 dict다:
    - {"status": "no_id"}: 학번/번호 열 자체를 못 찾음
    - {"status": "not_found"}: 점수 열 후보가 0개(또는 --column 지정 열을 못 찾음)
    - {"status": "ambiguous", "candidates": [{"letter","label","col_idx"}, ...]}:
      --column 없이 점수 열 후보가 2개 이상이라 자동 선택을 거부해야 하는 경우
    - {"status": "ok", "id_idx", "score_idx", "header_row_idx", "letter", "label"}
    """
    id_idx, header_row_idx = _find_label_row(rows)
    if id_idx is None:
        return {"status": "no_id"}

    candidates = _find_score_candidates(rows)
    candidate_map = dict(candidates)

    if column:
        column = column.strip()
        if _looks_like_column_letter(column):
            col_idx = _col_index_from_letter(column)
            label = candidate_map.get(col_idx)
            if label is None:
                # 점수 별칭이 아닌 열을 열 문자로 직접 지정한 경우 — 원래 헤더
                # 텍스트가 있으면 그것을, 없으면 열 문자 자체를 라벨로 쓴다.
                row = rows[header_row_idx] if header_row_idx < len(rows) else []
                raw = row[col_idx].strip() if col_idx < len(row) and row[col_idx] else ""
                label = raw or _col_letter(col_idx)
            return {
                "status": "ok", "id_idx": id_idx, "score_idx": col_idx,
                "header_row_idx": header_row_idx,
                "letter": _col_letter(col_idx), "label": label,
            }
        target = _normalize_header(column)
        row = rows[header_row_idx] if header_row_idx < len(rows) else []
        col_idx = next((j for j, c in enumerate(row) if _normalize_header(c) == target), None)
        if col_idx is None:
            return {"status": "not_found"}
        label = candidate_map.get(col_idx, column)
        return {
            "status": "ok", "id_idx": id_idx, "score_idx": col_idx,
            "header_row_idx": header_row_idx,
            "letter": _col_letter(col_idx), "label": label,
        }

    if not candidates:
        return {"status": "not_found"}
    if len(candidates) > 1:
        return {
            "status": "ambiguous",
            "candidates": [
                {"letter": _col_letter(idx), "label": label, "col_idx": idx}
                for idx, label in candidates
            ],
        }
    col_idx, label = candidates[0]
    return {
        "status": "ok", "id_idx": id_idx, "score_idx": col_idx,
        "header_row_idx": header_row_idx,
        "letter": _col_letter(col_idx), "label": label,
    }


def _resolve_named_column(rows, column):
    """'--grade-column'처럼 점수 별칭과 무관하게 열 문자 또는 헤더 이름을 직접
    지정하는 경우에 쓴다. 반환: (id_idx, header_row_idx, col_idx) — 못 찾으면
    col_idx가 None이다.
    """
    id_idx, header_row_idx = _find_label_row(rows)
    if id_idx is None:
        return None, -1, None
    column = (column or "").strip()
    if _looks_like_column_letter(column):
        return id_idx, header_row_idx, _col_index_from_letter(column)
    target = _normalize_header(column)
    row = rows[header_row_idx] if header_row_idx < len(rows) else []
    col_idx = next((j for j, c in enumerate(row) if _normalize_header(c) == target), None)
    return id_idx, header_row_idx, col_idx


def _parse_column_range(spec: str):
    """'E:J' 형식의 열 범위를 (시작 col_idx, 끝 col_idx) 0-based 포함 구간으로
    변환한다. 형식이 올바르지 않으면 None."""
    m = re.fullmatch(r"\s*([A-Za-z]{1,3})\s*:\s*([A-Za-z]{1,3})\s*", spec or "")
    if not m:
        return None
    start = _col_index_from_letter(m.group(1))
    end = _col_index_from_letter(m.group(2))
    if start > end:
        start, end = end, start
    return start, end


def _parse_number(cell):
    """숫자 문자열을 float로 변환한다. 비었거나 숫자가 아니면 None."""
    if not cell:
        return None
    try:
        return float(cell)
    except ValueError:
        return None


def _score_has_data(rows, column) -> bool:
    resolution = _resolve_score_column(rows, column=column)
    return resolution["status"] not in ("no_id", "not_found")


def _score_pairs_from_resolution(rows, resolution):
    id_idx = resolution["id_idx"]
    score_idx = resolution["score_idx"]
    header_row_idx = resolution["header_row_idx"]

    pairs = []
    no_score = 0
    for i, row in enumerate(rows):
        if i == header_row_idx:
            continue
        if id_idx >= len(row) or score_idx >= len(row):
            continue
        sid = row[id_idx]
        if not sid or not STUDENT_ID.fullmatch(sid):
            continue
        score = _parse_number(row[score_idx])
        if score is None:
            no_score += 1
            continue
        pairs.append((sid, score))
    return pairs, no_score


def _parse_score_from_rows(rows, column):
    resolution = _resolve_score_column(rows, column=column)

    if resolution["status"] in ("no_id", "not_found"):
        return None
    if resolution["status"] == "ambiguous":
        return {"ambiguous": True, "candidates": resolution["candidates"]}

    pairs, no_score = _score_pairs_from_resolution(rows, resolution)
    return pairs, no_score, resolution["letter"], resolution["label"], None


def parse_score_xlsx(path, column=None, sheet=None):
    """채점표 xlsx에서 점수 열을 찾아 (학번, 점수) 쌍을 추출한다. 시트별로
    독립 처리한다 — 열 인덱스는 시트마다 다르므로 시트를 넘나들며 인덱스를
    재사용하지 않는다(재사용하면 시트1의 헤더 인덱스로 시트2의 엉뚱한 열 값을
    점수로 읽는 사고가 난다).

    다중 헤더(제목 행 + 병합 구간 행 + 하위 문항 행)에서 점수 열 후보가 여럿
    이면(예: 소설 비평하기 점수, 시 비평하기 점수, 총점) 조용히 하나를 고르지
    않는다 — 반 전체의 톤 등급이 엉뚱한 열에서 파생되는 사고를 막기 위해서다.

    --sheet 없이 점수 데이터가 있는 시트가 2개 이상이면 명렬(roster)과 같은
    논리로 판단한다 — 진짜 위험은 "시트를 합치는 것"이 아니라 "시트1의 열
    인덱스로 시트2를 읽는 것"이었고 그건 시트별 독립 해석으로 이미 해결됐다.
    그래서 시트 간 학번 교집합이 없고 해석된 점수 열의 라벨이 같으면(예:
    "시 비평하기 > 점수") 자동으로 합쳐 진행한다. 학번이 겹치면(같은 학생에
    값이 둘이라 어느 쪽이 맞는지 알 수 없음) 또는 라벨이 다르면(열 문자만
    같고 의미가 다른 경우가 가장 위험함) 합치지 않고 멈춘다.

    반환:
    - None: 학번 열이 없거나, 점수 열 후보가 하나도 없거나, --column으로 지정한
      열을 찾지 못한 경우(호출부가 exit 1 + 안내로 처리)
    - {"invalid_sheet": True, "sheets": [...]}: --sheet로 지정한 시트가 없음
    - {"sheet_ambiguous": True, "sheets": [...]}: 시트 간 학번이 겹쳐 자동 병합
      불가 — --sheet로 지정해야 함
    - {"label_mismatch": True, "sheets": [{"sheet","letter","label"}, ...]}:
      시트마다 해석된 점수 열의 의미(라벨)가 달라 확인이 필요함
    - {"ambiguous": True, "candidates": [...]}: --column 없이 (한 시트 안에서)
      점수 열 후보가 2개 이상이라 자동 선택을 거부해야 하는 경우
    - (pairs, no_score, letter, label, merged_sheets): 정상 — pairs는 (학번,
      점수) 리스트, no_score는 학번은 있으나 점수가 비었거나 숫자가 아니어서
      건너뛴 행 수, letter/label은 실제로 사용한 열의 표시용 정보, merged_sheets
      는 자동 병합된 시트명 목록(단일 시트면 None)이다.
    """
    sheets = _sheets_from_xlsx(path)
    kind, payload = _pick_data_sheet(sheets, lambda rows: _score_has_data(rows, column), sheet=sheet)
    if kind == "invalid_sheet":
        return {"invalid_sheet": True, "sheets": payload}
    if kind == "none":
        return None
    if kind == "ok":
        return _parse_score_from_rows(payload, column)

    resolved = []
    for name, rows in payload:
        resolution = _resolve_score_column(rows, column=column)
        if resolution["status"] == "ambiguous":
            return {"ambiguous": True, "candidates": resolution["candidates"]}
        resolved.append((name, rows, resolution))

    merge_kind = _merge_keyed_sheets(resolved, _score_pairs_from_resolution)
    if merge_kind[0] == "id_overlap":
        return {"sheet_ambiguous": True, "sheets": merge_kind[1]}
    if merge_kind[0] == "label_mismatch":
        return {"label_mismatch": True, "sheets": merge_kind[1]}
    _, merged_pairs, total_no, letter, label, sheet_names = merge_kind
    return merged_pairs, total_no, letter, label, sheet_names


def _resolve_grade_column(rows, column) -> dict:
    id_idx, header_row_idx, col_idx = _resolve_named_column(rows, column)
    if id_idx is None or col_idx is None:
        return {"status": "not_found"}
    row = rows[header_row_idx] if header_row_idx < len(rows) else []
    raw = row[col_idx].strip() if col_idx < len(row) and row[col_idx] else ""
    label = raw or _col_letter(col_idx)
    return {
        "status": "ok", "id_idx": id_idx, "header_row_idx": header_row_idx,
        "col_idx": col_idx, "letter": _col_letter(col_idx), "label": label,
    }


def _grade_has_data(rows, column) -> bool:
    return _resolve_grade_column(rows, column)["status"] == "ok"


def _grade_pairs_from_resolution(rows, resolution):
    id_idx = resolution["id_idx"]
    col_idx = resolution["col_idx"]
    header_row_idx = resolution["header_row_idx"]

    pairs = []
    no_grade = 0
    for i, row in enumerate(rows):
        if i == header_row_idx:
            continue
        if id_idx >= len(row):
            continue
        sid = row[id_idx]
        if not sid or not STUDENT_ID.fullmatch(sid):
            continue
        grade = row[col_idx].strip() if col_idx < len(row) and row[col_idx] else ""
        if not grade:
            no_grade += 1
            continue
        pairs.append((sid, grade))
    return pairs, no_grade


def _parse_grade_from_rows(rows, column):
    resolution = _resolve_grade_column(rows, column)
    if resolution["status"] != "ok":
        return None
    pairs, no_grade = _grade_pairs_from_resolution(rows, resolution)
    return pairs, no_grade, None


def parse_grade_xlsx(path, column, sheet=None):
    """채점표 xlsx에서 등급 열의 값을 점수 대신 그대로 가져온다((학번, 등급) 쌍).
    시트별로 독립 처리한다(열 인덱스를 시트 간에 재사용하지 않는다).

    점수 열이 아예 없는 채점표(A/B/C, 상/중/하 등 등급만 있는 경우)를 위한
    경로다. column은 열 문자 또는 헤더 이름 중 하나다. --sheet 없이 데이터가
    있는 시트가 여럿이면 parse_score_xlsx와 같은 논리로 학번 교집합·라벨
    일치를 검사해 자동 병합하거나 멈춘다.

    반환: None(못 찾음) | {"invalid_sheet"...} | {"sheet_ambiguous"...}
    (학번 겹침) | {"label_mismatch"...}(시트마다 등급 열 헤더가 다름) |
    (pairs, no_grade_count, merged_sheets) — merged_sheets는 자동 병합된
    시트명 목록(단일 시트면 None).
    """
    sheets = _sheets_from_xlsx(path)
    kind, payload = _pick_data_sheet(sheets, lambda rows: _grade_has_data(rows, column), sheet=sheet)
    if kind == "invalid_sheet":
        return {"invalid_sheet": True, "sheets": payload}
    if kind == "none":
        return None
    if kind == "ok":
        return _parse_grade_from_rows(payload, column)

    resolved = [(name, rows, _resolve_grade_column(rows, column)) for name, rows in payload]
    merge_kind = _merge_keyed_sheets(resolved, _grade_pairs_from_resolution)
    if merge_kind[0] == "id_overlap":
        return {"sheet_ambiguous": True, "sheets": merge_kind[1]}
    if merge_kind[0] == "label_mismatch":
        return {"label_mismatch": True, "sheets": merge_kind[1]}
    _, merged_pairs, total_no, _letter, _label, sheet_names = merge_kind
    return merged_pairs, total_no, sheet_names


def _resolve_sum_column(rows, bounds) -> dict:
    id_idx, header_row_idx = _find_label_row(rows)
    if id_idx is None:
        return {"status": "not_found"}
    start, end = bounds
    row = rows[header_row_idx] if header_row_idx < len(rows) else []
    headers = []
    for c in range(start, min(end, len(row) - 1) + 1):
        cell = row[c].strip() if c < len(row) and row[c] else ""
        if cell:
            headers.append(cell)
    letter = f"{_col_letter(start)}~{_col_letter(end)}"
    label = " · ".join(headers) if headers else letter
    return {
        "status": "ok", "id_idx": id_idx, "header_row_idx": header_row_idx,
        "start": start, "end": end, "letter": letter, "label": label,
    }


def _sum_has_data(rows, bounds) -> bool:
    return _resolve_sum_column(rows, bounds)["status"] == "ok"


def _sum_pairs_from_resolution(rows, resolution):
    id_idx = resolution["id_idx"]
    header_row_idx = resolution["header_row_idx"]
    start, end = resolution["start"], resolution["end"]

    pairs = []
    no_score = 0
    for i, row in enumerate(rows):
        if i == header_row_idx:
            continue
        if id_idx >= len(row):
            continue
        sid = row[id_idx]
        if not sid or not STUDENT_ID.fullmatch(sid):
            continue
        total = 0.0
        any_value = False
        for c in range(start, min(end, len(row) - 1) + 1):
            v = _parse_number(row[c])
            if v is not None:
                total += v
                any_value = True
        if not any_value:
            no_score += 1
            continue
        pairs.append((sid, total))
    return pairs, no_score


def _parse_sum_from_rows(rows, bounds):
    resolution = _resolve_sum_column(rows, bounds)
    if resolution["status"] != "ok":
        return None
    pairs, no_score = _sum_pairs_from_resolution(rows, resolution)
    return pairs, no_score, None


def parse_sum_columns_xlsx(path, column_range, sheet=None):
    """채점표 xlsx에서 지정 범위(예: 'E:J') 열의 숫자를 합산해 점수로 쓴다.
    시트별로 독립 처리한다(열 인덱스를 시트 간에 재사용하지 않는다).

    점수 열이 따로 없고 문항별 점수만 있는 채점표를 위한 경로다. 범위 내
    비었거나 숫자가 아닌 셀은 0으로 취급하지 않고 그냥 건너뛰며(부분 합산),
    범위 전체가 비었으면 그 행은 no_score로 집계한다. --sheet 없이 데이터가
    있는 시트가 여럿이면 parse_score_xlsx와 같은 논리로 자동 병합하거나
    멈춘다 — 라벨은 그 범위 안의 헤더 텍스트를 이어붙인 값이라, 같은 범위라도
    시트마다 다른 문항을 가리키면(열이 밀린 경우) 라벨이 달라져 멈춘다.

    반환: None(학번 열을 못 찾거나 범위 형식이 잘못됨) | {"invalid_sheet"...} |
    {"sheet_ambiguous"...}(학번 겹침) | {"label_mismatch"...}(시트마다 범위
    안의 헤더가 다름) | (pairs, no_score_count, merged_sheets).
    """
    bounds = _parse_column_range(column_range)
    if bounds is None:
        return None
    sheets = _sheets_from_xlsx(path)
    kind, payload = _pick_data_sheet(sheets, lambda rows: _sum_has_data(rows, bounds), sheet=sheet)
    if kind == "invalid_sheet":
        return {"invalid_sheet": True, "sheets": payload}
    if kind == "none":
        return None
    if kind == "ok":
        return _parse_sum_from_rows(payload, bounds)

    resolved = [(name, rows, _resolve_sum_column(rows, bounds)) for name, rows in payload]
    merge_kind = _merge_keyed_sheets(resolved, _sum_pairs_from_resolution)
    if merge_kind[0] == "id_overlap":
        return {"sheet_ambiguous": True, "sheets": merge_kind[1]}
    if merge_kind[0] == "label_mismatch":
        return {"label_mismatch": True, "sheets": merge_kind[1]}
    _, merged_pairs, total_no, _letter, _label, sheet_names = merge_kind
    return merged_pairs, total_no, sheet_names


def parse_score_paste(obj) -> dict:
    """붙여넣기 입력(점수원본.json 스키마)에서 점수/등급 쌍을 추출한다.

    명렬입력.html 3단계("점수 (선택)")가 만드는 형태다: 교사가 채점표에서
    학번·점수(또는 학번·등급) 두 열을 화면에서 직접 골라 붙여넣으므로, 이
    경로에는 애초에 "점수 열 후보가 여럿이라 모호함"이라는 상태 자체가
    존재하지 않는다 — 실측에서 채점표 37개 중 20개(54.1%)가 점수 열 후보를
    정확히 하나만 내놓았는데 그 하나가 틀린 활동의 열이었고, 후보가 1개면
    _resolve_score_column의 ambiguous 분기가 아예 발동하지 않아 질문 트리거가
    원리적으로 생기지 않았다. 파일을 파싱하는 한 이 오염은 원천적으로 잡을
    수 없으므로, 교사가 이미 골라 놓은 값을 그대로 신뢰하는 것이 유일한 해법이다.

    스키마: {"항목": [{"학번": "30101", "점수": 15}, {"학번": "30102", "등급": "중"}]}.
    한 항목에는 "점수"(숫자) 또는 "등급"(문자열) 중 하나만 있어야 한다.

    반환은 항상 "status" 키를 포함한 dict다:
    - {"status": "error", "message": str}: 항목이 비어 있거나, 항목의 형식이
      올바르지 않거나(학번 누락, 점수/등급 동시 존재 또는 둘 다 부재, 타입 오류),
      파일 안에 점수 항목과 등급 항목이 섞여 있는 경우(톤 파생 규칙이 서로
      달라 하나로 통일해야 한다)
    - {"status": "ok", "mode": "score"|"grade", "pairs": [(학번, 값), ...]}
    """
    items = obj.get("항목") if isinstance(obj, dict) else None
    if not isinstance(items, list) or not items:
        return {"status": "error", "message": "붙여넣기 입력의 '항목'이 비어 있습니다."}

    score_pairs: list = []
    grade_pairs: list = []
    invalid = 0
    for it in items:
        if not isinstance(it, dict):
            invalid += 1
            continue
        sid = it.get("학번")
        has_score = "점수" in it
        has_grade = "등급" in it
        if not sid or (has_score and has_grade) or not (has_score or has_grade):
            invalid += 1
            continue
        if has_score:
            v = it["점수"]
            if isinstance(v, bool) or not isinstance(v, (int, float)):
                invalid += 1
                continue
            score_pairs.append((str(sid), v))
        else:
            v = it["등급"]
            if not isinstance(v, str) or not v.strip():
                invalid += 1
                continue
            grade_pairs.append((str(sid), v))

    if invalid > 0:
        return {
            "status": "error",
            "message": (
                f"형식이 올바르지 않은 항목 {invalid}건이 있습니다 "
                "(학번 누락, 점수·등급 동시 존재/부재, 또는 타입 오류)."
            ),
        }
    if score_pairs and grade_pairs:
        return {
            "status": "error",
            "message": (
                f"점수 항목 {len(score_pairs)}건과 등급 항목 {len(grade_pairs)}건이 섞여 "
                "있습니다. 한 가지로 통일해 주세요(점수와 등급은 톤 파생 규칙이 다릅니다)."
            ),
        }
    if score_pairs:
        return {"status": "ok", "mode": "score", "pairs": score_pairs}
    return {"status": "ok", "mode": "grade", "pairs": grade_pairs}


def _xlsx_is_empty(path) -> bool:
    """워크북에 실제 내용(비공백 셀)이 하나도 없는지 확인한다. 시트 중 하나라도
    내용이 있으면 비어있지 않은 것으로 판정한다."""
    sheets = _sheets_from_xlsx(path)
    return not any(any(c.strip() for c in row) for _, rows in sheets for row in rows)


MAPPING_GLOB = "매핑*.json"

# 가명처리 파이프라인이 로컬에 남기는 모든 중간 산출물. 명렬(전체 학급 실명
# 원장)도 매핑표와 마찬가지로 무기한 평문으로 남으면 안 되므로 파기·잔존
# 감지 대상에 포함한다.
SENSITIVE_GLOBS = ("매핑*.json", "명렬*.json", "관찰메모*.json", "점수*.json",
                   "토큰본*.json", "제출자*.json", "검수번들*.json", "검수본*.json")


def save_mapping(mapping: dict, path) -> None:
    """매핑표를 로컬에 저장한다. 이 파일은 가명처리의 '추가 정보'이므로 로컬 전용이다."""
    with open(path, "w", encoding="utf-8") as f:
        json.dump(mapping, f, ensure_ascii=False, indent=1)


def load_mapping(path):
    try:
        with open(path, encoding="utf-8-sig") as f:
            return json.load(f)
    except (OSError, ValueError):
        return None


def destroy_mapping(path) -> bool:
    """지정 경로의 매핑표를 파기한다. 파기 = 파일 부존재 확인 수준."""
    path = Path(path)
    if path.exists():
        path.unlink()
        return True
    return False


def detect_stale_artifacts(dir_path) -> list:
    """이전 실행이 비정상 종료돼 남은 가명처리 중간 산출물(매핑표·명렬·관찰메모·
    점수·토큰본)을 찾는다(실행 전 점검용). 이름·학번 등 내용은 반환하지 않고
    파일 경로만 반환한다.
    """
    found: list[Path] = []
    for pattern in SENSITIVE_GLOBS:
        found.extend(Path(dir_path).glob(pattern))
    return sorted(found)


# 이전 이름 유지(하위 호환) — 매핑표만이 아니라 명렬·관찰메모·점수·토큰본 등
# 모든 민감 산출물을 함께 찾는다.
detect_stale_mapping = detect_stale_artifacts


def destroy_artifacts(dir_path) -> list[str]:
    """지정 디렉터리의 민감 산출물(매핑표·명렬·관찰메모·점수·토큰본)을 모두
    파기한다. 파기한 파일명 목록을 반환한다(이름·학번 등 내용은 출력하지 않는다).
    """
    destroyed: list[str] = []
    for path in detect_stale_artifacts(dir_path):
        path.unlink()
        destroyed.append(path.name)
    return destroyed


# ---------------------------------------------------------------------------
# CLI — 명렬(실명·학번)이 LLM 컨텍스트로 들어가지 않도록, 에이전트는 이 CLI를
# subprocess로 실행하고 stdout의 요약(인원수·건수)만 읽는다.
# stdout/stderr에는 이름·학번을 절대 출력하지 않는다 — 이 계약이 기능의 전부다.
# ---------------------------------------------------------------------------

def _read_json(path):
    # utf-8-sig: 교사가 명렬.json/활동프로파일.json 등을 메모장에서 열어 다시
    # 저장하면 BOM이 붙는다 — 붙어 있으면 제거하고, 없으면 그대로 읽는다.
    with open(path, encoding="utf-8-sig") as f:
        return json.load(f)


def _write_json(obj, path) -> None:
    with open(path, "w", encoding="utf-8") as f:
        json.dump(obj, f, ensure_ascii=False, indent=1)


def _cmd_roster(args) -> int:
    roster = detect_roster(args.input)
    n = len(roster.get("students", []))
    if n == 0:
        읽기오류 = roster.get("읽기오류")
        if 읽기오류:
            # 0명의 원인을 밝히지 않으면 손상·미지원·빈 표가 구별되지 않는다.
            print(f"명렬 인식: 0명. 파일을 읽지 못했습니다 — {읽기오류}")
        else:
            print("명렬 인식: 0명. 학번·이름 두 열이 있는 표(엑셀) 또는 '학번 이름' 형식의 "
                  "텍스트로 다시 붙여넣어 주세요.")
        return 1
    _write_json(roster, args.out)
    print(f"명렬 인식: {n}명 (방식: {roster.get('방식', '?')}). 이름은 출력하지 않습니다.")
    시트합침 = roster.get("시트합침")
    if 시트합침:
        print(f"{len(시트합침)}개 시트를 합쳐 {n}명 인식했습니다.")
    if roster.get("성이름결합"):
        print("성 열과 이름 열을 합쳐 전체 이름으로 사용합니다.")
    duplicate = roster.get("중복", 0)
    if duplicate > 0:
        print(f"경고: 학번 중복으로 {duplicate}개 행을 채택하지 않았습니다. "
              f"원본의 학생 수가 {n}명이 맞는지 확인해 주세요.")
    for reason in roster.get("확인사유", []):
        print(f"확인필요: {reason}")
    skipped = roster.get("건너뜀", 0)
    if skipped > 0:
        msg = f"경고: {skipped}개 행을 건너뛰었습니다(학번 또는 이름이 비어 있음). 명단을 확인해 주세요."
        if roster.get("상태표기건너뜀", 0) > 0:
            msg += " (상태 표기 행 포함)"
        print(msg)
    short = count_short_names(roster)
    if short > 0:
        print(
            f"경고: 1자 이름 {short}건이 명렬에 있습니다. 본문에서 안전하게 가릴 수 없어"
            "(일반명사 충돌) mask/memo 단계에서 저장이 차단됩니다. 명렬의 이름 열을 성이 "
            "포함된 전체 이름으로 고쳐 주세요(성·이름이 분리된 명렬이면 합쳐 주세요)."
        )
    return 0


def _load_submitted_ids(args) -> list[str]:
    if args.submitted:
        return [s.strip() for s in args.submitted.split(",") if s.strip()]
    with open(args.submitted_from, encoding="utf-8-sig") as f:
        return [line.strip() for line in f if line.strip()]


def _cmd_issue(args) -> int:
    roster = _read_json(args.roster)
    submitted_ids = _load_submitted_ids(args)
    existing = None
    if Path(args.out).exists():
        existing = _read_json(args.out)
    mapping = issue_tokens(roster, submitted_ids, existing=existing)
    _write_json(mapping, args.out)
    total = len(roster.get("students", []))
    issued = len(mapping.get("map", {}))
    not_issued = max(total - issued, 0)
    print(f"토큰 발급: 제출자 {issued}명, 미발급(미제출) {not_issued}명.")
    return 0


def _cmd_mask(args) -> int:
    data = _read_json(args.input)
    roster = _read_json(args.roster)
    mapping = _read_json(args.mapping)

    short = count_short_names(roster)
    if short > 0:
        # 이 시점 이후로는 어떤 처리도 진행하지 않는다 — 결정론적 검사이므로
        # 데이터를 열어보지 않고도(값은 출력하지 않고) 바로 판단할 수 있다.
        print(SHORT_NAME_BLOCK_MESSAGE.format(n=short))
        return 1

    items = data.get("items", [])
    missing = sum(
        1 for item in items
        if mapping.get("map", {}).get(str(item.get("학번", ""))) is None
    )
    if missing > 0:
        # 학번 값 자체는 출력하지 않는다(계약 유지) — 건수만으로 원인을 짚어준다.
        # 명렬 인식 단계에서 이름 형식 오탐으로 학생이 누락됐을 가능성을 안내한다.
        print(
            f"가명화 중단: 매핑(토큰)이 없는 학번 {missing}건. "
            "명렬 인식에서 누락되었을 수 있습니다 — roster 출력의 인원수를 확인하세요."
        )
        return 1

    out_items = []
    total_warnings = 0
    for item in items:
        sid = str(item.get("학번", ""))
        token = mapping["map"][sid]
        text, warnings = pseudonymize_text(item.get("본문", ""), roster, mapping, owner_id=sid)
        total_warnings += len(warnings)
        out_items.append({"토큰": token, "본문": text})

    leak_count = 0
    for item in out_items:
        issues = scan_leak(item["본문"], roster, scope="본문")
        leak_count += sum(1 for level, code, _ in issues if level == "FAIL" and code == "ID_LEAK")

    if leak_count > 0:
        print(f"가명화 중단: 학번 유출 {leak_count}건이 발견되어 저장하지 않았습니다.")
        return 1

    _write_json({"items": out_items}, args.out)
    print(f"가명화: {len(out_items)}건 처리, 본문 이름 치환 경고 {total_warnings}건, 학번 유출 0건.")
    # 경고가 0건이어도 "탐지 한계 = 0건 = 안전"으로 오인되면 안 되므로 매번 출력한다.
    print(DETECTION_LIMIT_WARNING)
    return 0


def _sheet_error_message(result):
    """parse_memo_xlsx/parse_score_xlsx 등의 시트 선택 오류 반환값을 안내 문구로
    바꾼다. 시트 관련 오류가 아니면 None(호출부가 기존 처리를 계속하게 한다).

    데이터가 있는 시트가 여럿이어도 학번이 겹치지 않고 라벨이 같으면 호출부가
    이미 자동으로 합쳐 정상 반환했을 것이므로, 여기 도달하는 두 경우는 진짜
    확인이 필요한 경우뿐이다: 학번이 겹쳐 병합이 안전하지 않거나(sheet_ambiguous),
    시트마다 같은 열이 다른 의미를 가리키거나(label_mismatch)."""
    if not isinstance(result, dict):
        return None
    if result.get("invalid_sheet"):
        lines = ["지정한 시트를 찾을 수 없습니다. 유효한 시트명:"]
        lines.extend(f"  {name}" for name in result["sheets"])
        return "\n".join(lines)
    if result.get("sheet_ambiguous"):
        names = result["sheets"]
        lines = [
            f"시트 간 학번이 겹쳐 {len(names)}개 시트를 자동으로 합칠 수 없습니다. "
            "--sheet 로 지정해 주세요."
        ]
        lines.extend(f"  {name}" for name in names)
        lines.append(f"예: --sheet {names[0]}")
        return "\n".join(lines)
    if result.get("label_mismatch"):
        entries = result["sheets"]
        letters = {e["letter"] for e in entries}
        if len(letters) == 1:
            headline = f"시트마다 {next(iter(letters))}열의 의미가 다릅니다. --sheet 로 지정해 주세요."
        else:
            headline = "시트마다 열의 의미가 다릅니다. --sheet 로 지정해 주세요."
        lines = [headline]
        lines.extend(f"  {e['sheet']}  {e['letter']}열 — {e['label']}" for e in entries)
        return "\n".join(lines)
    return None


def _cmd_memo(args) -> int:
    """교사 관찰 메모(파일)를 토큰화한다.

    대화창에 실명·학번을 입력하는 채널을 원천 차단하기 위한 명령이다 — 메모는
    반드시 파일(채점표의 메모 열 또는 별도 텍스트)로 받고, 여기서 토큰화한 뒤에만
    에이전트에게 결과 파일을 넘긴다.
    """
    roster = _read_json(args.roster)
    mapping = _read_json(args.mapping)

    short = count_short_names(roster)
    if short > 0:
        print(SHORT_NAME_BLOCK_MESSAGE.format(n=short))
        return 1

    input_path = Path(args.input)
    suffix = input_path.suffix.lower()
    merged_sheets = None
    if suffix == ".xlsx":
        if _xlsx_is_empty(input_path):
            print("메모 파일이 비어 있습니다.")
            return 1
        result = parse_memo_xlsx(input_path, sheet=args.sheet)
        sheet_error = _sheet_error_message(result)
        if sheet_error is not None:
            print(sheet_error)
            return 1
        if result is None:
            print("메모 열 헤더를 '관찰 메모'로 지정해 주세요")
            return 1
        pairs, merged_sheets = result
    elif suffix in (".txt", ".md"):
        # utf-8-sig로 먼저 읽어 BOM 유무와 무관하게 "완전히 비었는지"를 정확히 판정한다.
        raw = input_path.read_text(encoding="utf-8-sig", errors="replace")
        if not raw.strip():
            print("메모 파일이 비어 있습니다.")
            return 1
        pairs = parse_memo_text(input_path)
    else:
        print("지원하지 않는 입력 형식입니다. xlsx 또는 txt/md 파일을 사용해 주세요.")
        return 1

    if not pairs:
        print(
            "메모를 한 건도 읽지 못했습니다. 텍스트 파일은 '학번: 메모' 형식인지, "
            "엑셀은 메모 열 헤더가 '관찰 메모'인지 확인해 주세요."
        )
        return 1

    out_items = []
    excluded = 0
    total_warnings = 0
    for sid, memo in pairs:
        token = mapping.get("map", {}).get(str(sid))
        if token is None:
            excluded += 1
            continue
        text, warnings = pseudonymize_text(memo, roster, mapping, owner_id=str(sid))
        total_warnings += len(warnings)
        out_items.append({"토큰": token, "메모": text})

    leak_count = 0
    for item in out_items:
        issues = scan_leak(item["메모"], roster, scope="본문")
        leak_count += sum(1 for level, code, _ in issues if level == "FAIL" and code == "ID_LEAK")

    if leak_count > 0:
        print(f"관찰 메모 중단: 학번 유출 {leak_count}건이 발견되어 저장하지 않았습니다.")
        return 1

    _write_json({"items": out_items}, args.out)

    merge_note = f" ({len(merged_sheets)}개 시트 합침)" if merged_sheets else ""

    if not out_items:
        # 후보는 있었으나 전부 매핑에 없는 학번(=미제출자)이라 제외된 경우.
        # 형식 오류로 인한 0건과 구분되도록 경고를 명확히 한다.
        print(f"수집 0건{merge_note} — 읽은 메모 {excluded}건이 모두 매핑에 없는 학번입니다(미제출자 확인 필요).")
        return 0

    print(
        f"관찰 메모: {len(out_items)}건 수집{merge_note}({excluded}건은 매핑 없음으로 제외). "
        f"본문 이름 치환 경고 {total_warnings}건, 학번 유출 0건."
    )
    return 0


def _to_json_number(v: float):
    """정수값이면 int로, 아니면 float 그대로 저장한다(15.0 -> 15, 15.5 -> 15.5)."""
    return int(v) if float(v).is_integer() else v


def _format_score_label(v) -> str:
    """분포 출력용 점수 표기(정수는 소수점 없이)."""
    return str(int(v)) if float(v).is_integer() else str(v)


def _cmd_score(args) -> int:
    """채점표(파일)의 점수를 토큰화한다.

    채점표를 에이전트가 직접 열지 않도록 하기 위한 명령이다 — 학번·이름이 함께
    있는 채점표 원본 대신, 이 CLI가 산출한 토큰↔점수(점수.json)만 에이전트에게
    넘긴다. 톤 등급은 이 산출물의 점수를 교사와 정한 구간표에 적용해 파생한다.

    --column/--grade-column/--sum-columns는 상호 배타다. 다중 헤더 채점표(예:
    문항별 소제목이 병합된 표)에서 점수 열 후보가 여럿이면(예: 소설 비평하기
    점수, 시 비평하기 점수, 총점) 조용히 하나를 고르지 않고 exit 1로 멈춘 뒤
    선택지를 제시한다 — 오선택은 반 전체의 톤 등급을 조용히 틀리게 만든다.

    --paste는 이 fail-closed 로직 자체를 우회하는 별도 입력 경로다(명렬입력.html
    3단계 산출물인 점수원본.json). 실측에서 채점표 37개 중 20개(54.1%)가 점수
    열 후보를 정확히 하나만 내놓았는데 그 하나가 틀린 열이었다 — 후보가 1개면
    모호성이 감지되지 않아 질문 트리거가 원리적으로 생기지 않으므로, 교사가
    화면에서 학번·점수 두 열을 직접 골라 붙여넣은 값을 그대로 신뢰하는 것이
    유일한 해법이다. --paste가 주어지면 xlsx를 열지 않고, 위치 인자(input)도
    필요 없다. --column/--grade-column/--sum-columns/--sheet와는 함께 쓸 수 없다
    (열 해석·시트 병합은 애초에 xlsx 경로에서만 의미가 있는 개념이기 때문이다).
    """
    if args.paste:
        if any([args.column, args.grade_column, args.sum_columns, args.sheet]):
            print(
                "--paste는 --column/--grade-column/--sum-columns/--sheet와 함께 쓸 수 없습니다."
            )
            return 1
        mapping = _read_json(args.mapping)
        return _cmd_score_paste(args, mapping)

    given = [args.column, args.grade_column, args.sum_columns]
    if sum(1 for v in given if v) > 1:
        print("--column, --grade-column, --sum-columns 중 하나만 지정해 주세요.")
        return 1

    if not args.input:
        print("채점표 파일 경로를 지정하거나 --paste 로 붙여넣기 입력을 사용해 주세요.")
        return 1

    mapping = _read_json(args.mapping)

    input_path = Path(args.input)
    if input_path.suffix.lower() != ".xlsx":
        print("지원하지 않는 입력 형식입니다. xlsx 파일을 사용해 주세요.")
        return 1

    if _xlsx_is_empty(input_path):
        print("채점표 파일이 비어 있습니다.")
        return 1

    if args.grade_column:
        return _cmd_score_grade(args, mapping, input_path)
    if args.sum_columns:
        return _cmd_score_sum(args, mapping, input_path)

    result = parse_score_xlsx(input_path, column=args.column, sheet=args.sheet)
    sheet_error = _sheet_error_message(result)
    if sheet_error is not None:
        print(sheet_error)
        return 1
    if result is None:
        print("점수 열 헤더를 '점수'로 지정하거나 --column 으로 알려 주세요.")
        return 1
    if isinstance(result, dict) and result.get("ambiguous"):
        candidates = result["candidates"]
        lines = [f"점수 열 후보가 {len(candidates)}개입니다. --column 으로 지정해 주세요."]
        for c in candidates:
            lines.append(f"  {c['letter']}열  — {c['label']}")
        lines.append("예: --column R")
        print("\n".join(lines))
        return 1

    pairs, no_score, letter, label, merged_sheets = result
    column_ref = f"{letter} — {label}"
    if merged_sheets:
        column_ref += f", {len(merged_sheets)}개 시트 합침"

    if not pairs and no_score == 0:
        print("점수를 한 건도 읽지 못했습니다. 학번 열과 점수 열이 있는지 확인해 주세요.")
        return 1

    out_items = []
    excluded_unmapped = 0
    for sid, score in pairs:
        token = mapping.get("map", {}).get(str(sid))
        if token is None:
            excluded_unmapped += 1
            continue
        out_items.append({"토큰": token, "점수": _to_json_number(score)})

    _write_json({"items": out_items}, args.out)

    if not out_items:
        print(
            f"점수 수집: 0명(열: {column_ref}). "
            f"제외: 매핑 없음 {excluded_unmapped}건, 점수 없음 {no_score}건."
        )
        return 0

    dist: dict = {}
    for item in out_items:
        dist[item["점수"]] = dist.get(item["점수"], 0) + 1
    dist_str = ", ".join(
        f"{_format_score_label(score)}점 {count}명"
        for score, count in sorted(dist.items(), key=lambda kv: -kv[0])
    )

    print(
        f"점수 수집: {len(out_items)}명(열: {column_ref}). 분포 — {dist_str}. "
        f"제외: 매핑 없음 {excluded_unmapped}건, 점수 없음 {no_score}건."
    )
    return 0


def _cmd_score_paste(args, mapping) -> int:
    """--paste: 붙여넣기로 만든 점수원본.json에서 점수/등급을 그대로 받는다.

    xlsx를 열지 않으므로 열 해석·시트 병합 개념이 없다 — 교사가 이미 화면에서
    학번·점수(또는 등급) 열을 직접 골랐으므로 그 값을 그대로 신뢰한다.
    """
    try:
        obj = _read_json(args.paste)
    except (OSError, ValueError) as exc:
        print(f"붙여넣기 입력 파일을 읽을 수 없습니다: {args.paste} ({type(exc).__name__}).")
        return 1

    result = parse_score_paste(obj)
    if result["status"] == "error":
        print(f"점수 붙여넣기 거부: {result['message']}")
        return 1

    mode = result["mode"]
    pairs = result["pairs"]
    out_items = []
    excluded_unmapped = 0
    for sid, value in pairs:
        token = mapping.get("map", {}).get(str(sid))
        if token is None:
            excluded_unmapped += 1
            continue
        if mode == "score":
            out_items.append({"토큰": token, "점수": _to_json_number(value)})
        else:
            out_items.append({"토큰": token, "등급": value})

    _write_json({"items": out_items}, args.out)

    label = "점수" if mode == "score" else "등급"

    if not out_items:
        print(f"{label} 수집: 0명(붙여넣기). 제외: 매핑 없음 {excluded_unmapped}건.")
        return 0

    dist: dict = {}
    for item in out_items:
        key = item.get("점수", item.get("등급"))
        dist[key] = dist.get(key, 0) + 1

    if mode == "score":
        dist_str = ", ".join(
            f"{_format_score_label(score)}점 {count}명"
            for score, count in sorted(dist.items(), key=lambda kv: -kv[0])
        )
    else:
        dist_str = ", ".join(f"{g} {c}명" for g, c in sorted(dist.items(), key=lambda kv: -kv[1]))

    print(
        f"{label} 수집: {len(out_items)}명(붙여넣기). 분포 — {dist_str}. "
        f"제외: 매핑 없음 {excluded_unmapped}건."
    )
    return 0


def _cmd_score_grade(args, mapping, input_path) -> int:
    """--grade-column: 점수 열이 없는 채점표에서 등급 문자열을 그대로 담는다."""
    result = parse_grade_xlsx(input_path, args.grade_column, sheet=args.sheet)
    sheet_error = _sheet_error_message(result)
    if sheet_error is not None:
        print(sheet_error)
        return 1
    if result is None:
        print(f"'{args.grade_column}' 열을 찾을 수 없습니다. 열 문자나 헤더 이름을 확인해 주세요.")
        return 1
    pairs, no_grade, merged_sheets = result
    grade_column_ref = args.grade_column
    if merged_sheets:
        grade_column_ref += f", {len(merged_sheets)}개 시트 합침"

    if not pairs and no_grade == 0:
        print("등급을 한 건도 읽지 못했습니다. 학번 열과 등급 열이 있는지 확인해 주세요.")
        return 1

    out_items = []
    excluded_unmapped = 0
    for sid, grade in pairs:
        token = mapping.get("map", {}).get(str(sid))
        if token is None:
            excluded_unmapped += 1
            continue
        out_items.append({"토큰": token, "등급": grade})

    _write_json({"items": out_items}, args.out)

    if not out_items:
        print(
            f"등급 수집: 0명(열: {grade_column_ref}). "
            f"제외: 매핑 없음 {excluded_unmapped}건, 등급 없음 {no_grade}건."
        )
        return 0

    dist: dict = {}
    for item in out_items:
        dist[item["등급"]] = dist.get(item["등급"], 0) + 1
    dist_str = ", ".join(f"{g} {c}명" for g, c in sorted(dist.items(), key=lambda kv: -kv[1]))

    print(
        f"등급 수집: {len(out_items)}명(열: {grade_column_ref}). 분포 — {dist_str}. "
        f"제외: 매핑 없음 {excluded_unmapped}건, 등급 없음 {no_grade}건."
    )
    return 0


def _cmd_score_sum(args, mapping, input_path) -> int:
    """--sum-columns: 점수 열이 없는 채점표에서 지정 범위 열을 합산해 점수로 쓴다."""
    result = parse_sum_columns_xlsx(input_path, args.sum_columns, sheet=args.sheet)
    sheet_error = _sheet_error_message(result)
    if sheet_error is not None:
        print(sheet_error)
        return 1
    if result is None:
        print("합산할 열 범위를 확인해 주세요(예: --sum-columns E:J).")
        return 1
    pairs, no_score, merged_sheets = result

    if not pairs and no_score == 0:
        print("점수를 한 건도 읽지 못했습니다. 학번 열과 합산 범위를 확인해 주세요.")
        return 1

    out_items = []
    excluded_unmapped = 0
    for sid, score in pairs:
        token = mapping.get("map", {}).get(str(sid))
        if token is None:
            excluded_unmapped += 1
            continue
        out_items.append({"토큰": token, "점수": _to_json_number(score)})

    _write_json({"items": out_items}, args.out)

    range_display = args.sum_columns.strip().upper().replace(" ", "").replace(":", "~") + "열"
    if merged_sheets:
        range_display += f", {len(merged_sheets)}개 시트 합침"

    if not out_items:
        print(
            f"합산: {range_display}. 점수 수집: 0명. "
            f"제외: 매핑 없음 {excluded_unmapped}건, 점수 없음 {no_score}건."
        )
        return 0

    dist: dict = {}
    for item in out_items:
        dist[item["점수"]] = dist.get(item["점수"], 0) + 1
    dist_str = ", ".join(
        f"{_format_score_label(score)}점 {count}명"
        for score, count in sorted(dist.items(), key=lambda kv: -kv[0])
    )

    print(
        f"합산: {range_display}. 점수 수집: {len(out_items)}명. 분포 — {dist_str}. "
        f"제외: 매핑 없음 {excluded_unmapped}건, 점수 없음 {no_score}건."
    )
    return 0


def _cmd_destroy(args) -> int:
    """가명처리 중간 산출물을 파기한다.

    되돌릴 수 없는 동작이므로 기본은 목록만 보여주고, 실제 삭제는 --yes를
    요구한다. 파일명은 출력하되 내용(실명·학번)은 열지도 출력하지도 않는다.
    """
    found = detect_stale_artifacts(args.dir)
    if not found:
        print("파기할 중간 산출물이 없습니다.")
        return 0

    names = ", ".join(p.name for p in found)
    if not args.yes:
        print(f"파기 대상 {len(found)}건: {names}")
        print("실제로 지우려면 --yes 를 붙여 다시 실행해 주세요(되돌릴 수 없습니다).")
        return 0

    destroyed = destroy_artifacts(args.dir)
    remaining = detect_stale_artifacts(args.dir)
    print(f"파기 완료: {len(destroyed)}건 — {', '.join(destroyed)}")
    if remaining:
        print(f"경고: {len(remaining)}건이 남아 있습니다 — {', '.join(p.name for p in remaining)}")
        return 1
    return 0


def _cmd_finalize(args) -> int:
    draft = _read_json(args.input)
    roster = _read_json(args.roster)
    mapping = _read_json(args.mapping)

    token_to_sid = {t: s for s, t in mapping.get("map", {}).items()}
    sid_to_name = {str(s.get("학번", "")): s.get("이름", "") for s in roster.get("students", [])}

    new_classes = []
    restored = 0
    for cls in draft.get("classes", []):
        new_students = []
        for student in cls.get("students", []):
            token = student.get("토큰")
            sid = token_to_sid.get(token)
            if sid is None:
                print("재결합 실패: 매핑에 없는 토큰이 있어 1:1 복원을 완료할 수 없습니다.")
                return 1
            new_student = {"학번": sid, "이름": sid_to_name.get(sid, "")}
            for key, value in student.items():
                if key == "토큰":
                    continue
                if isinstance(value, str):
                    value = reidentify(value, mapping)
                new_student[key] = value
            new_students.append(new_student)
            restored += 1
        new_classes.append({"name": cls.get("name", ""), "students": new_students})

    _write_json({"classes": new_classes}, args.out)
    print(f"재결합: {restored}명 복원 완료.")
    return 0


# ---------------------------------------------------------------------------
# review-bundle — 회고(SKILL.md §9-2)가 요구하는 "초안과 검수용을 학번 키로
# 매칭해 diff를 만든다"는 지금까지 CLI가 없어 에이전트가 실명 두 파일을 직접
# 열어야 했다. 검수 자체를 브라우저로 옮기고 에이전트에게는 토큰 기준 diff만
# 주기 위한 첫 조각이 이 번들이다 — 산출 파일은 실명을 담지만(로컬 전용),
# 이 CLI의 stdout에는 이름·학번이 절대 등장하지 않는다.
# ---------------------------------------------------------------------------

_BIGO_REVIEW_KEYWORDS = ("검토 필요", "추정", "교정", "확인")


def build_review_bundle(drafts: dict, mapping: dict, sources: dict | None = None,
                         profile: dict | None = None, rules: dict | None = None) -> dict:
    """실명 초안에 토큰을 붙여 브라우저 검수용 번들을 만든다.

    드라프트의 각 학생에 매핑표의 토큰을 붙인다(브라우저가 나중에 토큰 기준
    diff를 만들 수 있게). 매핑에 토큰이 없는 학번(미제출자 등)도 제외하지
    않고 토큰을 None으로 두며, 조용히 검수 대상에서 빠지지 않도록 우선검토
    사유를 붙인다.

    sources가 주어지면(토큰본.json) 토큰으로 원문을 찾아 함께 싣는다. 원문
    속 토큰은 reidentify로 학번으로 되돌린다 — 이 산출 파일 자체가 로컬 전용
    이므로 학번이 보여도 안전하고, 검수자가 문맥을 읽기에도 그 편이 낫다.

    profile은 활동명·목표바이트·상한바이트를 담는다(없으면 700/760 기본값).
    rules는 wiki/규칙.json에서 읽은 금지어휘·금지문자패턴이다(없으면
    verify_seteuk.load_rules()로 직접 읽는다) — 브라우저가 같은 규칙으로
    금지어·금지문자를 표시할 수 있게 그대로 싣는다.

    우선검토(판단이 필요해 검수 순서를 앞당길 사유)에는 최소한 다음을 담는다:
    예외 학생, 비고에 '검토 필요'/'추정'/'교정'/'확인' 포함, 바이트가 상한의
    95% 이상이거나 목표의 60% 이하, 매핑에 토큰이 없음. 해당 없으면 빈 목록.
    """
    profile = profile or {}
    target = int(profile.get("목표바이트", 700))
    limit = int(profile.get("상한바이트", 760))
    activity = profile.get("활동명") or mapping.get("활동") or ""

    if rules is None:
        from verify_seteuk import load_rules
        rules = load_rules()

    금지어 = list(rules.get("금지어휘", []))
    if rules.get("이는_경계검사", True) and "이는" not in 금지어:
        금지어 = 금지어 + ["이는"]
    # 정규식 패턴을 문자 목록처럼 실으면 소비자가 부분일치로 검사하다 아무것도
    # 잡지 못한다(패턴 문자열 자체가 본문에 나올 일이 없다). 패턴은 패턴으로 싣고,
    # 소비자가 정규식으로 쓰게 한다.
    금지문자패턴 = rules.get("금지문자패턴", "")

    token_map = mapping.get("map", {})
    source_by_token = {}
    if sources is not None:
        for item in sources.get("items", []):
            tok = item.get("토큰")
            if tok:
                source_by_token[tok] = item.get("본문", "")

    students_out = []
    for cls in drafts.get("classes", []):
        반 = cls.get("name", "")
        for s in cls.get("students", []):
            sid = str(s.get("학번", ""))
            token = token_map.get(sid)
            entry = {
                "토큰": token,
                "학번": sid,
                "이름": s.get("이름", ""),
                "반": 반,
                "톤등급": s.get("톤등급", ""),
                "핵심소재": s.get("핵심소재", ""),
                "세특": s.get("세특", ""),
                "비고": s.get("비고", ""),
                "예외": bool(s.get("예외", False)),
            }

            reasons: list[str] = []
            if entry["예외"]:
                reasons.append("예외 학생")
            비고 = entry["비고"] or ""
            for kw in _BIGO_REVIEW_KEYWORDS:
                if kw in 비고:
                    reasons.append(f"비고에 '{kw}' 포함")

            세특 = entry["세특"] or ""
            nbytes = len(세특.encode("utf-8"))
            if nbytes >= limit * 0.95:
                reasons.append("분량 상한 근접")
            elif nbytes <= target * 0.6:
                reasons.append("분량 목표 대비 크게 미달")

            if token is None:
                reasons.append("매핑에 토큰 없음")

            if sources is not None:
                본문 = source_by_token.get(token, "") if token else ""
                if 본문:
                    entry["원문"] = reidentify(본문, mapping)

            entry["우선검토"] = reasons
            students_out.append(entry)

    return {
        "활동명": activity,
        "목표바이트": target,
        "상한바이트": limit,
        "금지어": 금지어,
        "금지문자패턴": 금지문자패턴,
        "students": students_out,
    }


def _cmd_review_bundle(args) -> int:
    try:
        drafts = _read_json(args.drafts)
    except (OSError, ValueError):
        print("검수번들 생성 실패: 초안 파일을 읽지 못했습니다.")
        return 1

    total = sum(len(cls.get("students", [])) for cls in drafts.get("classes", []))
    if total == 0:
        print("검수번들 생성 실패: 초안에 학생이 없습니다.")
        return 1

    mapping = _read_json(args.mapping)
    sources = _read_json(args.sources) if args.sources else None
    profile = _read_json(args.profile) if args.profile else None

    bundle = build_review_bundle(drafts, mapping, sources=sources, profile=profile)
    _write_json(bundle, args.out)

    n = len(bundle["students"])
    priority = sum(1 for s in bundle["students"] if s.get("우선검토"))
    with_source = sum(1 for s in bundle["students"] if s.get("원문"))
    print(f"검수번들: {n}명(우선 검토 {priority}명, 원문 포함 {with_source}명). 저장: {args.out}")
    print("주의: 이 파일은 실명을 담은 민감 산출물입니다 — 로컬 전용, 검수 후 destroy로 파기하세요.")
    return 0


def main(argv=None) -> int:
    import argparse
    import sys

    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")

    parser = argparse.ArgumentParser(
        description="가명처리 CLI — 명렬(실명·학번)을 로컬에서만 다루고, "
                    "에이전트는 stdout 요약(인원수·건수)만 읽는다."
    )
    sub = parser.add_subparsers(dest="command", required=True)

    p_roster = sub.add_parser("roster", help="명렬 감지")
    p_roster.add_argument("input", help="명단 원본 파일(xlsx 또는 텍스트)")
    p_roster.add_argument("--out", required=True, help="명렬.json 저장 경로")
    p_roster.set_defaults(func=_cmd_roster)

    p_issue = sub.add_parser("issue", help="제출자 토큰 발급")
    p_issue.add_argument("--roster", required=True, help="명렬.json 경로")
    group = p_issue.add_mutually_exclusive_group(required=True)
    group.add_argument("--submitted", help="쉼표로 구분한 제출자 학번 목록")
    group.add_argument("--submitted-from", help="제출자 학번이 한 줄에 하나씩 있는 텍스트 파일")
    p_issue.add_argument("--out", required=True, help="매핑.json 저장 경로(기존 파일이 있으면 재사용·추가 발급)")
    p_issue.set_defaults(func=_cmd_issue)

    p_mask = sub.add_parser("mask", help="본문·학번 가명화")
    p_mask.add_argument("input", help='입력 JSON({"items":[{"학번","본문"}]})')
    p_mask.add_argument("--roster", required=True, help="명렬.json 경로")
    p_mask.add_argument("--mapping", required=True, help="매핑.json 경로")
    p_mask.add_argument("--out", required=True, help="토큰본.json 저장 경로")
    p_mask.set_defaults(func=_cmd_mask)

    p_memo = sub.add_parser("memo", help="교사 관찰 메모 토큰화(파일 입력 전용)")
    p_memo.add_argument("input", help="관찰 메모 원본 파일(xlsx 또는 txt/md)")
    p_memo.add_argument("--roster", required=True, help="명렬.json 경로")
    p_memo.add_argument("--mapping", required=True, help="매핑.json 경로")
    p_memo.add_argument("--out", required=True, help="관찰메모.json 저장 경로")
    p_memo.add_argument(
        "--sheet",
        help="처리할 시트명을 지정한다(xlsx 입력 전용). 지정하지 않으면 메모 데이터가 있는 "
             "시트가 1개일 때 자동 진행하고, 2개 이상이면 학번이 겹치지 않고 헤더가 같을 때 "
             "자동으로 합쳐 진행한다. 학번이 겹치거나 헤더가 다르면 exit 1로 멈춘 뒤 안내한다",
    )
    p_memo.set_defaults(func=_cmd_memo)

    p_score = sub.add_parser("score", help="채점표 점수 토큰화(파일 입력 전용)")
    p_score.add_argument(
        "input", nargs="?",
        help="채점표 원본 파일(xlsx). --paste 사용 시에는 생략한다",
    )
    p_score.add_argument("--roster", required=True, help="명렬.json 경로")
    p_score.add_argument("--mapping", required=True, help="매핑.json 경로")
    p_score.add_argument("--out", required=True, help="점수.json 저장 경로")
    p_score.add_argument(
        "--paste",
        help="붙여넣기로 만든 점수원본.json 경로(명렬입력.html 3단계 산출물). 지정하면 "
             "xlsx를 열지 않으며 위치 인자(input)도 필요 없다. 교사가 화면에서 학번·점수 "
             "(또는 학번·등급) 두 열을 직접 골라 붙여넣은 값을 그대로 신뢰한다 — 점수 열 "
             "후보가 하나뿐이라 모호성이 감지되지 않는 오염(실측 54.1%%)을 원천 차단한다. "
             "--column/--grade-column/--sum-columns/--sheet와 함께 쓸 수 없다",
    )
    p_score.add_argument(
        "--column",
        help="점수 열(엑셀 열 문자 J/R/AA 또는 헤더 이름, 기본: 점수/총점/합계/평가점수 자동 탐지, "
             "공백 무시). 후보가 여럿이면 이 옵션 없이는 exit 1로 멈추고 선택지를 안내한다",
    )
    p_score.add_argument(
        "--grade-column",
        help="점수 열이 없는 채점표용 — 지정 열(문자 또는 헤더 이름)의 값을 등급 문자열 그대로 담는다",
    )
    p_score.add_argument(
        "--sum-columns",
        help="점수 열이 없는 채점표용 — 지정 범위(예: E:J)의 숫자를 합산해 점수로 쓴다",
    )
    p_score.add_argument(
        "--sheet",
        help="처리할 시트명을 지정한다. 지정하지 않으면 점수(또는 등급·합산) 데이터가 있는 "
             "시트가 1개일 때 자동 진행하고, 2개 이상이면 학번이 겹치지 않고 열의 의미(라벨)가 "
             "같을 때 자동으로 합쳐 진행한다(예: 반별로 나뉜 채점표). 학번이 겹치거나 시트마다 "
             "열의 의미가 다르면 exit 1로 멈춘 뒤 안내한다",
    )
    p_score.set_defaults(func=_cmd_score)

    p_finalize = sub.add_parser("finalize", help="토큰 → 실명 재결합")
    p_finalize.add_argument("input", help='토큰 초안 JSON({"classes":[{"name","students":[{"토큰",...}]}]})')
    p_finalize.add_argument("--roster", required=True, help="명렬.json 경로")
    p_finalize.add_argument("--mapping", required=True, help="매핑.json 경로")
    p_finalize.add_argument("--out", required=True, help="실명초안.json 저장 경로")
    p_finalize.set_defaults(func=_cmd_finalize)

    p_review = sub.add_parser("review-bundle", help="브라우저 검수용 번들 생성(실명 산출물)")
    p_review.add_argument("--drafts", required=True,
                          help='실명 초안 JSON({"classes":[{"name","students":[{"학번","이름",...}]}]})')
    p_review.add_argument("--mapping", required=True, help="매핑.json 경로(토큰을 붙이는 데 필요)")
    p_review.add_argument("--sources",
                          help='토큰본.json 경로(선택, {"items":[{"토큰","본문"}]}). 있으면 학생별 '
                               "원문을 함께 싣는다(원문 속 토큰은 학번으로 되돌려 싣는다)")
    p_review.add_argument("--profile",
                          help="활동프로파일.json 경로(선택). 없으면 기본값(700/760)을 쓴다")
    p_review.add_argument("--out", required=True, help="검수번들.json 저장 경로(실명 포함 — 로컬 전용)")
    p_review.set_defaults(func=_cmd_review_bundle)

    p_destroy = sub.add_parser("destroy", help="가명처리 중간 산출물 파기")
    p_destroy.add_argument("--dir", default=".", help="검사할 폴더(기본: 현재 폴더)")
    p_destroy.add_argument("--yes", action="store_true",
                           help="실제로 삭제한다(없으면 목록만 보여준다)")
    p_destroy.set_defaults(func=_cmd_destroy)

    args = parser.parse_args(argv)
    return args.func(args)


if __name__ == "__main__":
    raise SystemExit(main())
