# -*- coding: utf-8 -*-
import json
import subprocess
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from verify_seteuk import check_text, find_name_intrusions, save_xlsx, verify_drafts

PROFILE = {"활동명": "가상 활동", "문두": "가상 활동에서", "목표바이트": 700, "상한바이트": 760,
           "평가자료": "가상 채점표(테스트)"}


def codes(issues):
    return [(lv, code) for lv, code, _ in issues]


def test_banned_word_detected():
    text = "가상 활동에서 작품을 읽고 또한 감상을 정리함."
    _, issues = check_text(text, PROFILE, exempt=True)
    assert ("FAIL", "BANNED_WORD") in codes(issues)


def test_ineun_word_boundary_no_false_positive():
    ok = "가상 활동에서 완벽해 보이는 존재를 분석함."
    _, issues = check_text(ok, PROFILE, exempt=True)
    assert ("FAIL", "BANNED_WORD") not in codes(issues)


def test_ineun_detected_as_word():
    bad = "가상 활동에서 분석함. 이는 큰 문제임을 밝힘."
    _, issues = check_text(bad, PROFILE, exempt=True)
    assert ("FAIL", "BANNED_WORD") in codes(issues)


def test_banned_char_hyphen():
    text = "가상 활동에서 두-세 가지 관점을 정리함."
    _, issues = check_text(text, PROFILE, exempt=True)
    assert ("FAIL", "BANNED_CHAR") in codes(issues)


def test_banned_char_straight_double_quote():
    text = '가상 활동에서 "인용문"을 정리함.'
    _, issues = check_text(text, PROFILE, exempt=True)
    assert ("FAIL", "BANNED_CHAR") in codes(issues)


def test_single_quote_allowed_for_book_title():
    text = "가상 활동에서 '가상의 책(작가)'을 읽고 감상을 정리함."
    _, issues = check_text(text, PROFILE, exempt=True)
    assert ("FAIL", "BANNED_CHAR") not in codes(issues)


def test_byte_over_fails():
    text = "가상 활동에서 " + "분석함. " * 80  # 760바이트 초과 유도
    _, issues = check_text(text, PROFILE)
    assert ("FAIL", "BYTE_OVER") in codes(issues)


def test_byte_under_warns_not_fails():
    text = "가상 활동에서 작품을 읽고 감상을 정리함."
    _, issues = check_text(text, PROFILE)
    assert ("WARN", "BYTE_UNDER") in codes(issues)
    assert not any(lv == "FAIL" and c == "BYTE_UNDER" for lv, c in codes(issues))


def test_opening_prefix_enforced():
    text = "이 학생은 작품을 열심히 분석함." + " 서술을 이어감." * 40
    _, issues = check_text(text, PROFILE)
    assert ("FAIL", "OPENING") in codes(issues)


def test_exempt_skips_opening_and_under():
    text = "가상 활동에 참여함."
    _, issues = check_text(text, PROFILE, exempt=True)
    assert ("FAIL", "OPENING") not in codes(issues)
    assert ("WARN", "BYTE_UNDER") not in codes(issues)


def test_name_intrusion_detected():
    assert find_name_intrusions("김가상 학생이 발표함.", ["김가상", "이허구"]) == ["김가상"]
    assert find_name_intrusions("작품을 분석함.", ["김가상", "이허구"]) == []


GOOD_TEXT = (
    "가상 활동에서 '가상의 책(작가)'을 선정하여 인물의 갈등에 주목하며 감상문을 작성함. "
    "서술 시점의 효과를 짚고 인물의 내적 갈등이 심화되는 과정을 정리함. "
    "작품에 반영된 사회 현실을 비판적으로 읽어냄. "
    "자신의 경험과 견주어 삶의 태도를 성찰하는 다짐을 밝힘. "
    "감상의 근거를 본문에서 찾아 제시하는 태도가 돋보임. "
    "작품을 자기 이해의 계기로 삼는 모습을 보임."
)


def make_drafts(text=GOOD_TEXT, exempt=False):
    return {
        "classes": [
            {
                "name": "1반",
                "students": [
                    {
                        "학번": "10101",
                        "이름": "김가상",
                        "핵심소재": "가상의 책(작가)",
                        "톤등급": "중",
                        "세특": text,
                        "비고": "",
                        "예외": exempt,
                    }
                ],
            }
        ]
    }


def test_verify_drafts_pass():
    report = verify_drafts(make_drafts(), PROFILE)
    assert report["fail"] == 0


def test_verify_drafts_counts_fail():
    report = verify_drafts(make_drafts(text=GOOD_TEXT + " 또한 정리함."), PROFILE)
    assert report["fail"] >= 1


def test_save_xlsx_writes_file(tmp_path):
    drafts = make_drafts()
    report = verify_drafts(drafts, PROFILE)
    out = tmp_path / "out.xlsx"
    save_xlsx(drafts, report, str(out))
    assert out.exists()


def test_saved_xlsx_byte_column_is_live_formula(tmp_path):
    from openpyxl import load_workbook

    drafts = make_drafts()
    report = verify_drafts(drafts, PROFILE)
    out = tmp_path / "out.xlsx"
    save_xlsx(drafts, report, str(out))
    ws = load_workbook(str(out))["1반"]
    value = ws["F2"].value
    assert isinstance(value, str) and value.startswith("=")
    assert "LENB(E2)" in value and "LEN(E2)" in value


def test_cli_blocks_save_on_fail(tmp_path):
    drafts_path = tmp_path / "d.json"
    profile_path = tmp_path / "p.json"
    out = tmp_path / "out.xlsx"
    drafts_path.write_text(
        json.dumps(make_drafts(text=GOOD_TEXT + " 또한 정리함."), ensure_ascii=False),
        encoding="utf-8",
    )
    profile_path.write_text(json.dumps(PROFILE, ensure_ascii=False), encoding="utf-8")
    script = str(Path(__file__).resolve().parents[1] / "verify_seteuk.py")
    proc = subprocess.run(
        [sys.executable, script, str(drafts_path), "--profile", str(profile_path), "--save", str(out)],
        capture_output=True,
        text=True,
        encoding="utf-8",
    )
    assert proc.returncode == 1
    assert not out.exists()
    assert "FAIL" in proc.stdout
    assert "저장 차단" in proc.stdout


def test_cli_rejects_empty_drafts(tmp_path):
    drafts_path = tmp_path / "d.json"
    profile_path = tmp_path / "p.json"
    out = tmp_path / "out.xlsx"
    drafts_path.write_text(json.dumps({"classes": []}, ensure_ascii=False), encoding="utf-8")
    profile_path.write_text(json.dumps(PROFILE, ensure_ascii=False), encoding="utf-8")
    script = str(Path(__file__).resolve().parents[1] / "verify_seteuk.py")
    proc = subprocess.run(
        [sys.executable, script, str(drafts_path), "--profile", str(profile_path), "--save", str(out)],
        capture_output=True, text=True, encoding="utf-8",
    )
    assert proc.returncode == 1
    assert not out.exists()


def test_cli_success_path_saves(tmp_path):
    drafts_path = tmp_path / "d.json"
    profile_path = tmp_path / "p.json"
    out = tmp_path / "out.xlsx"
    drafts_path.write_text(json.dumps(make_drafts(), ensure_ascii=False), encoding="utf-8")
    profile_path.write_text(json.dumps(PROFILE, ensure_ascii=False), encoding="utf-8")
    script = str(Path(__file__).resolve().parents[1] / "verify_seteuk.py")
    proc = subprocess.run(
        [sys.executable, script, str(drafts_path), "--profile", str(profile_path), "--save", str(out)],
        capture_output=True, text=True, encoding="utf-8",
    )
    assert proc.returncode == 0
    assert out.exists()
    assert "저장 완료" in proc.stdout
