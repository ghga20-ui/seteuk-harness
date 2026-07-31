# -*- coding: utf-8 -*-
"""토큰 "값" 신뢰 회귀 테스트(등장인물·학번 전부 가상 — 김가상/30101 등).

계약: 이 스크립트의 stdout은 에이전트(LLM)가 읽는다. detect_stage는 "토큰"
키의 존재만 보고 값은 보지 않으므로, 예전에는 빈 토큰이 FAIL 0으로 통과하고
학번 형태 값("30101")이 행별 출력에 그대로 찍힐 수 있었다 — 검증기 자체가
유출 경로가 되는 구멍이다. 여기서 못 박는 계약:

- 형식 위반 토큰 값은 차단 메시지에도 절대 되뱉지 않는다(TOKEN_INVALID).
  위반 학생의 식별은 위치(반 이름 + 반 내 순번)로만 한다.
- 토큰 단계 초안에 실명 키(이름/학번)가 섞이면 값 없이 차단한다(PRIVACY_FIELD).
- --mapping의 값도 무검증이면 오염 매핑 값이 실명 단계 행별 출력에 나간다 —
  형식 위반 값은 개수만 세어 FAIL(MAPPING_INVALID), 중복 값은 MAPPING_DUP.
- --save xlsx는 생성 순간부터 0o600(POSIX 전용 assert — Windows는 ACL 몫).
"""
import json
import os
import re
import stat
import sys
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
import verify_seteuk as V

PROFILE = {"활동명": "가상 활동", "문두": "가상 활동에서", "목표바이트": 700, "상한바이트": 760,
           "평가자료": "가상 채점표(테스트)"}

GOOD_TEXT = (
    "가상 활동에서 '가상의 책(작가)'을 선정하여 인물의 갈등에 주목하며 감상문을 작성함. "
    "서술 시점의 효과를 짚고 인물의 내적 갈등이 심화되는 과정을 정리함. "
    "작품에 반영된 사회 현실을 비판적으로 읽어냄. "
    "자신의 경험과 견주어 삶의 태도를 성찰하는 다짐을 밝힘. "
    "감상의 근거를 본문에서 찾아 제시하는 태도가 돋보임. "
    "작품을 자기 이해의 계기로 삼는 모습을 보임."
)

# 4자리 이상 숫자 = 학번 후보(test_stage_and_stdout.py와 같은 그물). 이 파일의
# 시나리오에서 바이트 수(3자리)·건수(1~2자리)는 4자리에 닿지 않는다.
SID_LIKE = re.compile(r"\d{4,}")

posix_only = pytest.mark.skipif(
    os.name != "posix", reason="파일 mode 비트는 POSIX에서만 의미가 있다")


def token_student(token, text=GOOD_TEXT, exempt=False):
    return {"토큰": token, "핵심소재": "가상의 책(작가)", "톤등급": "중",
            "세특": text, "비고": "", "예외": exempt}


def named_student(sid, name, text=GOOD_TEXT, exempt=False):
    return {"학번": sid, "이름": name, "핵심소재": "가상의 책(작가)", "톤등급": "중",
            "세특": text, "비고": "", "예외": exempt}


def wrap(students):
    return {"classes": [{"name": "1반", "students": students}]}


def write_json(path, data):
    path.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")


def run_main(tmp_path, capsys, drafts, roster=None, mapping=None,
             save=False, expected_count=None):
    """main()을 같은 프로세스에서 실행하고 (exit코드, stdout)을 돌려준다."""
    write_json(tmp_path / "d.json", drafts)
    write_json(tmp_path / "p.json", PROFILE)
    argv = [str(tmp_path / "d.json"), "--profile", str(tmp_path / "p.json")]
    if roster is not None:
        write_json(tmp_path / "r.json", roster)
        argv += ["--roster", str(tmp_path / "r.json")]
    if mapping is not None:
        write_json(tmp_path / "m.json", mapping)
        argv += ["--mapping", str(tmp_path / "m.json")]
    if save:
        argv += ["--save", str(tmp_path / "out.xlsx")]
    if expected_count is not None:
        argv += ["--expected-count", str(expected_count)]
    rc = V.main(argv)
    return rc, capsys.readouterr().out


# ---------------------------------------------------------------------------
# ① 토큰 단계 선검증 — TOKEN_INVALID
# ---------------------------------------------------------------------------

def test_empty_token_fails_token_invalid(tmp_path, capsys):
    """빈 토큰은 더 이상 FAIL 0으로 통과하지 못한다 — 위치로만 식별해 차단."""
    rc, out = run_main(tmp_path, capsys, wrap([token_student("")]))
    assert rc == 1
    assert "FAIL 1반 1번째 학생 [TOKEN_INVALID] 토큰 형식 위반(빈 값 또는 S-XXXX 아님)" in out
    assert "가명화 산출물에서 초안을 다시 만드세요" in out


def test_sid_like_token_value_never_printed(tmp_path, capsys):
    """학번 형태 토큰("30101")은 TOKEN_INVALID로 막히고 값이 stdout 어디에도 없다."""
    drafts = wrap([token_student("S-AB12"), token_student("S-CD34"),
                   token_student("30101")])
    rc, out = run_main(tmp_path, capsys, drafts)
    assert rc == 1
    assert "FAIL 1반 3번째 학생 [TOKEN_INVALID]" in out
    assert "30101" not in out
    assert not SID_LIKE.search(out), out


def test_invalid_token_student_other_fails_use_position(tmp_path, capsys):
    """TOKEN_INVALID 학생의 다른 FAIL(금지어 등) 행별 출력에도 토큰 값이 안 나간다."""
    drafts = wrap([token_student("30101", text=GOOD_TEXT + " 또한 정리함.")])
    rc, out = run_main(tmp_path, capsys, drafts)
    assert rc == 1
    assert "TOKEN_INVALID" in out
    assert "FAIL 1반 1번째 학생 [BANNED_WORD]" in out  # 본 검사도 위치로 찍는다
    assert "30101" not in out


def test_dup_token_only_among_valid_tokens(tmp_path, capsys):
    """형식 위반 값끼리 겹쳐도 DUP_TOKEN이 아니라 TOKEN_INVALID로만 잡는다."""
    drafts = wrap([token_student("30101"), token_student("30101")])
    rc, out = run_main(tmp_path, capsys, drafts)
    assert rc == 1
    assert "DUP_TOKEN" not in out
    assert out.count("TOKEN_INVALID") == 2
    assert "30101" not in out


def test_valid_dup_token_still_caught(tmp_path, capsys):
    """유효 토큰의 중복 감지는 종전대로 살아 있어야 한다(회귀 방지)."""
    drafts = wrap([token_student("S-AB12"), token_student("S-AB12")])
    rc, out = run_main(tmp_path, capsys, drafts)
    assert rc == 1
    assert "DUP_TOKEN" in out


# ---------------------------------------------------------------------------
# ② 토큰 단계 선검증 — PRIVACY_FIELD(실명 키 혼입)
# ---------------------------------------------------------------------------

def test_token_student_with_name_key_fails_privacy_field(tmp_path, capsys):
    """토큰 학생에게 "이름" 키가 있으면 차단하되 이름 값은 stdout에 없다."""
    student = token_student("S-AB12")
    student["이름"] = "김가상"
    rc, out = run_main(tmp_path, capsys, wrap([student]))
    assert rc == 1
    assert "FAIL 1반 1번째 학생 [PRIVACY_FIELD] 토큰 단계 초안에 실명 키(이름/학번)가 들어 있음" in out
    assert "해당 키를 제거하세요" in out
    assert "김가상" not in out


def test_token_student_with_sid_key_blocked_without_value(tmp_path, capsys):
    """"학번" 키 동시 보유는 STAGE_MIXED가 먼저 잡는다 — 어느 쪽이든 값은 안 나간다."""
    student = token_student("S-AB12")
    student["학번"] = "30101"
    rc, out = run_main(tmp_path, capsys, wrap([student]))
    assert rc == 1
    assert "STAGE_MIXED" in out or "PRIVACY_FIELD" in out
    assert "30101" not in out


def test_verify_token_drafts_direct_call_catches_sid_key():
    """함수 직접 호출 경로(STAGE_MIXED 게이트 없음)에서도 PRIVACY_FIELD로 잡힌다."""
    student = token_student("S-AB12")
    student["학번"] = "30101"
    student["이름"] = "김가상"
    report = V.verify_token_drafts(wrap([student]), PROFILE)
    codes = [c for r in report["rows"] for _, c, _ in r["issues"]]
    assert "PRIVACY_FIELD" in codes
    for r in report["rows"]:
        for _, _, msg in r["issues"]:
            assert "30101" not in msg and "김가상" not in msg


def test_invalid_token_value_not_stored_in_rows():
    """형식 위반 값은 rows에도 싣지 않는다 — 하류 어디서도 찍을 수 없어야 한다."""
    report = V.verify_token_drafts(wrap([token_student("30101")]), PROFILE)
    dumped = json.dumps(report, ensure_ascii=False)
    assert "30101" not in dumped
    assert report["rows"][0]["토큰"] == ""
    assert report["rows"][0]["표시"] == "1번째 학생"


# ---------------------------------------------------------------------------
# ③ 실명 단계 — 매핑 값 검증(MAPPING_INVALID / MAPPING_DUP)
# ---------------------------------------------------------------------------

def test_corrupt_mapping_value_fails_without_leak(tmp_path, capsys):
    """오염 매핑 값("40202")은 행별 출력·차단 메시지 어디에도 나가면 안 된다."""
    drafts = wrap([named_student("30101", "김가상")])
    roster = {"students": [{"학번": "30101", "이름": "김가상"}]}
    mapping = {"활동": "가상 활동", "map": {"30101": "40202"}}
    rc, out = run_main(tmp_path, capsys, drafts, roster=roster, mapping=mapping)
    assert rc == 1
    assert "FAIL [MAPPING_INVALID] 매핑에 형식 위반 토큰 1건" in out
    assert "매핑 파일 손상 의심" in out
    assert "40202" not in out
    assert "30101" not in out
    assert not SID_LIKE.search(out), out


def test_mapping_duplicate_value_fails_mapping_dup(tmp_path, capsys):
    """같은 토큰 값이 두 학번에 걸리면 MAPPING_DUP로 집계돼 저장이 막힌다."""
    drafts = wrap([named_student("30101", "김가상"),
                   named_student("30102", "이가상")])
    roster = {"students": [{"학번": "30101", "이름": "김가상"},
                           {"학번": "30102", "이름": "이가상"}]}
    mapping = {"활동": "가상 활동", "map": {"30101": "S-AB12", "30102": "S-AB12"}}
    rc, out = run_main(tmp_path, capsys, drafts, roster=roster, mapping=mapping)
    assert rc == 1
    assert "FAIL [MAPPING_DUP] 같은 토큰 값이 여러 학번에 걸림 1건" in out
    assert not SID_LIKE.search(out), out


def test_clean_mapping_flow_unchanged(tmp_path, capsys):
    """정상 S-XXXX 매핑 흐름은 종전대로 행별 토큰 출력 + FAIL 0으로 통과한다."""
    drafts = wrap([named_student("30101", "김가상")])
    roster = {"students": [{"학번": "30101", "이름": "김가상"}]}
    mapping = {"활동": "가상 활동", "map": {"30101": "S-AB12"}}
    rc, out = run_main(tmp_path, capsys, drafts, roster=roster, mapping=mapping)
    assert rc == 0
    assert "MAPPING_INVALID" not in out and "MAPPING_DUP" not in out
    assert "결과: FAIL 0건" in out
    assert not SID_LIKE.search(out), out


# ---------------------------------------------------------------------------
# ④ 정상 토큰 흐름 회귀 + --save xlsx 권한
# ---------------------------------------------------------------------------

def test_valid_token_stage_flow_unchanged(tmp_path, capsys):
    """정상 S-XXXX 토큰 초안은 종전대로 통과한다(선검증이 오탐하지 않는다)."""
    rc, out = run_main(tmp_path, capsys, wrap([token_student("S-AB12"),
                                               token_student("S-CD34")]))
    assert rc == 0
    assert "TOKEN_INVALID" not in out and "PRIVACY_FIELD" not in out
    assert "결과: FAIL 0건" in out


def _run_save(tmp_path, capsys):
    drafts = wrap([named_student("30101", "김가상")])
    roster = {"students": [{"학번": "30101", "이름": "김가상"}]}
    rc, out = run_main(tmp_path, capsys, drafts, roster=roster,
                       save=True, expected_count=1)
    assert rc == 0, out
    assert "저장 완료" in out
    return tmp_path / "out.xlsx"


def test_save_xlsx_succeeds_with_intact_content(tmp_path, capsys):
    """0o600 적용 후에도 저장 성공·내용 무결성·임시 파일 정리는 그대로다(전 OS)."""
    from openpyxl import load_workbook

    out_path = _run_save(tmp_path, capsys)
    assert out_path.exists()
    assert not (tmp_path / "out.xlsx.tmp").exists()
    ws = load_workbook(str(out_path))["1반"]
    assert ws["A2"].value == "30101"
    assert ws["B2"].value == "김가상"
    assert ws["E2"].value == GOOD_TEXT


@posix_only
def test_save_xlsx_mode_is_0600(tmp_path, capsys):
    """검수 xlsx는 실명 산출물 — 생성 순간부터 소유자 전용(0o600)이어야 한다."""
    out_path = _run_save(tmp_path, capsys)
    assert stat.S_IMODE(os.stat(out_path).st_mode) == 0o600
