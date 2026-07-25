# -*- coding: utf-8 -*-
"""pseudonymize.py CLI 서브커맨드 테스트.

핵심 계약: 명렬(실명·학번)이 이 CLI를 거치는 동안 stdout에 절대 등장하지 않는다.
LLM(에이전트)은 이 CLI의 stdout만 읽으므로, stdout에 이름·학번이 새면 이 기능의
존재 이유가 사라진다.
"""
import json
import subprocess
import sys
from pathlib import Path

SCRIPT = str(Path(__file__).resolve().parents[1] / "pseudonymize.py")

ROSTER = {"students": [
    {"학번": "10101", "이름": "김가상"},
    {"학번": "10102", "이름": "이허구"},
    {"학번": "10103", "이름": "박미정"},
]}

NAMES = ["김가상", "이허구", "박미정"]
IDS = ["10101", "10102", "10103"]


def run(*args):
    return subprocess.run(
        [sys.executable, SCRIPT, *args],
        capture_output=True, text=True, encoding="utf-8",
    )


def assert_no_pii(text):
    for name in NAMES:
        assert name not in text
    for sid in IDS:
        assert sid not in text


def _make_roster_xlsx(tmp_path):
    from openpyxl import Workbook

    p = tmp_path / "명단.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["학번", "이름"])
    for s in ROSTER["students"]:
        ws.append([s["학번"], s["이름"]])
    wb.save(p)
    return p


# ---------------------------------------------------------------------------
# roster
# ---------------------------------------------------------------------------

def test_roster_cli_happy_path(tmp_path):
    src = _make_roster_xlsx(tmp_path)
    out = tmp_path / "명렬.json"
    proc = run("roster", str(src), "--out", str(out))
    assert proc.returncode == 0
    assert out.exists()
    saved = json.loads(out.read_text(encoding="utf-8"))
    assert len(saved["students"]) == 3
    assert "명렬 인식: 3명" in proc.stdout
    assert "표헤더" in proc.stdout
    assert_no_pii(proc.stdout)
    assert_no_pii(proc.stderr)


def test_roster_cli_pattern_mode_adds_warning(tmp_path):
    from openpyxl import Workbook

    p = tmp_path / "명단.xlsx"
    wb = Workbook()
    ws = wb.active
    for s in ROSTER["students"][:2]:
        ws.append([s["학번"], s["이름"]])
    wb.save(p)
    out = tmp_path / "명렬.json"
    proc = run("roster", str(p), "--out", str(out))
    assert proc.returncode == 0
    assert "패턴" in proc.stdout
    assert "확인" in proc.stdout
    assert_no_pii(proc.stdout)


def test_roster_cli_skipped_rows_warns_without_pii(tmp_path):
    """건너뛴 행이 있으면 인원수만으로 경고하고, 이름·학번은 출력하지 않는다."""
    from openpyxl import Workbook

    p = tmp_path / "명단.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["학번", "이름"])
    ws.append(["10101", "김가상"])
    ws.append(["10102", ""])  # 이름이 비어 있어 건너뜀
    ws.append(["10103", "이허구"])
    wb.save(p)

    out = tmp_path / "명렬.json"
    proc = run("roster", str(p), "--out", str(out))
    assert proc.returncode == 0
    assert "경고: 1개 행을 건너뛰었습니다(학번 또는 이름이 비어 있음). 명단을 확인해 주세요." in proc.stdout
    assert_no_pii(proc.stdout)
    assert_no_pii(proc.stderr)


def test_roster_cli_short_name_warns_but_exits_zero(tmp_path):
    """1자 이름이 감지되면 mask/memo 이전에 미리 경고하되(교사가 고칠 기회), 인식
    자체는 성공(exit 0)으로 처리해 명렬.json은 저장된다."""
    from openpyxl import Workbook

    p = tmp_path / "명단.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["학번", "이름"])
    ws.append(["10104", "봄"])
    ws.append(["10105", "한소율"])
    wb.save(p)

    out = tmp_path / "명렬.json"
    proc = run("roster", str(p), "--out", str(out))
    assert proc.returncode == 0
    assert out.exists()
    assert "1자 이름" in proc.stdout
    assert "봄" not in proc.stdout
    assert "한소율" not in proc.stdout


def test_roster_cli_surname_column_reports_combination_without_values(tmp_path):
    """성 열과 이름 열을 결합했다는 사실은 알리되 값은 출력하지 않는다."""
    from openpyxl import Workbook

    p = tmp_path / "명단.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["학번", "성", "이름"])
    ws.append(["10104", "김", "봄"])
    wb.save(p)

    out = tmp_path / "명렬.json"
    proc = run("roster", str(p), "--out", str(out))
    assert proc.returncode == 0
    assert "성 열과 이름 열을 합쳐 전체 이름으로 사용합니다." in proc.stdout
    assert "김" not in proc.stdout
    assert "봄" not in proc.stdout
    saved = json.loads(out.read_text(encoding="utf-8"))
    assert saved["students"] == [{"학번": "10104", "이름": "김봄"}]


def test_roster_cli_status_word_row_warns_and_excludes(tmp_path):
    """이름 칸이 '자퇴'인 행은 명렬에서 제외되고, 건너뜀 경고에 상태 표기
    행이 포함되었음을 알린다. 이름 값('자퇴' 포함)은 출력하지 않는다."""
    from openpyxl import Workbook

    p = tmp_path / "명단.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["학번", "이름"])
    ws.append(["10101", "김가상"])
    ws.append(["10102", "이허구"])
    ws.append(["10114", "자퇴"])
    wb.save(p)

    out = tmp_path / "명렬.json"
    proc = run("roster", str(p), "--out", str(out))
    assert proc.returncode == 0
    saved = json.loads(out.read_text(encoding="utf-8"))
    assert len(saved["students"]) == 2
    assert "명렬 인식: 2명" in proc.stdout
    assert "건너뛰었습니다" in proc.stdout
    assert "(상태 표기 행 포함)" in proc.stdout
    assert "자퇴" not in proc.stdout
    assert_no_pii(proc.stdout)
    assert_no_pii(proc.stderr)


def test_roster_cli_zero_detected_fails(tmp_path):
    p = tmp_path / "빈파일.txt"
    p.write_text("아무 명단도 없는 문서입니다.\n", encoding="utf-8")
    out = tmp_path / "명렬.json"
    proc = run("roster", str(p), "--out", str(out))
    assert proc.returncode == 1
    assert not out.exists()
    assert "0명" in proc.stdout


# ---------------------------------------------------------------------------
# issue
# ---------------------------------------------------------------------------

def _write_roster_json(tmp_path):
    p = tmp_path / "명렬.json"
    p.write_text(json.dumps(ROSTER, ensure_ascii=False), encoding="utf-8")
    return p


def test_issue_cli_happy_path(tmp_path):
    roster_path = _write_roster_json(tmp_path)
    out = tmp_path / "매핑.json"
    proc = run("issue", "--roster", str(roster_path), "--submitted", "10101,10102", "--out", str(out))
    assert proc.returncode == 0
    assert out.exists()
    mapping = json.loads(out.read_text(encoding="utf-8"))
    assert set(mapping["map"]) == {"10101", "10102"}
    assert "제출자 2명" in proc.stdout
    assert "미발급(미제출) 1명" in proc.stdout
    assert_no_pii(proc.stdout)


def test_issue_cli_submitted_from_file(tmp_path):
    roster_path = _write_roster_json(tmp_path)
    submitted_file = tmp_path / "제출자.txt"
    submitted_file.write_text("10101\n10102\n", encoding="utf-8")
    out = tmp_path / "매핑.json"
    proc = run("issue", "--roster", str(roster_path), "--submitted-from", str(submitted_file), "--out", str(out))
    assert proc.returncode == 0
    mapping = json.loads(out.read_text(encoding="utf-8"))
    assert set(mapping["map"]) == {"10101", "10102"}
    assert_no_pii(proc.stdout)


def test_issue_cli_reuses_existing_mapping(tmp_path):
    roster_path = _write_roster_json(tmp_path)
    out = tmp_path / "매핑.json"
    run("issue", "--roster", str(roster_path), "--submitted", "10101", "--out", str(out))
    first = json.loads(out.read_text(encoding="utf-8"))
    proc = run("issue", "--roster", str(roster_path), "--submitted", "10101,10102", "--out", str(out))
    assert proc.returncode == 0
    second = json.loads(out.read_text(encoding="utf-8"))
    assert second["map"]["10101"] == first["map"]["10101"]
    assert "10102" in second["map"]
    assert_no_pii(proc.stdout)


# ---------------------------------------------------------------------------
# mask
# ---------------------------------------------------------------------------

def _write_mapping(tmp_path, roster_path, submitted):
    out = tmp_path / "매핑.json"
    run("issue", "--roster", str(roster_path), "--submitted", ",".join(submitted), "--out", str(out))
    return out, json.loads(out.read_text(encoding="utf-8"))


def test_mask_cli_happy_path(tmp_path):
    roster_path = _write_roster_json(tmp_path)
    mapping_path, mapping = _write_mapping(tmp_path, roster_path, ["10101", "10102"])

    input_json = tmp_path / "입력.json"
    input_json.write_text(json.dumps({"items": [
        {"학번": "10101", "본문": "10101 김가상은 이허구와 함께 발표함."},
        {"학번": "10102", "본문": "봄을 노래한 시를 분석함."},
    ]}, ensure_ascii=False), encoding="utf-8")

    out = tmp_path / "토큰본.json"
    proc = run("mask", str(input_json), "--roster", str(roster_path), "--mapping", str(mapping_path), "--out", str(out))
    assert proc.returncode == 0
    assert out.exists()
    saved = json.loads(out.read_text(encoding="utf-8"))
    assert len(saved["items"]) == 2
    for item in saved["items"]:
        assert "토큰" in item
        assert item["토큰"].startswith("S-")
    assert_no_pii(proc.stdout)
    assert_no_pii(proc.stderr)
    # 저장된 파일 자체에도 실명·학번이 없어야 한다(치환이 실제로 됐는지 확인)
    saved_text = out.read_text(encoding="utf-8")
    assert_no_pii(saved_text)
    assert "가명화: 2건 처리" in proc.stdout
    assert "학번 유출 0건" in proc.stdout
    assert "탐지 한계" in proc.stdout  # 경고가 있어도 탐지 한계 문구는 항상 출력된다


def test_mask_cli_always_prints_detection_limit_warning_even_with_zero_name_warnings(tmp_path):
    """골든리포트 §6: 경고 0건이어도 '경고 0건 = 안전'으로 오인되지 않도록,
    mask 성공 stdout 끝에 탐지 한계 문구가 항상 출력되어야 한다."""
    roster_path = _write_roster_json(tmp_path)
    mapping_path, mapping = _write_mapping(tmp_path, roster_path, ["10101", "10102"])

    input_json = tmp_path / "입력.json"
    input_json.write_text(json.dumps({"items": [
        {"학번": "10101", "본문": "봄을 노래한 시를 분석함."},
        {"학번": "10102", "본문": "가을을 노래한 시를 분석함."},
    ]}, ensure_ascii=False), encoding="utf-8")

    out = tmp_path / "토큰본.json"
    proc = run("mask", str(input_json), "--roster", str(roster_path), "--mapping", str(mapping_path), "--out", str(out))
    assert proc.returncode == 0
    assert "치환 경고 0건" in proc.stdout
    assert "탐지 한계" in proc.stdout
    assert "0으로 만들지 않습니다" in proc.stdout
    assert_no_pii(proc.stdout)


def test_mask_cli_leak_blocks_save(tmp_path):
    """미발급자(미제출자)의 학번이 다른 학생 본문에 그대로 남으면 저장하지 않고 exit 1."""
    roster_path = _write_roster_json(tmp_path)
    # 10103(박미정)은 미제출 → 토큰 미발급
    mapping_path, mapping = _write_mapping(tmp_path, roster_path, ["10101"])

    input_json = tmp_path / "입력.json"
    input_json.write_text(json.dumps({"items": [
        {"학번": "10101", "본문": "10103 학생의 답안을 참고해 정리함."},
    ]}, ensure_ascii=False), encoding="utf-8")

    out = tmp_path / "토큰본.json"
    proc = run("mask", str(input_json), "--roster", str(roster_path), "--mapping", str(mapping_path), "--out", str(out))
    assert proc.returncode == 1
    assert not out.exists()
    assert_no_pii(proc.stdout)
    assert_no_pii(proc.stderr)


def test_mask_cli_unissued_ids_message_includes_count_without_pii(tmp_path):
    """매핑 없는 학번이 여러 건이면 건수를 밝히되 학번 값 자체는 출력하지 않는다."""
    roster_path = _write_roster_json(tmp_path)
    mapping_path, mapping = _write_mapping(tmp_path, roster_path, ["10101"])

    input_json = tmp_path / "입력.json"
    input_json.write_text(json.dumps({"items": [
        {"학번": "10102", "본문": "발표를 준비함."},
        {"학번": "10103", "본문": "토론에 참여함."},
    ]}, ensure_ascii=False), encoding="utf-8")

    out = tmp_path / "토큰본.json"
    proc = run("mask", str(input_json), "--roster", str(roster_path), "--mapping", str(mapping_path), "--out", str(out))
    assert proc.returncode == 1
    assert not out.exists()
    assert "매핑(토큰)이 없는 학번 2건" in proc.stdout
    assert "명렬 인식에서 누락되었을 수 있습니다" in proc.stdout
    assert_no_pii(proc.stdout)
    assert_no_pii(proc.stderr)


def test_mask_cli_unissued_own_id_fails_closed(tmp_path):
    """입력 항목의 학번 자체가 매핑에 없으면(토큰 발급 누락) 저장하지 않고 exit 1."""
    roster_path = _write_roster_json(tmp_path)
    mapping_path, mapping = _write_mapping(tmp_path, roster_path, ["10101"])

    input_json = tmp_path / "입력.json"
    input_json.write_text(json.dumps({"items": [
        {"학번": "10102", "본문": "발표를 준비함."},
    ]}, ensure_ascii=False), encoding="utf-8")

    out = tmp_path / "토큰본.json"
    proc = run("mask", str(input_json), "--roster", str(roster_path), "--mapping", str(mapping_path), "--out", str(out))
    assert proc.returncode == 1
    assert not out.exists()
    assert_no_pii(proc.stdout)


# ---------------------------------------------------------------------------
# mask — owner_id 정책(동명이인 오귀속 차단, 1자 이름 경고) CLI 레벨 검증
# ---------------------------------------------------------------------------

def test_mask_cli_duplicate_names_each_item_gets_own_token(tmp_path):
    """동명이인 골든 재현: mask가 owner_id를 넘겨 각 항목의 주인 이름만 주인
    토큰이 되고, 절대 남의 토큰이 되지 않아야 한다(오귀속 회귀 방지)."""
    roster = {"students": [
        {"학번": "30105", "이름": "이서준"},
        {"학번": "30110", "이름": "이서준"},
    ]}
    roster_path = tmp_path / "명렬.json"
    roster_path.write_text(json.dumps(roster, ensure_ascii=False), encoding="utf-8")
    mapping_path, mapping = _write_mapping(tmp_path, roster_path, ["30105", "30110"])
    token_a = mapping["map"]["30105"]
    token_b = mapping["map"]["30110"]

    input_json = tmp_path / "입력.json"
    input_json.write_text(json.dumps({"items": [
        {"학번": "30105", "본문": "이서준은 '나목'을 분석함."},
        {"학번": "30110", "본문": "이서준. '아몬드(손원평)'를 선정했다"},
    ]}, ensure_ascii=False), encoding="utf-8")

    out = tmp_path / "토큰본.json"
    proc = run("mask", str(input_json), "--roster", str(roster_path), "--mapping", str(mapping_path), "--out", str(out))
    assert proc.returncode == 0
    saved = json.loads(out.read_text(encoding="utf-8"))
    body_a = next(item["본문"] for item in saved["items"] if item["토큰"] == token_a)
    body_b = next(item["본문"] for item in saved["items"] if item["토큰"] == token_b)

    assert token_a in body_a
    assert token_b not in body_a
    assert token_b in body_b
    assert token_a not in body_b
    assert "이서준" not in body_a and "이서준" not in body_b


def test_mask_cli_single_char_name_blocks_pipeline_without_leaking_name_value(tmp_path):
    """1자 이름('봄')이 명렬에 남아 있으면(성 결합으로도 해소되지 않으면) 그 이름을
    본문에서 안전하게 가릴 수 없으므로 경고로 넘어가지 않고 저장 자체를 차단한다
    (fail-closed). 차단 메시지에도 이름 값 자체는 등장하지 않아야 한다."""
    roster = {"students": [
        {"학번": "10104", "이름": "봄"},
        {"학번": "10105", "이름": "한소율"},
    ]}
    roster_path = tmp_path / "명렬.json"
    roster_path.write_text(json.dumps(roster, ensure_ascii=False), encoding="utf-8")
    mapping_path, mapping = _write_mapping(tmp_path, roster_path, ["10104", "10105"])

    input_json = tmp_path / "입력.json"
    input_json.write_text(json.dumps({"items": [
        {"학번": "10104", "본문": "'봄봄(김유정)'을 선정하여 해학성을 분석함."},
        {"학번": "10105", "본문": "'봄봄(김유정)'에 나타난 해학성을 다룸."},
    ]}, ensure_ascii=False), encoding="utf-8")

    out = tmp_path / "토큰본.json"
    proc = run("mask", str(input_json), "--roster", str(roster_path), "--mapping", str(mapping_path), "--out", str(out))
    assert proc.returncode == 1
    assert not out.exists()
    assert "중단" in proc.stdout
    assert "1자 이름" in proc.stdout
    assert "봄" not in proc.stdout
    assert "한소율" not in proc.stdout
    assert_no_pii(proc.stderr)


def test_mask_cli_surname_combined_roster_no_longer_blocks(tmp_path):
    """성 열+이름 열이 결합되어 명렬에 2자 이상 이름만 남으면 정상 진행된다."""
    roster = {"students": [{"학번": "10104", "이름": "김봄"}]}
    roster_path = tmp_path / "명렬.json"
    roster_path.write_text(json.dumps(roster, ensure_ascii=False), encoding="utf-8")
    mapping_path, mapping = _write_mapping(tmp_path, roster_path, ["10104"])

    input_json = tmp_path / "입력.json"
    input_json.write_text(json.dumps({"items": [
        {"학번": "10104", "본문": "'봄봄(김유정)'을 선정하여 해학성을 분석함."},
    ]}, ensure_ascii=False), encoding="utf-8")

    out = tmp_path / "토큰본.json"
    proc = run("mask", str(input_json), "--roster", str(roster_path), "--mapping", str(mapping_path), "--out", str(out))
    assert proc.returncode == 0
    assert out.exists()
    saved = json.loads(out.read_text(encoding="utf-8"))
    assert "봄봄(김유정)" in saved["items"][0]["본문"]


# ---------------------------------------------------------------------------
# finalize
# ---------------------------------------------------------------------------

def test_finalize_cli_happy_path(tmp_path):
    roster_path = _write_roster_json(tmp_path)
    mapping_path, mapping = _write_mapping(tmp_path, roster_path, ["10101", "10102"])
    token1 = mapping["map"]["10101"]
    token2 = mapping["map"]["10102"]

    draft = {"classes": [{"name": "1반", "students": [
        {"토큰": token1, "핵심소재": "가상의 책", "톤등급": "중", "세특": "작품을 분석함.", "비고": "", "예외": False},
        {"토큰": token2, "핵심소재": "가상의 시", "톤등급": "상", "세특": "감상을 정리함.", "비고": "", "예외": False},
    ]}]}
    draft_path = tmp_path / "토큰초안.json"
    draft_path.write_text(json.dumps(draft, ensure_ascii=False), encoding="utf-8")

    out = tmp_path / "실명초안.json"
    proc = run("finalize", str(draft_path), "--roster", str(roster_path), "--mapping", str(mapping_path), "--out", str(out))
    assert proc.returncode == 0
    assert out.exists()
    saved = json.loads(out.read_text(encoding="utf-8"))
    students = saved["classes"][0]["students"]
    assert {"학번": "10101", "이름": "김가상"}.items() <= students[0].items()
    assert {"학번": "10102", "이름": "이허구"}.items() <= students[1].items()
    assert "재결합: 2명 복원 완료" in proc.stdout
    # stdout 자체에는 실명·학번이 등장하면 안 된다(에이전트가 읽는 것은 stdout뿐)
    assert_no_pii(proc.stdout)
    assert_no_pii(proc.stderr)


def test_finalize_cli_unmapped_token_fails(tmp_path):
    roster_path = _write_roster_json(tmp_path)
    mapping_path, mapping = _write_mapping(tmp_path, roster_path, ["10101"])

    draft = {"classes": [{"name": "1반", "students": [
        {"토큰": "S-FFFF", "핵심소재": "가상의 책", "톤등급": "중", "세특": "작품을 분석함.", "비고": "", "예외": False},
    ]}]}
    draft_path = tmp_path / "토큰초안.json"
    draft_path.write_text(json.dumps(draft, ensure_ascii=False), encoding="utf-8")

    out = tmp_path / "실명초안.json"
    proc = run("finalize", str(draft_path), "--roster", str(roster_path), "--mapping", str(mapping_path), "--out", str(out))
    assert proc.returncode == 1
    assert not out.exists()
    assert_no_pii(proc.stdout)


# ---------------------------------------------------------------------------
# memo — 교사 관찰 메모를 대화가 아니라 파일로 받아 토큰화한다.
# 핵심 계약: 이름·학번·메모 내용이 stdout/stderr에 등장하면 안 된다.
# ---------------------------------------------------------------------------

def test_memo_cli_xlsx_happy_path(tmp_path):
    from openpyxl import Workbook

    roster_path = _write_roster_json(tmp_path)
    mapping_path, mapping = _write_mapping(tmp_path, roster_path, ["10101", "10102"])

    p = tmp_path / "채점표.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["학번", "관찰 메모"])
    ws.append(["10101", "10101 김가상은 발표 때 질문이 좋았다."])
    ws.append(["10102", "이허구와 함께 성실히 참여함."])
    wb.save(p)

    out = tmp_path / "관찰메모.json"
    proc = run("memo", str(p), "--roster", str(roster_path), "--mapping", str(mapping_path), "--out", str(out))
    assert proc.returncode == 0
    assert out.exists()

    saved = json.loads(out.read_text(encoding="utf-8"))
    assert len(saved["items"]) == 2
    for item in saved["items"]:
        assert item["토큰"].startswith("S-")
        assert "메모" in item

    saved_text = out.read_text(encoding="utf-8")
    assert_no_pii(saved_text)
    assert_no_pii(proc.stdout)
    assert_no_pii(proc.stderr)
    assert "관찰 메모: 2건 수집" in proc.stdout
    assert "학번 유출 0건" in proc.stdout


def test_memo_cli_xlsx_header_variants(tmp_path):
    """헤더 공백·표기 변형(특기사항, 관찰메모)도 자동 탐지한다."""
    from openpyxl import Workbook

    roster_path = _write_roster_json(tmp_path)
    mapping_path, mapping = _write_mapping(tmp_path, roster_path, ["10101"])

    p = tmp_path / "채점표.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["학번", "특기사항"])
    ws.append(["10101", "성실히 참여함."])
    wb.save(p)

    out = tmp_path / "관찰메모.json"
    proc = run("memo", str(p), "--roster", str(roster_path), "--mapping", str(mapping_path), "--out", str(out))
    assert proc.returncode == 0
    saved = json.loads(out.read_text(encoding="utf-8"))
    assert len(saved["items"]) == 1
    assert_no_pii(proc.stdout)


def test_memo_cli_xlsx_missing_header_fails(tmp_path):
    from openpyxl import Workbook

    roster_path = _write_roster_json(tmp_path)
    mapping_path, mapping = _write_mapping(tmp_path, roster_path, ["10101"])

    p = tmp_path / "채점표.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["학번", "비고"])
    ws.append(["10101", "성실히 참여함."])
    wb.save(p)

    out = tmp_path / "관찰메모.json"
    proc = run("memo", str(p), "--roster", str(roster_path), "--mapping", str(mapping_path), "--out", str(out))
    assert proc.returncode == 1
    assert not out.exists()
    assert "메모 열 헤더를 '관찰 메모'로 지정해 주세요" in proc.stdout
    assert_no_pii(proc.stdout)


def test_memo_cli_xlsx_skips_empty_memo_cells(tmp_path):
    from openpyxl import Workbook

    roster_path = _write_roster_json(tmp_path)
    mapping_path, mapping = _write_mapping(tmp_path, roster_path, ["10101", "10102"])

    p = tmp_path / "채점표.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["학번", "관찰 메모"])
    ws.append(["10101", "발표가 인상적이었다."])
    ws.append(["10102", ""])
    wb.save(p)

    out = tmp_path / "관찰메모.json"
    proc = run("memo", str(p), "--roster", str(roster_path), "--mapping", str(mapping_path), "--out", str(out))
    assert proc.returncode == 0
    saved = json.loads(out.read_text(encoding="utf-8"))
    assert len(saved["items"]) == 1
    assert "관찰 메모: 1건 수집" in proc.stdout


def test_memo_cli_text_happy_path(tmp_path):
    roster_path = _write_roster_json(tmp_path)
    mapping_path, mapping = _write_mapping(tmp_path, roster_path, ["10101", "10102"])

    p = tmp_path / "메모.txt"
    p.write_text(
        "10101: 김가상은 발표 때 질문이 좋았다.\n"
        "10102 이허구와 함께 성실히 참여함.\n",
        encoding="utf-8",
    )

    out = tmp_path / "관찰메모.json"
    proc = run("memo", str(p), "--roster", str(roster_path), "--mapping", str(mapping_path), "--out", str(out))
    assert proc.returncode == 0
    saved = json.loads(out.read_text(encoding="utf-8"))
    assert len(saved["items"]) == 2
    for item in saved["items"]:
        assert item["토큰"].startswith("S-")

    saved_text = out.read_text(encoding="utf-8")
    assert_no_pii(saved_text)
    assert_no_pii(proc.stdout)
    assert_no_pii(proc.stderr)
    assert "관찰 메모: 2건 수집" in proc.stdout


def test_memo_cli_unmapped_id_excluded(tmp_path):
    """명렬에 없거나(미제출) 매핑이 없는 학번의 메모는 건수만 반영하고 제외한다."""
    roster_path = _write_roster_json(tmp_path)
    # 10103(박미정)은 미제출 → 토큰 미발급
    mapping_path, mapping = _write_mapping(tmp_path, roster_path, ["10101"])

    p = tmp_path / "메모.txt"
    p.write_text(
        "10101: 발표가 좋았다.\n"
        "10103: 박미정은 미제출이지만 메모가 남아 있다.\n",
        encoding="utf-8",
    )

    out = tmp_path / "관찰메모.json"
    proc = run("memo", str(p), "--roster", str(roster_path), "--mapping", str(mapping_path), "--out", str(out))
    assert proc.returncode == 0
    saved = json.loads(out.read_text(encoding="utf-8"))
    assert len(saved["items"]) == 1
    assert saved["items"][0]["토큰"] == mapping["map"]["10101"]
    assert "1건 수집(1건은 매핑 없음으로 제외)" in proc.stdout
    assert_no_pii(proc.stdout)
    assert_no_pii(out.read_text(encoding="utf-8"))


def test_memo_cli_leak_blocks_save(tmp_path):
    """메모 본문에 매핑 없는(미제출자) 학번이 그대로 남으면 저장하지 않고 exit 1."""
    roster_path = _write_roster_json(tmp_path)
    # 10103(박미정)은 미제출 → 토큰 미발급이라 pseudonymize_text가 치환할 수 없다
    mapping_path, mapping = _write_mapping(tmp_path, roster_path, ["10101"])

    p = tmp_path / "메모.txt"
    p.write_text("10101: 10103 학생과 비교하면 발표가 좋았다.\n", encoding="utf-8")

    out = tmp_path / "관찰메모.json"
    proc = run("memo", str(p), "--roster", str(roster_path), "--mapping", str(mapping_path), "--out", str(out))
    assert proc.returncode == 1
    assert not out.exists()
    assert_no_pii(proc.stdout)
    assert_no_pii(proc.stderr)


# ---------------------------------------------------------------------------
# memo — BOM 내성 (메모장·PowerShell 5.1 -Encoding utf8 기본값은 UTF-8 BOM을 붙인다)
# ---------------------------------------------------------------------------

def test_memo_cli_text_with_bom_still_parses(tmp_path):
    """메모장에서 저장한 BOM 붙은 텍스트 메모도 조용히 0건 처리되지 않고 정상 수집된다."""
    roster_path = _write_roster_json(tmp_path)
    mapping_path, mapping = _write_mapping(tmp_path, roster_path, ["10101"])

    p = tmp_path / "메모.txt"
    p.write_bytes("10101: 모둠 진행을 맡음.\n".encode("utf-8-sig"))

    out = tmp_path / "관찰메모.json"
    proc = run("memo", str(p), "--roster", str(roster_path), "--mapping", str(mapping_path), "--out", str(out))
    assert proc.returncode == 0
    assert out.exists()
    saved = json.loads(out.read_text(encoding="utf-8"))
    assert len(saved["items"]) == 1
    assert "관찰 메모: 1건 수집" in proc.stdout
    assert_no_pii(proc.stdout)


# ---------------------------------------------------------------------------
# memo — 침묵 실패 제거: 0건 수집이 조용히 성공으로 보이면 안 된다
# ---------------------------------------------------------------------------

def test_memo_cli_text_zero_candidates_fails(tmp_path):
    """내용은 있지만 '학번: 메모' 형식이 하나도 없으면 조용히 0건 처리하지 않고 exit 1."""
    roster_path = _write_roster_json(tmp_path)
    mapping_path, mapping = _write_mapping(tmp_path, roster_path, ["10101"])

    p = tmp_path / "메모.txt"
    p.write_text(
        "오늘 수업은 순조롭게 진행되었다.\n형식이 학번으로 시작하지 않는다.\n",
        encoding="utf-8",
    )

    out = tmp_path / "관찰메모.json"
    proc = run("memo", str(p), "--roster", str(roster_path), "--mapping", str(mapping_path), "--out", str(out))
    assert proc.returncode == 1
    assert not out.exists()
    assert "메모를 한 건도 읽지 못했습니다" in proc.stdout
    assert_no_pii(proc.stdout)


def test_memo_cli_all_candidates_unmapped_warns_but_succeeds(tmp_path):
    """읽은 메모가 있으나 전부 미제출자(매핑 없음)라 수집 0건이면 exit 0을 유지하되
    경고 문구가 명확해야 한다(단순 '0건 수집'과 구분)."""
    roster_path = _write_roster_json(tmp_path)
    # 10103(박미정)만 미제출 → 매핑 없음
    mapping_path, mapping = _write_mapping(tmp_path, roster_path, ["10101"])

    p = tmp_path / "메모.txt"
    p.write_text("10103: 박미정은 미제출이지만 메모가 남아 있다.\n", encoding="utf-8")

    out = tmp_path / "관찰메모.json"
    proc = run("memo", str(p), "--roster", str(roster_path), "--mapping", str(mapping_path), "--out", str(out))
    assert proc.returncode == 0
    assert "수집 0건" in proc.stdout
    assert "미제출자 확인 필요" in proc.stdout
    assert_no_pii(proc.stdout)


def test_memo_cli_empty_file_fails(tmp_path):
    """메모 파일이 완전히 비어 있으면 exit 1 + 명확한 안내."""
    roster_path = _write_roster_json(tmp_path)
    mapping_path, mapping = _write_mapping(tmp_path, roster_path, ["10101"])

    p = tmp_path / "메모.txt"
    p.write_text("", encoding="utf-8")

    out = tmp_path / "관찰메모.json"
    proc = run("memo", str(p), "--roster", str(roster_path), "--mapping", str(mapping_path), "--out", str(out))
    assert proc.returncode == 1
    assert not out.exists()
    assert "메모 파일이 비어 있습니다" in proc.stdout
    assert_no_pii(proc.stdout)


# ---------------------------------------------------------------------------
# memo — 1자 이름 fail-closed (mask와 동일한 정책이 memo에도 적용되어야 한다)
# ---------------------------------------------------------------------------

def test_memo_cli_single_char_name_blocks_pipeline_without_leaking_name_value(tmp_path):
    """1자 이름이 명렬에 남아 있으면 관찰 메모 토큰화도 저장하지 않고 exit 1로
    중단한다. 차단 메시지에 이름 값이 등장하면 안 된다."""
    roster = {"students": [{"학번": "10104", "이름": "봄"}]}
    roster_path = tmp_path / "명렬.json"
    roster_path.write_text(json.dumps(roster, ensure_ascii=False), encoding="utf-8")
    mapping_path, mapping = _write_mapping(tmp_path, roster_path, ["10104"])

    p = tmp_path / "메모.txt"
    p.write_text("10104: 질문이 꾸준했고 봄의 발표 태도가 안정적임.\n", encoding="utf-8")

    out = tmp_path / "관찰메모.json"
    proc = run("memo", str(p), "--roster", str(roster_path), "--mapping", str(mapping_path), "--out", str(out))
    assert proc.returncode == 1
    assert not out.exists()
    assert "중단" in proc.stdout
    assert "1자 이름" in proc.stdout
    assert "봄" not in proc.stdout
    assert_no_pii(proc.stderr)


def test_memo_cli_normal_roster_unaffected_by_short_name_gate(tmp_path):
    """정상(2자 이상) 명렬에서는 회귀 없이 종전대로 관찰 메모 토큰화가 통과한다."""
    roster_path = _write_roster_json(tmp_path)
    mapping_path, mapping = _write_mapping(tmp_path, roster_path, ["10101"])

    p = tmp_path / "메모.txt"
    p.write_text("10101: 성실히 참여함.\n", encoding="utf-8")

    out = tmp_path / "관찰메모.json"
    proc = run("memo", str(p), "--roster", str(roster_path), "--mapping", str(mapping_path), "--out", str(out))
    assert proc.returncode == 0
    assert out.exists()


# ---------------------------------------------------------------------------
# score — 채점표 점수를 대화나 직접 열람이 아니라 CLI로만 받는다.
# 핵심 계약: 학번·이름이 stdout/stderr/산출 파일에 등장하면 안 된다.
# 단, 점수 분포(집계값)는 교사 확인 편의를 위해 stdout에 출력한다.
# ---------------------------------------------------------------------------

def _make_score_xlsx(tmp_path, rows, header=("학번", "점수"), filename="채점표.xlsx"):
    from openpyxl import Workbook

    p = tmp_path / filename
    wb = Workbook()
    ws = wb.active
    ws.append(list(header))
    for row in rows:
        ws.append(list(row))
    wb.save(p)
    return p


def test_score_cli_happy_path(tmp_path):
    """정상 경로: 토큰↔점수 매핑이 정확히 대응한다."""
    roster_path = _write_roster_json(tmp_path)
    mapping_path, mapping = _write_mapping(tmp_path, roster_path, ["10101", "10102", "10103"])

    p = _make_score_xlsx(tmp_path, [
        ("10101", "15"),
        ("10102", "13"),
        ("10103", "15"),
    ])

    out = tmp_path / "점수.json"
    proc = run("score", str(p), "--roster", str(roster_path), "--mapping", str(mapping_path), "--out", str(out))
    assert proc.returncode == 0
    assert out.exists()

    saved = json.loads(out.read_text(encoding="utf-8"))
    assert len(saved["items"]) == 3
    by_token = {item["토큰"]: item["점수"] for item in saved["items"]}
    assert by_token[mapping["map"]["10101"]] == 15
    assert by_token[mapping["map"]["10102"]] == 13
    assert by_token[mapping["map"]["10103"]] == 15


def test_score_cli_no_pii_in_stdout_stderr_and_output(tmp_path):
    """핵심 계약: stdout·stderr·산출 파일 어디에도 이름·학번이 없어야 한다."""
    roster_path = _write_roster_json(tmp_path)
    mapping_path, mapping = _write_mapping(tmp_path, roster_path, ["10101", "10102"])

    p = _make_score_xlsx(tmp_path, [
        ("10101", "15"),
        ("10102", "9"),
    ])

    out = tmp_path / "점수.json"
    proc = run("score", str(p), "--roster", str(roster_path), "--mapping", str(mapping_path), "--out", str(out))
    assert proc.returncode == 0
    assert_no_pii(proc.stdout)
    assert_no_pii(proc.stderr)
    assert_no_pii(out.read_text(encoding="utf-8"))


def test_score_cli_reports_distribution(tmp_path):
    """점수 분포는 집계값이라 식별정보가 아니므로 stdout에 출력한다."""
    roster_path = _write_roster_json(tmp_path)
    mapping_path, mapping = _write_mapping(tmp_path, roster_path, ["10101", "10102", "10103"])

    p = _make_score_xlsx(tmp_path, [
        ("10101", "15"),
        ("10102", "13"),
        ("10103", "15"),
    ])

    out = tmp_path / "점수.json"
    proc = run("score", str(p), "--roster", str(roster_path), "--mapping", str(mapping_path), "--out", str(out))
    assert proc.returncode == 0
    assert "점수 수집: 3명" in proc.stdout
    assert "열: B — 점수" in proc.stdout
    assert "15점 2명" in proc.stdout
    assert "13점 1명" in proc.stdout
    assert_no_pii(proc.stdout)


def test_score_cli_missing_header_fails(tmp_path):
    """점수 열 헤더('점수'/'총점'/'합계'/'평가점수')가 없으면 exit 1 + 안내."""
    roster_path = _write_roster_json(tmp_path)
    mapping_path, mapping = _write_mapping(tmp_path, roster_path, ["10101"])

    p = _make_score_xlsx(tmp_path, [("10101", "성실히 참여함")], header=("학번", "비고"))

    out = tmp_path / "점수.json"
    proc = run("score", str(p), "--roster", str(roster_path), "--mapping", str(mapping_path), "--out", str(out))
    assert proc.returncode == 1
    assert not out.exists()
    assert "점수 열 헤더를 '점수'로 지정하거나 --column 으로 알려 주세요." in proc.stdout
    assert_no_pii(proc.stdout)


def test_score_cli_excludes_unmapped_and_missing_scores_with_counts(tmp_path):
    """미제출자(매핑 없음)와 빈 점수는 제외하되 건수로 보고한다."""
    roster_path = _write_roster_json(tmp_path)
    # 10103(박미정)은 미제출 → 토큰 미발급
    mapping_path, mapping = _write_mapping(tmp_path, roster_path, ["10101", "10102"])

    p = _make_score_xlsx(tmp_path, [
        ("10101", "15"),
        ("10102", ""),       # 점수 없음
        ("10103", "12"),     # 매핑 없음(미제출자)
    ])

    out = tmp_path / "점수.json"
    proc = run("score", str(p), "--roster", str(roster_path), "--mapping", str(mapping_path), "--out", str(out))
    assert proc.returncode == 0
    saved = json.loads(out.read_text(encoding="utf-8"))
    assert len(saved["items"]) == 1
    assert saved["items"][0]["토큰"] == mapping["map"]["10101"]
    assert "제외: 매핑 없음 1건, 점수 없음 1건" in proc.stdout
    assert_no_pii(proc.stdout)
    assert_no_pii(out.read_text(encoding="utf-8"))


def test_score_cli_explicit_column_option(tmp_path):
    """--column으로 비표준 헤더명을 명시 지정하면 그 열을 점수 열로 인식한다."""
    roster_path = _write_roster_json(tmp_path)
    mapping_path, mapping = _write_mapping(tmp_path, roster_path, ["10101"])

    p = _make_score_xlsx(tmp_path, [("10101", "18")], header=("학번", "수행평가점수"))

    out = tmp_path / "점수.json"
    proc = run(
        "score", str(p), "--roster", str(roster_path), "--mapping", str(mapping_path),
        "--out", str(out), "--column", "수행평가점수",
    )
    assert proc.returncode == 0
    saved = json.loads(out.read_text(encoding="utf-8"))
    assert len(saved["items"]) == 1
    assert saved["items"][0]["점수"] == 18
    assert "열: B — 수행평가점수" in proc.stdout
    assert_no_pii(proc.stdout)


# ---------------------------------------------------------------------------
# score — 다중 헤더 + 다중 후보 (FIX 1): 점수 후보가 여럿이면 자동 선택하지
# 않는다. 실제 채점표(소설 비평하기/시 비평하기 두 활동이 한 표에 있고 각각
# '점수' 열이 있으며 '총점' 열도 따로 있는 구조)를 재현한다.
# ---------------------------------------------------------------------------

def _make_multi_header_score_xlsx(tmp_path, student_rows, filename="채점표.xlsx"):
    """1행 제목 + 2행 라벨(병합 그룹 포함) + 3행 하위 문항의 3중 헤더 채점표.

    열 배치: A=학번, B=이름, C~I=소설 비평하기(문항1~6, 기본점수), J=점수,
    K~Q=시 비평하기(문항1~6, 기본점수), R=점수, S=총점.
    """
    from openpyxl import Workbook

    p = tmp_path / filename
    wb = Workbook()
    ws = wb.active
    ws.append(["2026학년도 2학년 문학 수행평가 채점표"])
    ws.append([
        "학번", "이름",
        "소설 비평하기", "", "", "", "", "", "",
        "점수",
        "시 비평하기", "", "", "", "", "", "",
        "점수",
        "총점",
    ])
    ws.append([
        "", "",
        "문항1", "문항2", "문항3", "문항4", "문항5", "문항6", "기본점수",
        "",
        "문항1", "문항2", "문항3", "문항4", "문항5", "문항6", "기본점수",
        "", "",
    ])
    for row in student_rows:
        ws.append(list(row))
    wb.save(p)
    return p


def _blank_novel_and_poem_scores(sid, novel_score, poem_score, total_score):
    """소설 항목 6개+기본점수는 비워 두고 소설점수(J)·시점수(R)·총점(S)만 채운다."""
    return (
        sid, "",  # 학번, 이름(비움 — 이 테스트는 명렬과 무관)
        "", "", "", "", "", "", "",  # C~I 소설 문항 원점수(생략)
        novel_score,  # J
        "", "", "", "", "", "", "",  # K~Q 시 문항 원점수(생략)
        poem_score,  # R
        total_score,  # S
    )


def test_score_cli_multi_header_ambiguous_exits_and_lists_candidates_without_pii(tmp_path):
    """점수 후보 3개(J열 소설 점수, R열 시 점수, S열 총점)를 조용히 하나 고르지
    않고 exit 1로 멈춰 선택지(열 문자 + 병합 라벨)를 제시해야 한다. 학생 이름·
    학번은 절대 stdout에 등장하면 안 된다."""
    roster_path = _write_roster_json(tmp_path)
    mapping_path, mapping = _write_mapping(tmp_path, roster_path, ["10101", "10102", "10103"])

    p = _make_multi_header_score_xlsx(tmp_path, [
        _blank_novel_and_poem_scores("10101", 15, 13, 28),
        _blank_novel_and_poem_scores("10102", 12, 14, 26),
        _blank_novel_and_poem_scores("10103", 15, 15, 30),
    ])

    out = tmp_path / "점수.json"
    proc = run("score", str(p), "--roster", str(roster_path), "--mapping", str(mapping_path), "--out", str(out))

    assert proc.returncode == 1
    assert not out.exists()
    assert "점수 열 후보가 3개입니다" in proc.stdout
    assert "--column" in proc.stdout
    assert "J열" in proc.stdout
    assert "소설 비평하기 > 점수" in proc.stdout
    assert "R열" in proc.stdout
    assert "시 비평하기 > 점수" in proc.stdout
    assert "S열" in proc.stdout
    assert "총점" in proc.stdout
    # 핵심 계약: 후보 목록을 보여주는 중에도 학생 이름·학번은 절대 노출되지 않는다.
    assert_no_pii(proc.stdout)
    assert_no_pii(proc.stderr)


def test_score_cli_column_letter_selects_candidate(tmp_path):
    """--column R(엑셀 열 문자)로 지정하면 헤더 탐지를 건너뛰고 그 열을 바로 쓴다."""
    roster_path = _write_roster_json(tmp_path)
    mapping_path, mapping = _write_mapping(tmp_path, roster_path, ["10101", "10102", "10103"])

    p = _make_multi_header_score_xlsx(tmp_path, [
        _blank_novel_and_poem_scores("10101", 15, 13, 28),
        _blank_novel_and_poem_scores("10102", 12, 14, 26),
        _blank_novel_and_poem_scores("10103", 15, 15, 30),
    ])

    out = tmp_path / "점수.json"
    proc = run(
        "score", str(p), "--roster", str(roster_path), "--mapping", str(mapping_path),
        "--out", str(out), "--column", "R",
    )
    assert proc.returncode == 0
    saved = json.loads(out.read_text(encoding="utf-8"))
    by_token = {item["토큰"]: item["점수"] for item in saved["items"]}
    assert by_token[mapping["map"]["10101"]] == 13
    assert by_token[mapping["map"]["10102"]] == 14
    assert by_token[mapping["map"]["10103"]] == 15
    assert "열: R — 시 비평하기 > 점수" in proc.stdout
    assert_no_pii(proc.stdout)


def test_score_cli_column_header_name_총점_selects_candidate(tmp_path):
    """--column 총점(헤더 이름)으로 지정해도 정상 동작한다."""
    roster_path = _write_roster_json(tmp_path)
    mapping_path, mapping = _write_mapping(tmp_path, roster_path, ["10101", "10102", "10103"])

    p = _make_multi_header_score_xlsx(tmp_path, [
        _blank_novel_and_poem_scores("10101", 15, 13, 28),
        _blank_novel_and_poem_scores("10102", 12, 14, 26),
        _blank_novel_and_poem_scores("10103", 15, 15, 30),
    ])

    out = tmp_path / "점수.json"
    proc = run(
        "score", str(p), "--roster", str(roster_path), "--mapping", str(mapping_path),
        "--out", str(out), "--column", "총점",
    )
    assert proc.returncode == 0
    saved = json.loads(out.read_text(encoding="utf-8"))
    by_token = {item["토큰"]: item["점수"] for item in saved["items"]}
    assert by_token[mapping["map"]["10101"]] == 28
    assert by_token[mapping["map"]["10102"]] == 26
    assert by_token[mapping["map"]["10103"]] == 30
    assert "열: S — 총점" in proc.stdout
    assert_no_pii(proc.stdout)


def test_score_cli_single_candidate_still_auto_proceeds(tmp_path):
    """회귀: 점수 열 후보가 1개뿐이면(기존 단일 헤더 채점표) --column 없이도
    지금까지처럼 자동으로 진행한다."""
    roster_path = _write_roster_json(tmp_path)
    mapping_path, mapping = _write_mapping(tmp_path, roster_path, ["10101"])

    p = _make_score_xlsx(tmp_path, [("10101", "17")])

    out = tmp_path / "점수.json"
    proc = run("score", str(p), "--roster", str(roster_path), "--mapping", str(mapping_path), "--out", str(out))
    assert proc.returncode == 0
    saved = json.loads(out.read_text(encoding="utf-8"))
    assert saved["items"][0]["점수"] == 17
    assert "열: B — 점수" in proc.stdout
    assert_no_pii(proc.stdout)


# ---------------------------------------------------------------------------
# score --grade-column / --sum-columns (FIX 2): 점수 열이 아예 없는 채점표
# ---------------------------------------------------------------------------

def test_score_cli_grade_column_collects_grades_and_distribution(tmp_path):
    """--grade-column: 점수 대신 등급 문자열을 그대로 담고, 분포는 등급별로 집계한다."""
    roster_path = _write_roster_json(tmp_path)
    mapping_path, mapping = _write_mapping(tmp_path, roster_path, ["10101", "10102", "10103"])

    p = _make_score_xlsx(tmp_path, [
        ("10101", "상"),
        ("10102", "중"),
        ("10103", "상"),
    ], header=("학번", "등급"))

    out = tmp_path / "점수.json"
    proc = run(
        "score", str(p), "--roster", str(roster_path), "--mapping", str(mapping_path),
        "--out", str(out), "--grade-column", "등급",
    )
    assert proc.returncode == 0
    saved = json.loads(out.read_text(encoding="utf-8"))
    by_token = {item["토큰"]: item["등급"] for item in saved["items"]}
    assert by_token[mapping["map"]["10101"]] == "상"
    assert by_token[mapping["map"]["10102"]] == "중"
    assert by_token[mapping["map"]["10103"]] == "상"
    assert "등급 수집: 3명" in proc.stdout
    assert "상 2명" in proc.stdout
    assert "중 1명" in proc.stdout
    assert_no_pii(proc.stdout)


def test_score_cli_sum_columns_sums_range(tmp_path):
    """--sum-columns E:J: 지정 범위의 숫자를 합산해 점수로 쓴다."""
    from openpyxl import Workbook

    p = tmp_path / "채점표.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["학번", "이름", "비고", "비고2", "문항1", "문항2", "문항3", "문항4", "문항5", "기본점수"])
    ws.append(["10101", "", "", "", 3, 3, 3, 3, 3, 5])   # E~J 합 20
    ws.append(["10102", "", "", "", 2, 2, 2, 2, 2, 5])   # E~J 합 15
    wb.save(p)

    roster_path = _write_roster_json(tmp_path)
    mapping_path, mapping = _write_mapping(tmp_path, roster_path, ["10101", "10102"])

    out = tmp_path / "점수.json"
    proc = run(
        "score", str(p), "--roster", str(roster_path), "--mapping", str(mapping_path),
        "--out", str(out), "--sum-columns", "E:J",
    )
    assert proc.returncode == 0
    saved = json.loads(out.read_text(encoding="utf-8"))
    by_token = {item["토큰"]: item["점수"] for item in saved["items"]}
    assert by_token[mapping["map"]["10101"]] == 20
    assert by_token[mapping["map"]["10102"]] == 15
    assert "합산: E~J열" in proc.stdout
    assert_no_pii(proc.stdout)


def test_score_cli_mutually_exclusive_options_fail(tmp_path):
    """--column/--grade-column/--sum-columns 중 2개 이상 지정하면 exit 1."""
    roster_path = _write_roster_json(tmp_path)
    mapping_path, mapping = _write_mapping(tmp_path, roster_path, ["10101"])

    p = _make_score_xlsx(tmp_path, [("10101", "17")])

    out = tmp_path / "점수.json"
    proc = run(
        "score", str(p), "--roster", str(roster_path), "--mapping", str(mapping_path),
        "--out", str(out), "--column", "점수", "--grade-column", "등급",
    )
    assert proc.returncode == 1
    assert not out.exists()
    assert_no_pii(proc.stdout)
