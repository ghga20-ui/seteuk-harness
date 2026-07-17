# -*- coding: utf-8 -*-
"""평가 선행 게이트와 AI 초안 표시 테스트.

교사의 평가 자료가 확인되지 않은 활동은 세특을 저장할 수 없고,
저장되는 파일에는 AI 초안임이 명시되어야 한다.
"""
import json
import subprocess
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

PROFILE_WITH_EVAL = {
    "활동명": "가상 활동", "문두": "가상 활동에서",
    "목표바이트": 700, "상한바이트": 760,
    "평가자료": "가상 채점표.xlsx (2026-07-17 확인)",
}
PROFILE_NO_EVAL = {k: v for k, v in PROFILE_WITH_EVAL.items() if k != "평가자료"}

GOOD_TEXT = (
    "가상 활동에서 '가상의 책(작가)'을 선정하여 인물의 갈등에 주목하며 감상문을 작성함. "
    "서술 시점의 효과를 짚고 인물의 내적 갈등이 심화되는 과정을 정리함. "
    "작품에 반영된 사회 현실을 비판적으로 읽어냄. "
    "자신의 경험과 견주어 삶의 태도를 성찰하는 다짐을 밝힘. "
    "감상의 근거를 본문에서 찾아 제시하는 태도가 돋보임. "
    "작품을 자기 이해의 계기로 삼는 모습을 보임."
)
DRAFTS = {
    "classes": [{"name": "1반", "students": [
        {"학번": "10101", "이름": "김가상", "핵심소재": "가상의 책(작가)",
         "톤등급": "중", "세특": GOOD_TEXT, "비고": "", "예외": False}
    ]}]
}


def _run_cli(tmp_path, profile):
    drafts_path = tmp_path / "d.json"
    profile_path = tmp_path / "p.json"
    out = tmp_path / "out.xlsx"
    drafts_path.write_text(json.dumps(DRAFTS, ensure_ascii=False), encoding="utf-8")
    profile_path.write_text(json.dumps(profile, ensure_ascii=False), encoding="utf-8")
    script = str(Path(__file__).resolve().parents[1] / "verify_seteuk.py")
    proc = subprocess.run(
        [sys.executable, script, str(drafts_path), "--profile", str(profile_path), "--save", str(out)],
        capture_output=True, text=True, encoding="utf-8",
    )
    return proc, out


def test_cli_blocks_without_evaluation_evidence(tmp_path):
    proc, out = _run_cli(tmp_path, PROFILE_NO_EVAL)
    assert proc.returncode == 1
    assert not out.exists()
    assert "평가" in proc.stdout


def test_cli_passes_with_evaluation_evidence(tmp_path):
    proc, out = _run_cli(tmp_path, PROFILE_WITH_EVAL)
    assert proc.returncode == 0
    assert out.exists()


def test_saved_xlsx_has_ai_draft_notice_sheet(tmp_path):
    from openpyxl import load_workbook

    proc, out = _run_cli(tmp_path, PROFILE_WITH_EVAL)
    assert proc.returncode == 0
    wb = load_workbook(str(out))
    assert wb.sheetnames[0] == "안내"
    text = " ".join(str(c.value) for row in wb["안내"].iter_rows() for c in row if c.value)
    assert "AI 초안" in text
    assert "검수" in text
