#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""seteuk-harness 결정론 채점기 + 저장 게이트.

기계 규칙은 wiki/규칙.json(단일 정본)에서 읽는다. 파일이 없으면 내장 기본값을 쓴다.
규칙정본.md(사람용)를 개정할 때는 규칙.json도 함께 갱신한다.
"""
from __future__ import annotations

import json
import re
from pathlib import Path

# wiki/규칙.json이 없거나 손상됐을 때의 내장 기본값
DEFAULT_RULES = {
    "금지어휘": ["단순히", "또한", "이를 통해", "바탕으로", "무엇보다", "단순한", "넘어"],
    "이는_경계검사": True,
    "금지문자패턴": "[·\\-<>\"“”‘’*※①-⑳#&@]",
    "목표바이트": 700,
    "상한바이트": 760,
}
RULES_PATH = Path(__file__).resolve().parent.parent / "wiki" / "규칙.json"
INEUN = re.compile(r"(?:^|[ ,.'])이는(?=[ ,.]|$)")


def load_rules(path=RULES_PATH) -> dict:
    """규칙 파일을 읽어 기본값 위에 덮는다. 실패하면 기본값을 반환한다."""
    rules = dict(DEFAULT_RULES)
    try:
        # utf-8-sig: 규칙 파일을 메모장 등에서 편집하면 BOM이 붙을 수 있다.
        with open(path, encoding="utf-8-sig") as f:
            rules.update(json.load(f))
    except (OSError, ValueError):
        pass
    return rules


RULES = load_rules()
FIRST_QUOTED = re.compile(r"'([^']*)'")


def extract_subject(text: str) -> str:
    """세특 본문에서 첫 번째 작은따옴표 안 내용(작품명)을 추출한다.

    세특 규칙상 작은따옴표는 작품명 표기 전용이므로 첫 '…' 안 내용이 곧 작품명이다.
    작은따옴표가 없거나 닫히지 않으면(예외 처리된 빈 답안 등) 빈 문자열을 반환한다.
    중첩·비정상 따옴표가 있어도 첫 쌍만 취하고 예외를 내지 않는다.
    """
    if not text:
        return ""
    match = FIRST_QUOTED.search(text)
    if not match:
        return ""
    return match.group(1).strip()


def autofill_subjects(drafts: dict) -> tuple[int, int]:
    """핵심소재가 빈 학생에 한해 세특 본문 첫 작품명으로 채운다.

    이미 값이 있으면(교사가 채운 값 포함) 절대 덮어쓰지 않는다.
    반환값: (채운 인원수, 시도했으나 빈칸으로 남은 인원수).
    """
    filled = 0
    blank = 0
    for cls in drafts.get("classes", []):
        for student in cls.get("students", []):
            existing = str(student.get("핵심소재", "") or "").strip()
            if existing:
                continue
            subject = extract_subject(student.get("세특", ""))
            if subject:
                student["핵심소재"] = subject
                filled += 1
            else:
                blank += 1
    return filled, blank


def check_text(text: str, profile: dict, exempt: bool = False, rules: dict | None = None):
    """단일 세특 본문을 검사한다. 반환: (utf8 바이트수, [(레벨, 코드, 메시지)])."""
    if rules is None:
        rules = RULES
    # 활동프로파일이 같은 이름의 규칙 키를 갖고 있으면 그 활동에 한해 우선한다.
    # 규칙 파일은 전역 하나뿐이라, 수학처럼 부등호가 필요한 과목이 국어 규칙까지
    # 바꿔야 하는 문제가 있었다 — 과목·활동별 차이는 프로파일에 싣는다.
    for key in ("금지어휘", "금지문자패턴", "이는_경계검사"):
        if key in profile:
            rules = {**rules, key: profile[key]}
    issues: list[tuple[str, str, str]] = []
    nbytes = len(text.encode("utf-8"))

    # 문두는 교사가 인터뷰에서 확정한 값이므로 그 구간은 금지 검사에서 면제한다.
    # 활동명에 가운뎃점('시·소설 비교')이 있으면 문두 강제와 금지문자 검사가
    # 서로 물려 어느 쪽으로 써도 FAIL이 되는 교착이 있었다 — 검사는 본문에만 건다.
    prefix = profile.get("문두", "")
    scanned = text[len(prefix):] if (prefix and text.startswith(prefix)) else text

    for word in rules.get("금지어휘", []):
        if word in scanned:
            issues.append(("FAIL", "BANNED_WORD", f"금지 어휘 '{word}' 포함"))
    if rules.get("이는_경계검사", True) and INEUN.search(scanned):
        issues.append(("FAIL", "BANNED_WORD", "금지 어휘 '이는' 포함"))

    match = re.search(rules.get("금지문자패턴", DEFAULT_RULES["금지문자패턴"]), scanned)
    if match:
        issues.append(("FAIL", "BANNED_CHAR", f"금지 특수문자 '{match.group()}' 포함"))

    # 바이트 한도 및 문두 검사
    limit = int(profile.get("상한바이트", 760))
    target = int(profile.get("목표바이트", 700))
    if nbytes > limit:
        issues.append(("FAIL", "BYTE_OVER", f"{nbytes}바이트로 상한 {limit} 초과"))
    elif not exempt and nbytes < target - 80:
        issues.append(("WARN", "BYTE_UNDER", f"{nbytes}바이트로 목표 {target}에 크게 미달"))

    if not exempt:
        if prefix and not text.startswith(prefix):
            issues.append(("FAIL", "OPENING", f"문두가 '{prefix}'로 시작하지 않음"))

    return nbytes, issues


def find_name_intrusions(text: str, names: list[str]) -> list[str]:
    """본문에 등장한 학생 이름 목록을 반환한다(본인 포함 — 이름 기재 자체가 금지)."""
    return [name for name in names if name and name in text]


def load_submitted_ids(path) -> set[str]:
    """제출 원본에서 학번 집합을 뽑는다. 명렬과 무관한 별도 경로 데이터를 전제로 한다.

    핵심 설계: 명렬이 틀렸을 가능성을 명렬로 검증할 수는 없다. 이 함수가 읽는
    파일은 반드시 명렬에서 파생되지 않은 데이터여야 한다(제출 폴더의 파일 목록,
    또는 mask 단계가 만든 매핑.json — 그 키는 "실제로 제출한 학번"이며 명렬의
    이름 열과는 독립적으로 결정된다).

    지원 형식(모두 학번만 사용하고 이름은 참조하지 않는다):
    - 학번 문자열/숫자 리스트: ["10101", "10102"]
    - {"학번목록": [...]}
    - {"map": {"10101": "S-AB12", ...}} (매핑.json과 동일한 모양 — 키가 학번)
    - {"students": [{"학번": "10101"}, ...]}
    - {"classes": [{"students": [{"학번": "10101"}, ...]}]} (초안·명렬과 동일한 모양)
    """
    with open(path, encoding="utf-8-sig") as f:
        data = json.load(f)

    ids: set[str] = set()

    def add(value) -> None:
        if value is not None and str(value).strip():
            ids.add(str(value).strip())

    if isinstance(data, list):
        for item in data:
            if isinstance(item, dict):
                add(item.get("학번"))
            else:
                add(item)
    elif isinstance(data, dict):
        # 여러 모양이 한 파일에 같이 있으면 합집합을 만들지 않고 가장 구체적인
        # 것 하나만 쓴다. 명렬(students)과 제출자(학번목록)가 함께 든 파일에서
        # 합집합을 만들면 명렬 전원이 제출자가 되어 미제출자 귀속 게이트가
        # 조용히 죽는다 — 게이트가 살아 있는 것처럼 보이면서 아무것도 막지 않는다.
        if data.get("학번목록"):
            for v in data["학번목록"]:
                add(v)
        elif data.get("map"):
            for sid in data["map"]:
                add(sid)
        elif data.get("students"):
            for s in data["students"]:
                add(s.get("학번"))
        elif data.get("classes"):
            for cls in data["classes"]:
                for s in cls.get("students", []):
                    add(s.get("학번"))
    return ids


def verify_drafts(drafts: dict, profile: dict, roster: dict | None = None,
                   submitted: set[str] | None = None) -> dict:
    """전체 초안을 검사해 보고서를 만든다. 이름 혼입은 반 전체 명단으로 검사한다.

    roster({"students":[{"학번","이름"}]})를 주면 명렬 대조를 수행한다:
    초안에만 있는 학번·이름 불일치는 FAIL(ROSTER), 명렬에만 있는 학생은
    보고서의 "미제출" 목록으로 반환한다(잘못된 학생 귀속과 조용한 누락 방지).

    **명렬 대조(ROSTER)는 형식 일치만 증명한다.** 초안의 이름은 finalize
    단계에서 바로 이 명렬로부터 붙여진 것이므로, 여기서 다시 명렬과 비교하면
    명렬을 명렬로 검증하는 순환이 된다 — 명렬 자체가 잘못됐을 때(예: 이름 열이
    한 칸씩 밀린 경우) 전원의 이름이 서로 어긋나 있어도 이 대조는 100% 통과한다.
    실제 오귀속을 잡으려면 명렬에서 파생되지 않은 외부 정보가 필요하다:
    교사가 선언하는 인원수(verify_seteuk.py의 --expected-count 게이트)와
    제출 원본 대조(submitted 인자)가 그 역할을 한다.

    submitted(학번 집합)를 주면 명렬과 독립된 제출자 목록과 대조한다:
    초안에는 있는데 제출자 목록에 없는 학번은 FAIL(NOT_SUBMITTED — 제출하지
    않은 학생에게 세특이 생성된 것으로 의심), 제출자인데 초안에 없는 학번은
    WARN(SUBMITTED_NOT_DRAFTED)으로 건수만 보고한다.

    **issue 메시지 문자열에 학번을 보간하지 않는다.** 이 함수의 반환값은 로컬
    파이썬끼리만 오가지만, main()이 메시지를 stdout에 그대로 찍는 순간
    에이전트(LLM)에게 전송된다 — 학생 식별은 row의 학번(데이터)으로 하고,
    메시지는 코드(DUP_ID 등)와 일반 서술만 담는다.
    """
    roster_map = None
    경고: list[tuple[str, str, str]] = []
    if roster is not None:
        roster_map = {str(s.get("학번", "")): s.get("이름", "") for s in roster.get("students", [])}
    else:
        경고.append(("WARN", "NO_ROSTER", "명렬 미제공 — 학번 잔존 검사와 미제출자 감지가 생략됨"))

    if submitted is None:
        경고.append(("WARN", "NO_SUBMITTED_CHECK",
                     "제출 원본 대조를 건너뛰었습니다(--submitted 미지정) — "
                     "명렬과 독립된 경로로 제출 여부를 확인하지 않았으므로, "
                     "제출하지 않은 학생에게 세특이 생성돼도 이 검사로는 잡지 못합니다."))

    rows = []
    fail = warn = 0
    seen_ids = set()
    for cls in drafts.get("classes", []):
        names = [s.get("이름", "") for s in cls.get("students", [])]
        for student in cls.get("students", []):
            sid = str(student.get("학번", ""))
            sname = student.get("이름", "")
            text = student.get("세특", "")
            exempt = bool(student.get("예외", False))
            nbytes, issues = check_text(text, profile, exempt=exempt)
            if sid and sid in seen_ids:
                # 같은 학번이 두 반에 있으면 오귀속이 확정적으로 일어난다 —
                # 명렬 오인(출석번호를 학번으로) 또는 반 중복 편성 흔적이므로 막는다.
                issues.append(("FAIL", "DUP_ID",
                               "같은 학번이 초안에 두 번 이상 등장 — 오귀속 의심, 명렬을 확인하세요"))
            seen_ids.add(sid)
            for name in find_name_intrusions(text, names):
                issues.append(("FAIL", "NAME", "학생 이름이 본문에 혼입됨"))

            # 토큰 잔존 검사
            from pseudonymize import scan_token_residue
            for token in scan_token_residue(text):
                issues.append(("FAIL", "TOKEN_RESIDUE", f"토큰 '{token}'이 세특 본문에 남아 있음"))

            # 명렬이 제공되면 본문에 학번이 남아 있는지 검사
            if roster is not None:
                from pseudonymize import scan_id_in_narrative
                for _found_sid in scan_id_in_narrative(text, roster):
                    issues.append(("FAIL", "ID_IN_BODY",
                                   "세특 본문에 학번이 그대로 노출됨 — 식별자 없는 표현으로 재서술하세요"))

            if roster_map is not None:
                if sid not in roster_map:
                    issues.append(("FAIL", "ROSTER", "학번이 명렬에 없음 — 오전사 또는 오매핑 의심"))
                elif roster_map[sid] != sname:
                    issues.append(("FAIL", "ROSTER",
                                   "이름이 명렬과 불일치 — 학생 귀속 확인 필요"
                                   "(명렬 파일에서 직접 대조하세요)"))
                # else: 형식 일치만 확인된 것 — 이 이름은 애초에 같은 명렬에서
                # 왔으므로(finalize), 일치한다는 사실이 귀속 정확성을 증명하지
                # 않는다. 명렬 자체가 밀렸다면 여기서도 항상 일치로 보인다.

            if submitted is not None and sid not in submitted:
                # 미제출자에게도 활동 참여 사실만 담은 예외 세특을 쓰는 것이 규칙이므로
                # 예외 표시가 있으면 정상이다. 문제는 제출 기록이 없는데 정상 세특이
                # 있는 경우다 — 근거 없는 서술이 만들어졌다는 뜻이다.
                if exempt:
                    issues.append(("WARN", "NOT_SUBMITTED_EXEMPT",
                                   "미제출 — 예외 세특으로 처리됨"))
                else:
                    issues.append(("FAIL", "NOT_SUBMITTED",
                                    "제출 기록이 없는데 정상 세특이 있음 — "
                                    "근거 없는 서술 의심"))

            fail += sum(1 for lv, _, _ in issues if lv == "FAIL")
            warn += sum(1 for lv, _, _ in issues if lv == "WARN")
            rows.append(
                {"반": cls.get("name", ""), "학번": sid,
                 "이름": sname, "바이트": nbytes, "issues": issues}
            )

    missing = []
    if roster_map is not None:
        missing = [{"학번": sid, "이름": name} for sid, name in roster_map.items() if sid not in seen_ids]

    if submitted is not None:
        orphan_count = sum(1 for sid in submitted if sid not in seen_ids)
        if orphan_count:
            경고.append(("WARN", "SUBMITTED_NOT_DRAFTED",
                         f"제출자 중 초안에 없는 인원 {orphan_count}명 — 누락 후보(명렬과 무관한 대조)"))

    warn += len(경고)
    return {"rows": rows, "fail": fail, "warn": warn, "미제출": missing, "경고": 경고}


def detect_stage(drafts: dict) -> str:
    """초안 학생들의 키 모양으로 파이프라인 단계를 판별한다.

    - 전원 "토큰" 키 → "token"(§6 — finalize 이전, 실명·학번이 아직 없다)
    - 전원 "학번" 키 → "named"(§7 — finalize가 실명을 재결합한 뒤)
    - 혼재는 "mixed": 두 종류 학생이 공존, 한 학생이 두 키를 다 보유, 둘 다
      없음 — 어느 쪽이든 단계가 확정되지 않으므로 검사를 시작하면 안 된다
      (토큰 검사로 실명 초안을 통과시키거나, 실명 검사가 토큰 학생을 빈
      학번으로 오인해 전원 ROSTER FAIL을 내는 사고가 실제로 있었다).
    """
    kinds: set[str] = set()
    for cls in drafts.get("classes", []):
        for student in cls.get("students", []):
            has_token = "토큰" in student
            has_sid = "학번" in student
            if has_token == has_sid:  # 둘 다 보유 또는 둘 다 없음
                return "mixed"
            kinds.add("token" if has_token else "named")
    if len(kinds) != 1:
        return "mixed"
    return kinds.pop()


def verify_token_drafts(drafts: dict, profile: dict) -> dict:
    """토큰 초안(§6)을 검사한다. 식별자가 토큰뿐이라 stdout에 그대로 찍어도 안전하다.

    수행: check_text 전체(바이트·문두·금지어 등), 토큰 잔존(자기 토큰 포함 —
    본문에 남은 토큰은 finalize에서 학번으로 복원되어 그대로 노출된다),
    중복 토큰(DUP_TOKEN — 같은 토큰이 두 학생에게 붙었으면 오귀속 확정).
    건너뜀: 이름 혼입·명렬 대조·미제출 감지 — 초안에 이름·학번이 없어 검사
    자체가 성립하지 않는다. finalize 후 실명 단계(§7)에서 수행한다.
    """
    rows = []
    fail = warn = 0
    seen_tokens: set[str] = set()
    for cls in drafts.get("classes", []):
        for student in cls.get("students", []):
            token = str(student.get("토큰", ""))
            text = student.get("세특", "")
            exempt = bool(student.get("예외", False))
            nbytes, issues = check_text(text, profile, exempt=exempt)
            if token and token in seen_tokens:
                issues.append(("FAIL", "DUP_TOKEN",
                               "같은 토큰이 초안에 두 번 이상 등장 — 오귀속 의심, 매핑·초안을 확인하세요"))
            seen_tokens.add(token)

            # 토큰 잔존 검사 — 자기 토큰도 잡는다. 세특 본문에는 어떤 식별자도
            # 남으면 안 되고, 남은 토큰은 재결합 때 학번 숫자가 되어 노출된다.
            from pseudonymize import scan_token_residue
            for residue in scan_token_residue(text):
                issues.append(("FAIL", "TOKEN_RESIDUE", f"토큰 '{residue}'이 세특 본문에 남아 있음"))

            fail += sum(1 for lv, _, _ in issues if lv == "FAIL")
            warn += sum(1 for lv, _, _ in issues if lv == "WARN")
            rows.append({"반": cls.get("name", ""), "토큰": token, "바이트": nbytes, "issues": issues})
    return {"rows": rows, "fail": fail, "warn": warn}


def save_xlsx(drafts: dict, report: dict, out_path: str) -> None:
    """검증 보고서와 함께 반별 시트로 저장한다. 호출 전 FAIL 0을 보장할 것.

    바이트수 열은 정적 값이 아니라 엑셀 수식으로 넣는다 — 교사가 검수 중
    세특을 고치면 바이트가 실시간으로 재계산되어 분량 조절이 수월하다.
    수식 (LENB-LEN)*2+LEN 은 한글 3바이트 환산으로 UTF-8 기준과 일치한다.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    wb.remove(wb.active)
    notice = wb.create_sheet("안내")
    notice["A1"] = "AI 초안 — 교사 검수 전"
    notice["A2"] = "이 파일의 세특은 AI가 생성한 초안입니다. 교사가 실제 수행과의 일치, 허위·과장 여부,"
    notice["A3"] = "기재요령 준수를 검수·승인하기 전에는 NEIS에 입력할 수 없습니다."
    notice["A1"].font = Font(bold=True, size=14, color="C00000")
    notice.column_dimensions["A"].width = 90
    for cls in drafts.get("classes", []):
        ws = wb.create_sheet(cls.get("name", "시트"))
        ws.append(["학번", "이름", "핵심소재", "톤등급", "세특", "바이트수", "비고"])
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9E1F2")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row_idx, s in enumerate(cls.get("students", []), start=2):
            ws.append([s.get("학번", ""), s.get("이름", ""), s.get("핵심소재", ""),
                       s.get("톤등급", ""), s.get("세특", ""),
                       f"=(LENB(E{row_idx})-LEN(E{row_idx}))*2+LEN(E{row_idx})",
                       s.get("비고", "")])
        for col, width in zip("ABCDEFG", [8, 8, 24, 7, 90, 8, 30]):
            ws.column_dimensions[col].width = width
        for row in ws.iter_rows(min_row=2):
            row[4].alignment = Alignment(wrap_text=True, vertical="top")
            row[6].alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(out_path)


def main(argv=None) -> int:
    import argparse
    import json
    import sys

    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")

    parser = argparse.ArgumentParser(description="세특 결정론 채점기 + 저장 게이트")
    parser.add_argument("drafts", help="초안 JSON 경로")
    parser.add_argument("--profile", required=True, help="활동프로파일 JSON 경로")
    parser.add_argument("--roster", help="명렬 JSON 경로(선택) — 학번·이름 대조와 미제출자 보고")
    parser.add_argument("--submitted",
                         help="제출 원본(토큰본 또는 제출자 학번 목록) JSON 경로(선택) — "
                              "명렬과 독립된 경로로 제출 여부를 대조한다")
    parser.add_argument("--mapping",
                         help='매핑 JSON 경로(선택, {"map": {"학번": "S-XXXX"}} — pseudonymize.py '
                              "issue의 산출물). 실명 단계 행별 출력의 식별자를 토큰으로 번역해 "
                              "stdout에 학번이 나가지 않게 한다. 없으면 코드별 집계만 출력한다")
    parser.add_argument("--expected-count", type=int, dest="expected_count",
                         help="교사가 선언하는 이번 활동의 대상 인원(명렬·초안에서 세지 말 것). "
                              "--save 시 이 값 또는 활동프로파일의 '인원'이 반드시 있어야 한다")
    parser.add_argument("--save", help="검증 통과 시 저장할 xlsx 경로")
    args = parser.parse_args(argv)

    # utf-8-sig: 교사가 초안·프로파일·명렬 JSON을 메모장에서 편집하면 BOM이 붙을 수 있다.
    with open(args.drafts, encoding="utf-8-sig") as f:
        drafts = json.load(f)
    with open(args.profile, encoding="utf-8-sig") as f:
        profile = json.load(f)
    total_students = sum(len(cls.get("students", [])) for cls in drafts.get("classes", []))
    if total_students == 0:
        print("오류: 초안에 학생이 없습니다. 초안 JSON의 classes/students 구조를 확인하세요.")
        return 1

    # 단계 판별 — 혼재면 어떤 검사도 시작하지 않는다(잘못된 단계의 검사가
    # 부분 통과를 남기면 그 stdout이 곧 오판의 근거가 된다).
    stage = detect_stage(drafts)
    if stage == "mixed":
        print("FAIL [STAGE_MIXED] 토큰 학생과 실명 학생이 섞여 있습니다"
              "(또는 두 키를 다 가진/둘 다 없는 학생이 있습니다). "
              "§6 토큰 단계면 전원 '토큰' 키, §7 실명 단계면 전원 '학번' 키여야 합니다. "
              "다른 검사는 진행하지 않았습니다.")
        return 1

    if stage == "token" and (args.roster or args.submitted or args.save):
        # 토큰 초안에는 학번이 없어 명렬 대조가 성립하지 않고(전원 ROSTER FAIL이
        # 되는 사고), 저장은 실명 재결합 전이라 애초에 산출물이 될 수 없다.
        print("실행 거부: 명렬 대조와 저장은 실명 재결합(finalize) 후의 실명 단계 전용입니다. "
              "토큰 단계에서는 --roster/--submitted/--save 없이 다시 실행하세요.")
        return 1

    if args.save and not args.roster:
        print("저장 거부: --roster 명렬.json 이 필요합니다. 명렬 없이는 학번 잔존·오귀속·미제출자 검사를 할 수 없어 저장하지 않습니다.")
        return 1

    if not str(profile.get("평가자료", "")).strip():
        print("차단: 교사의 평가 자료(채점표, 루브릭 점수, 관찰 기록 등)가 확인되지 않았습니다.")
        print("이 도구는 평가를 대신하지 않습니다. 채점을 먼저 완료하고 활동프로파일의 평가자료 항목에 출처를 기록하세요.")
        return 1

    if stage == "token":
        report = verify_token_drafts(drafts, profile)
        for row in report["rows"]:
            for level, code, msg in row["issues"]:
                print(f"{level} {row['반']} {row['토큰']} [{code}] {msg}")
        print("이름 혼입·명렬 대조·미제출 감지는 실명 단계에서 수행됩니다")
        print(f"결과: FAIL {report['fail']}건, WARN {report['warn']}건")
        if report["fail"] > 0:
            print("FAIL을 모두 해소한 뒤 다시 실행하세요.")
            return 1
        return 0

    roster = None
    if args.roster:
        with open(args.roster, encoding="utf-8-sig") as f:
            roster = json.load(f)
    submitted = None
    if args.submitted:
        submitted = load_submitted_ids(args.submitted)
    token_map = None
    if args.mapping:
        with open(args.mapping, encoding="utf-8-sig") as f:
            token_map = {str(sid): tok for sid, tok in (json.load(f).get("map") or {}).items()}

    if roster is not None:
        # 이 문구는 ROSTER 대조가 오귀속을 증명하지 못한다는 사실을 매 실행마다
        # 드러낸다(적대적 감사 대응) — 초안의 이름이 바로 이 명렬에서 왔으므로,
        # 명렬과 이름이 일치한다는 결과는 "형식이 맞다"는 것 이상을 의미하지 않는다.
        print("명렬 대조: 형식 일치만 확인했습니다(이름 출처가 명렬이므로 오귀속은 증명되지 않습니다).")

    report = verify_drafts(drafts, profile, roster=roster, submitted=submitted)

    # 실명 단계 stdout 비식별화 — 이 스크립트의 stdout은 에이전트(LLM)가 읽는
    # 유일한 창구라, 학번이 어떤 형태로도 나가면 개인정보 계약 위반이다.
    # 매핑이 있으면 행별 식별자를 학번→토큰으로 번역해 찍고, 없으면(파기 후
    # 재검증 등) 거부하는 대신 코드별 집계만 찍는다. 반 이름은 허용.
    if token_map is not None:
        unmapped_rows = 0
        for row in report["rows"]:
            token = token_map.get(str(row["학번"]))
            if token is None:
                unmapped_rows += 1
                continue
            for level, code, msg in row["issues"]:
                print(f"{level} {row['반']} {token} [{code}] {msg}")
        if unmapped_rows:
            # 행을 숨긴 채 통과시키면 감춰진 FAIL이 조용히 저장까지 가므로,
            # 번역 불가 자체를 FAIL로 계산해 저장 게이트에 걸리게 한다.
            print(f"매핑에 없는 학번 {unmapped_rows}건 — 매핑·명렬을 확인하세요")
            report["fail"] += unmapped_rows
    else:
        counts: dict[tuple[str, str], int] = {}
        for row in report["rows"]:
            for level, code, _ in row["issues"]:
                counts[(level, code)] = counts.get((level, code), 0) + 1
        if counts:
            order = sorted(counts, key=lambda k: (k[0] != "FAIL", k[1]))
            print(", ".join(f"{level} {code} {counts[(level, code)]}건" for level, code in order))
            print("학생별 상세를 보려면 --mapping 매핑.json을 지정하세요")
    for level, code, msg in report.get("경고", []):
        print(f"{level} [{code}] {msg}")
    unmapped_missing = 0
    for s in report.get("미제출", []):
        token = (token_map or {}).get(str(s["학번"]))
        if token:
            print(f"미제출 {token} — 원본 없음, 별도 확인 필요")
        else:
            # 미제출자는 토큰이 발급되지 않는 것이 정상이라(issue는 제출자 전용)
            # 매핑 유무와 무관하게 건수로만 알린다 — 명단은 로컬에서 대조한다.
            unmapped_missing += 1
    if unmapped_missing:
        print(f"미제출 {unmapped_missing}명 — 원본 없음, 별도 확인 필요(명단은 명렬과 초안을 로컬에서 대조하세요)")
    print(f"결과: FAIL {report['fail']}건, WARN {report['warn']}건, 미제출 {len(report.get('미제출', []))}명")

    if report["fail"] > 0:
        print("저장 차단: FAIL을 모두 해소한 뒤 다시 실행하세요.")
        return 1
    if args.save:
        import os

        # ① 교사 선언 인원수 대조 — 명렬·초안 어느 쪽에서도 파생되지 않은 유일한
        # 외부 정보다. 명렬이 통째로 잘못돼도(행 밀림 등) 이 값은 영향받지 않으므로
        # 저장될 행 수와 어긋나면 오귀속·누락을 구조적으로 의심할 수 있다.
        declared_count = None
        count_source = None
        if args.expected_count is not None:
            declared_count = args.expected_count
            count_source = "cli"
        else:
            raw = profile.get("인원")
            if raw not in (None, ""):
                try:
                    declared_count = int(raw)
                    count_source = "profile"
                except (TypeError, ValueError):
                    declared_count = None

        if declared_count is None:
            print("저장하려면 이 활동의 대상 인원을 알려주세요: --expected-count 48")
            print("명렬이나 초안에서 세지 말고, 교사가 아는 실제 인원을 넣어 주세요.")
            print("(예: 학급 정원에서 전출·자퇴를 뺀 수)")
            return 1

        if count_source == "profile":
            print(f"프로파일에 기록된 인원 {declared_count}명을 재사용합니다. "
                  "다른 값이면 --expected-count로 다시 지정하세요.")
        else:
            print(f"선언된 대상 인원: {declared_count}명")

        if declared_count != total_students:
            diff = declared_count - total_students
            if diff > 0:
                detail = f"{diff}명이 비었습니다."
            else:
                detail = f"{-diff}명이 초과했습니다."
            print(f"인원 불일치: 선언 {declared_count}명, 저장 대상 {total_students}명. "
                  f"{detail} 명렬 인식 단계를 다시 확인해 주세요.")
            return 1

        if count_source == "cli":
            # 다음 회차에 --expected-count 없이도 재사용할 수 있도록 프로파일에 기록한다.
            try:
                updated_profile = dict(profile)
                updated_profile["인원"] = declared_count
                with open(args.profile, "w", encoding="utf-8") as f:
                    json.dump(updated_profile, f, ensure_ascii=False, indent=2)
            except OSError:
                pass

        filled, blank = autofill_subjects(drafts)
        print(f"핵심소재 자동 기입: {filled}명(빈칸 {blank}명은 그대로).")

        tmp_path = args.save + ".tmp"
        try:
            save_xlsx(drafts, report, tmp_path)
            os.replace(tmp_path, args.save)
        except Exception:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
            raise
        print(f"저장 완료: {args.save}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
